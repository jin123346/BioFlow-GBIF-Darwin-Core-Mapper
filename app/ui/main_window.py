from pathlib import Path
import json

import pandas as pd

from PySide6.QtCore import QSettings, Qt
from PySide6.QtCore import QUrl
from PySide6.QtGui import QColor, QDesktopServices, QIcon, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSplitter,
    QSpinBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.config.app_info import (
    APP_NAME,
    APP_SUBTITLE,
    APP_VERSION,
    AUTHOR_EMAIL,
    AUTHOR_GITHUB,
    AUTHOR_NAME,
    COPYRIGHT_TEXT,
)
from app.config.country_code import ISO_COUNTRY_CODES
from app.config.auto_mapping import (
    CUSTOM_AUTO_MAPPING_PATH,
    get_default_auto_mapping_rules,
    load_auto_mapping_rules,
    reset_auto_mapping_rules,
    save_auto_mapping_rules,
)
from app.config.dwc_fields import (
    ALL_DWC_FIELDS,
    get_default_dwc_fields,
    get_dwc_field,
    get_optional_dwc_fields,
)

from app.services.excel_service import ExcelService
from app.services.gbif_service import GbifService
from app.services.mapping_service import MappingService, auto_match_column
from app.ui.paste_table_widget import PasteTableWidget
from app.utils.paths import LOGO_PATH, OUTPUT_DIR


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION} - {APP_SUBTITLE}")
        self.resize(1280, 820)
        if LOGO_PATH.exists():
            self.setWindowIcon(QIcon(str(LOGO_PATH)))

        self.raw_df = None
        self.source_df = None
        self.result_df = None
        self.gbif_df = None
        self.gbif_search_summary = None
        self.current_file_path = None
        self.current_sheet_name = None
        self.selected_header_row = None
        self.settings = QSettings(AUTHOR_NAME, APP_NAME)
        self.active_field_keys = [field["key"] for field in get_default_dwc_fields()]
        self.combo_boxes = {}
        self.manual_inputs = {}

        self._init_ui()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(14, 14, 14, 14)
        root_layout.setSpacing(10)

        control_panel = QFrame()
        control_panel.setObjectName("controlPanel")
        control_layout = QVBoxLayout(control_panel)
        control_layout.setContentsMargins(14, 10, 14, 10)
        control_layout.setSpacing(8)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        self.btn_upload = QPushButton("파일 업로드")
        self.btn_upload.setProperty("role", "primary")
        self.btn_upload.clicked.connect(self.load_excel)

        self.btn_apply_header = QPushButton("선택 행을 헤더로 적용")
        self.btn_apply_header.setProperty("role", "secondary")
        self.btn_apply_header.clicked.connect(self.apply_selected_header)
        self.btn_apply_header.setEnabled(False)

        self.btn_add_field = QPushButton("컬럼 추가")
        self.btn_add_field.setProperty("role", "secondary")
        self.btn_add_field.clicked.connect(self.add_optional_field)
        self.btn_add_field.setEnabled(False)

        self.btn_source_coordinate_preview = QPushButton("원본 좌표 확인")
        self.btn_source_coordinate_preview.setProperty("role", "secondary")
        self.btn_source_coordinate_preview.clicked.connect(self.preview_source_coordinates_on_map)
        self.btn_source_coordinate_preview.setEnabled(False)

        self.btn_preview_result = QPushButton("결과 확인")
        self.btn_preview_result.setProperty("role", "primary")
        self.btn_preview_result.clicked.connect(self.preview_mapped_result)
        self.btn_preview_result.setEnabled(False)

        self.btn_coordinate_preview = QPushButton("좌표 확인")
        self.btn_coordinate_preview.setProperty("role", "secondary")
        self.btn_coordinate_preview.clicked.connect(self.preview_coordinates_on_map)
        self.btn_coordinate_preview.setEnabled(False)

        self.btn_export = QPushButton("매핑 결과 엑셀 저장")
        self.btn_export.setProperty("role", "success")
        self.btn_export.clicked.connect(self.export_mapped_excel)
        self.btn_export.setEnabled(False)

        self.btn_about = QPushButton("정보")
        self.btn_about.setProperty("role", "secondary")
        self.btn_about.clicked.connect(self.show_about)

        self.lbl_file = QLabel("업로드된 파일 없음")
        self.lbl_file.setObjectName("metaValue")

        self.lbl_sheet = QLabel("선택된 시트 없음")
        self.lbl_sheet.setObjectName("metaValue")

        self.lbl_header = QLabel("선택된 헤더 행: 없음")
        self.lbl_header.setObjectName("metaValueAccent")

        button_layout.addWidget(self.btn_upload)
        button_layout.addWidget(self.btn_apply_header)
        button_layout.addWidget(self.btn_add_field)
        button_layout.addWidget(self.btn_source_coordinate_preview)
        button_layout.addWidget(self.btn_preview_result)
        button_layout.addWidget(self.btn_coordinate_preview)
        button_layout.addWidget(self.btn_export)
        button_layout.addWidget(self.btn_about)
        button_layout.addStretch()

        status_layout = QHBoxLayout()
        status_layout.setSpacing(10)
        status_layout.addWidget(self._build_meta_block("파일", self.lbl_file))
        status_layout.addWidget(self._build_meta_block("시트", self.lbl_sheet))
        status_layout.addWidget(self._build_meta_block("헤더 상태", self.lbl_header))
        status_layout.addStretch()

        title_layout = QHBoxLayout()
        title_layout.setSpacing(10)

        logo_label = QLabel()
        logo_label.setObjectName("appLogo")
        logo_label.setFixedSize(74, 54)
        logo_label.setAlignment(Qt.AlignCenter)
        if LOGO_PATH.exists():
            logo_pixmap = QPixmap(str(LOGO_PATH))
            logo_label.setPixmap(
                logo_pixmap.scaled(
                    logo_label.size(),
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation,
                )
            )
        logo_label.setVisible(LOGO_PATH.exists())

        title_text_layout = QVBoxLayout()
        title_text_layout.setSpacing(2)

        app_title = QLabel(APP_NAME)
        app_title.setObjectName("appTitle")

        app_subtitle = QLabel(APP_SUBTITLE)
        app_subtitle.setObjectName("appSubtitle")

        title_text_layout.addWidget(app_title)
        title_text_layout.addWidget(app_subtitle)
        title_layout.addWidget(logo_label)
        title_layout.addLayout(title_text_layout)
        title_layout.addSpacing(18)
        title_layout.addLayout(button_layout)
        title_layout.addStretch()

        control_layout.addLayout(title_layout)
        control_layout.addLayout(status_layout)
        root_layout.addWidget(control_panel)

        content_splitter = QSplitter(Qt.Horizontal)
        content_splitter.setChildrenCollapsible(False)

        left_panel = QWidget()
        left_panel.setMinimumWidth(300)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(10)
        left_layout.addWidget(
            self._build_section_header(
                "원본 엑셀 미리보기",
                "행을 클릭해서 실제 헤더로 사용할 줄을 고르세요.",
            )
        )

        self.preview_position_label = QLabel("현재 선택 위치: 없음")
        self.preview_position_label.setObjectName("metaValueAccent")
        left_layout.addWidget(self.preview_position_label)

        self.preview_table = QTableWidget()
        self.preview_table.cellClicked.connect(self.select_header_row)
        self._configure_table(self.preview_table)
        self.preview_table.verticalHeader().setVisible(True)
        self.preview_table.verticalHeader().setDefaultSectionSize(28)
        left_layout.addWidget(self.preview_table)

        right_panel = QWidget()
        right_panel.setMinimumWidth(620)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        workspace_splitter = QSplitter(Qt.Vertical)
        workspace_splitter.setChildrenCollapsible(False)

        mapping_panel = QWidget()
        mapping_panel_layout = QVBoxLayout(mapping_panel)
        mapping_panel_layout.setContentsMargins(0, 0, 0, 0)
        mapping_panel_layout.setSpacing(10)
        mapping_panel_layout.addWidget(
            self._build_section_header(
                "GBIF 필드 매핑",
                "기본 필드를 먼저 보여주고, 필요하면 컬럼 추가로 더 넣을 수 있습니다.",
            )
        )

        self.mapping_summary = QLabel("현재 0개 필드")
        self.mapping_summary.setObjectName("metaValue")
        mapping_panel_layout.addWidget(self.mapping_summary)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QFrame.NoFrame)

        self.mapping_container = QWidget()
        self.mapping_layout = QVBoxLayout(self.mapping_container)
        self.mapping_layout.setContentsMargins(0, 0, 0, 0)
        self.mapping_layout.setSpacing(10)
        self.mapping_layout.setAlignment(Qt.AlignTop)
        self.scroll_area.setWidget(self.mapping_container)
        mapping_panel_layout.addWidget(self.scroll_area)

        result_panel = QWidget()
        result_layout = QVBoxLayout(result_panel)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(10)
        result_layout.addWidget(
            self._build_section_header(
                "변환 결과 미리보기",
                "추가한 GBIF 필드도 결과 테이블과 저장 엑셀에 함께 포함됩니다.",
            )
        )

        self.result_table = PasteTableWidget()
        self.result_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed
        )
        self._configure_table(self.result_table)
        self.result_table.verticalHeader().setVisible(True)
        self.result_table.verticalHeader().setDefaultSectionSize(30)
        result_layout.addWidget(self.result_table)

        self.missing_summary_card = QFrame()
        self.missing_summary_card.setProperty("class", "sectionCard")
        self.missing_summary_card.setMinimumHeight(92)
        self.missing_summary_card.setMaximumHeight(110)
        missing_layout = QVBoxLayout(self.missing_summary_card)
        missing_layout.setContentsMargins(16, 14, 16, 14)
        missing_layout.setSpacing(6)

        self.missing_summary_title = QLabel("빈칸 점검")
        self.missing_summary_title.setProperty("class", "sectionTitle")

        self.missing_summary_label = QLabel("결과 확인 후 빈칸 요약이 여기에 표시됩니다.")
        self.missing_summary_label.setWordWrap(True)
        self.missing_summary_label.setProperty("class", "sectionDescription")
        self.missing_summary_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        self.missing_summary_scroll = QScrollArea()
        self.missing_summary_scroll.setWidgetResizable(True)
        self.missing_summary_scroll.setFrameShape(QFrame.NoFrame)
        self.missing_summary_scroll.setMinimumHeight(44)
        self.missing_summary_scroll.setMaximumHeight(54)

        self.missing_summary_content = QWidget()
        self.missing_summary_content_layout = QVBoxLayout(self.missing_summary_content)
        self.missing_summary_content_layout.setContentsMargins(0, 0, 0, 0)
        self.missing_summary_content_layout.setSpacing(4)
        self.missing_summary_content_layout.addWidget(self.missing_summary_label)
        self.missing_summary_scroll.setWidget(self.missing_summary_content)

        missing_layout.addWidget(self.missing_summary_title)
        missing_layout.addWidget(self.missing_summary_scroll)
        result_layout.addWidget(self.missing_summary_card)

        workspace_splitter.addWidget(mapping_panel)
        workspace_splitter.addWidget(result_panel)
        workspace_splitter.setStretchFactor(0, 5)
        workspace_splitter.setStretchFactor(1, 5)
        workspace_splitter.setSizes([360, 360])

        self.main_tabs = QTabWidget()

        mapping_tab = QWidget()
        mapping_tab_layout = QVBoxLayout(mapping_tab)
        mapping_tab_layout.setContentsMargins(0, 0, 0, 0)
        mapping_tab_layout.addWidget(workspace_splitter)

        self.main_tabs.addTab(mapping_tab, "매핑 작업")
        self.main_tabs.addTab(self._build_gbif_analysis_tab(), "GBIF 분석")
        self.main_tabs.addTab(self._build_auto_mapping_settings_tab(), "자동 매핑 설정")

        right_layout.addWidget(self.main_tabs)

        content_splitter.addWidget(left_panel)
        content_splitter.addWidget(right_panel)
        content_splitter.setStretchFactor(0, 4)
        content_splitter.setStretchFactor(1, 9)
        content_splitter.setSizes([330, 900])

        root_layout.addWidget(content_splitter)
        root_layout.setStretchFactor(content_splitter, 1)

        footer_label = QLabel(f"{APP_NAME} v{APP_VERSION} · {COPYRIGHT_TEXT}")
        footer_label.setObjectName("footerCredit")
        footer_label.setAlignment(Qt.AlignCenter)
        root_layout.addWidget(footer_label)

        self.setStyleSheet(
            """
            QMainWindow {
                background: #eef3f7;
            }
            QWidget {
                color: #1e293b;
                font-family: "Segoe UI", "Malgun Gothic", sans-serif;
                font-size: 12px;
            }
            QFrame#controlPanel {
                background: #fbfdff;
                border: 1px solid #d6e0ea;
                border-radius: 12px;
            }
            QLabel#appTitle {
                color: #102a43;
                font-size: 19px;
                font-weight: 800;
            }
            QLabel#appSubtitle {
                color: #52667a;
                font-size: 11px;
                font-weight: 600;
            }
            QLabel#footerCredit {
                color: #8a9aab;
                font-size: 10px;
                font-weight: 600;
                padding: 0px;
            }
            QWidget[class="metaBlock"] {
                background: #f5f8fb;
                border: 1px solid #e0e7ef;
                border-radius: 8px;
            }
            QFrame[class="sectionCard"] {
                background: #fbfdff;
                border: 1px solid #d8e2ec;
                border-radius: 10px;
            }
            QLabel[class="sectionTitle"] {
                font-size: 15px;
                font-weight: 700;
                color: #132f4c;
            }
            QLabel[class="sectionDescription"] {
                color: #64758a;
                font-size: 12px;
            }
            QLabel[class="metaLabel"] {
                color: #718096;
                font-size: 10px;
                font-weight: 700;
            }
            QLabel#metaValue {
                color: #26384c;
                font-size: 12px;
                font-weight: 700;
            }
            QLabel#metaValueAccent {
                color: #0f766e;
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton {
                background: #ffffff;
                color: #244158;
                border: 1px solid #cdd8e4;
                border-radius: 8px;
                padding: 7px 13px;
                font-weight: 700;
                min-height: 18px;
            }
            QPushButton[role="primary"] {
                background: #123a55;
                color: #ffffff;
                border: none;
            }
            QPushButton[role="success"] {
                background: #0f766e;
                color: #ffffff;
                border: none;
            }
            QPushButton[role="secondary"] {
                background: #ffffff;
                color: #244158;
                border: 1px solid #c9d6e2;
            }
            QPushButton[role="link"] {
                background: transparent;
                color: #0f766e;
                border: none;
                border-radius: 4px;
                padding: 2px 0px;
                font-weight: 700;
                text-align: left;
                min-height: 16px;
            }
            QPushButton:disabled {
                background: #e3e9f0;
                color: #93a3b5;
                border: 1px solid #d5dee8;
            }
            QPushButton:hover:!disabled {
                background: #eaf2f8;
                border: 1px solid #b6cadb;
            }
            QPushButton[role="primary"]:hover:!disabled {
                background: #174a6d;
                border: none;
            }
            QPushButton[role="success"]:hover:!disabled {
                background: #0d9488;
                border: none;
            }
            QScrollArea, QTableWidget {
                background: #fbfdff;
                border: 1px solid #d7e1ec;
                border-radius: 10px;
            }
            QTableWidget {
                alternate-background-color: #f6f9fc;
                gridline-color: #e2e8f0;
                selection-background-color: #dff3ef;
                selection-color: #102a43;
            }
            QTableWidget::item {
                padding: 6px;
                border-bottom: 1px solid #edf2f7;
            }
            QTableWidget::item:selected {
                background: #dff3ef;
                color: #102a43;
            }
            QHeaderView::section {
                background: #edf4f8;
                color: #28445d;
                border: none;
                border-right: 1px solid #d7e1ec;
                border-bottom: 1px solid #d7e1ec;
                padding: 9px;
                font-weight: 700;
            }
            QFrame[frameShape="6"] {
                background: #fbfdff;
                border: 1px solid #dde6ef;
                border-radius: 10px;
            }
            QFrame[class="mappingCard"] {
                background: #ffffff;
                border: 1px solid #dbe5ee;
                border-radius: 8px;
            }
            QLabel[class="fieldName"] {
                color: #172b4d;
                font-size: 13px;
                font-weight: 800;
            }
            QLabel[class="fieldDescription"] {
                color: #64748b;
                font-size: 11px;
            }
            QComboBox, QLineEdit {
                border: 1px solid #cbd8e5;
                border-radius: 8px;
                padding: 8px 10px;
                background: #ffffff;
                color: #20364d;
                min-height: 20px;
            }
            QComboBox:hover, QLineEdit:hover {
                border: 1px solid #95adbf;
            }
            QComboBox:focus, QLineEdit:focus {
                border: 1px solid #0f766e;
                background: #fbfffe;
            }
            QComboBox[autoMatched="true"] {
                background: #ecfdf5;
                border: 1px solid #8bd4bd;
                color: #0f5132;
            }
            QSplitter::handle {
                background: #dde7ef;
                border-radius: 2px;
            }
            QSplitter::handle:hover {
                background: #b9cad8;
            }
            QTabWidget::pane {
                border: none;
                background: transparent;
            }
            QTabBar::tab {
                background: #e5edf4;
                color: #52667a;
                border: 1px solid #ccd8e3;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 8px 16px;
                font-weight: 700;
                min-width: 100px;
            }
            QTabBar::tab:selected {
                background: #fbfdff;
                color: #123a55;
            }
            QScrollBar:vertical {
                background: #f2f6fa;
                width: 12px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #b9c7d5;
                border-radius: 6px;
                min-height: 26px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QScrollBar:horizontal {
                background: #f2f6fa;
                height: 12px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #b9c7d5;
                border-radius: 6px;
                min-width: 26px;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
            """
        )

        self._update_mapping_summary()

    def _build_meta_block(self, title: str, value_label: QLabel) -> QWidget:
        wrapper = QWidget()
        wrapper.setProperty("class", "metaBlock")
        layout = QHBoxLayout(wrapper)
        layout.setContentsMargins(9, 4, 9, 4)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setProperty("class", "metaLabel")
        value_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.addStretch()
        return wrapper

    def _build_section_header(self, title: str, description: str) -> QFrame:
        frame = QFrame()
        frame.setProperty("class", "sectionCard")

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(4)

        title_label = QLabel(title)
        title_label.setProperty("class", "sectionTitle")

        desc_label = QLabel(description)
        desc_label.setProperty("class", "sectionDescription")
        desc_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(desc_label)
        return frame

    def _build_gbif_analysis_tab(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        layout.addWidget(
            self._build_section_header(
                "GBIF 종 데이터 분석",
                "GBIF에서 찾는 종의 좌표 자료를 가져와 지도와 연도별 기록 변화를 확인합니다.",
            )
        )

        search_card = QFrame()
        search_card.setProperty("class", "sectionCard")
        search_layout = QVBoxLayout(search_card)
        search_layout.setContentsMargins(16, 14, 16, 14)
        search_layout.setSpacing(10)

        input_layout = QHBoxLayout()
        input_layout.setSpacing(8)

        self.gbif_species_input = QLineEdit()
        self.gbif_species_input.setPlaceholderText("학명 또는 종명 예: Quercus mongolica")

        self.gbif_country_input = QLineEdit()
        self.gbif_country_input.setPlaceholderText("국가코드 선택 입력")
        self.gbif_country_input.setMaxLength(2)
        self.gbif_country_input.setFixedWidth(120)

        self.gbif_limit_input = QSpinBox()
        self.gbif_limit_input.setRange(10, 100000)
        self.gbif_limit_input.setSingleStep(1000)
        self.gbif_limit_input.setValue(1000)
        self.gbif_limit_input.setSuffix("건")
        self.gbif_limit_input.setFixedWidth(116)

        self.gbif_chart_type_combo = QComboBox()
        self.gbif_chart_type_combo.addItem("막대 그래프", "bar")
        self.gbif_chart_type_combo.addItem("선 그래프", "line")
        self.gbif_chart_type_combo.setFixedWidth(120)

        self.gbif_period_combo = QComboBox()
        self.gbif_period_combo.addItem("종합", "summary")
        self.gbif_period_combo.addItem("연도별", "year")
        self.gbif_period_combo.addItem("월별", "month")
        self.gbif_period_combo.addItem("계절별", "season")
        self.gbif_period_combo.setFixedWidth(96)
        self.gbif_period_combo.currentIndexChanged.connect(self.update_gbif_period_controls)

        self.gbif_year_from_input = QSpinBox()
        self.gbif_year_from_input.setRange(0, 9999)
        self.gbif_year_from_input.setSpecialValueText("전체")
        self.gbif_year_from_input.setFixedWidth(86)

        self.gbif_year_to_input = QSpinBox()
        self.gbif_year_to_input.setRange(0, 9999)
        self.gbif_year_to_input.setSpecialValueText("전체")
        self.gbif_year_to_input.setFixedWidth(86)

        self.gbif_month_from_input = QSpinBox()
        self.gbif_month_from_input.setRange(1, 12)
        self.gbif_month_from_input.setValue(1)
        self.gbif_month_from_input.setSuffix("월")
        self.gbif_month_from_input.setFixedWidth(70)

        self.gbif_month_to_input = QSpinBox()
        self.gbif_month_to_input.setRange(1, 12)
        self.gbif_month_to_input.setValue(12)
        self.gbif_month_to_input.setSuffix("월")
        self.gbif_month_to_input.setFixedWidth(70)

        self.btn_fetch_gbif = QPushButton("GBIF 가져오기")
        self.btn_fetch_gbif.setProperty("role", "primary")
        self.btn_fetch_gbif.clicked.connect(self.fetch_gbif_analysis)

        self.btn_open_gbif_report = QPushButton("지도/그래프 열기")
        self.btn_open_gbif_report.setProperty("role", "secondary")
        self.btn_open_gbif_report.clicked.connect(self.open_gbif_analysis_report)
        self.btn_open_gbif_report.setEnabled(False)

        self.btn_export_gbif_csv = QPushButton("CSV 저장")
        self.btn_export_gbif_csv.setProperty("role", "secondary")
        self.btn_export_gbif_csv.clicked.connect(self.export_gbif_analysis_csv)
        self.btn_export_gbif_csv.setEnabled(False)

        self.btn_export_gbif_excel = QPushButton("엑셀 저장")
        self.btn_export_gbif_excel.setProperty("role", "secondary")
        self.btn_export_gbif_excel.clicked.connect(self.export_gbif_analysis_excel)
        self.btn_export_gbif_excel.setEnabled(False)

        self.btn_export_gbif_geojson = QPushButton("QGIS 저장")
        self.btn_export_gbif_geojson.setProperty("role", "secondary")
        self.btn_export_gbif_geojson.clicked.connect(self.export_gbif_analysis_geojson)
        self.btn_export_gbif_geojson.setEnabled(False)

        input_layout.addWidget(self.gbif_species_input, 1)
        input_layout.addWidget(self.gbif_country_input)
        input_layout.addWidget(self.gbif_limit_input)
        input_layout.addWidget(self.gbif_period_combo)
        input_layout.addWidget(self.gbif_chart_type_combo)
        input_layout.addWidget(self.btn_fetch_gbif)
        input_layout.addWidget(self.btn_open_gbif_report)
        input_layout.addWidget(self.btn_export_gbif_csv)
        input_layout.addWidget(self.btn_export_gbif_excel)
        input_layout.addWidget(self.btn_export_gbif_geojson)
        search_layout.addLayout(input_layout)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(8)
        filter_layout.addWidget(QLabel("연도 범위"))
        filter_layout.addWidget(self.gbif_year_from_input)
        filter_layout.addWidget(QLabel("~"))
        filter_layout.addWidget(self.gbif_year_to_input)
        filter_layout.addSpacing(14)
        filter_layout.addWidget(QLabel("월 범위"))
        filter_layout.addWidget(self.gbif_month_from_input)
        filter_layout.addWidget(QLabel("~"))
        filter_layout.addWidget(self.gbif_month_to_input)
        filter_layout.addStretch()
        search_layout.addLayout(filter_layout)

        download_layout = QHBoxLayout()
        download_layout.setSpacing(8)

        self.gbif_username_input = QLineEdit()
        self.gbif_username_input.setPlaceholderText("GBIF username")
        self.gbif_username_input.setText(self.settings.value("gbif/username", "", str))

        self.gbif_email_input = QLineEdit()
        self.gbif_email_input.setPlaceholderText("GBIF email")
        self.gbif_email_input.setText(self.settings.value("gbif/email", "", str))

        self.gbif_password_input = QLineEdit()
        self.gbif_password_input.setPlaceholderText("GBIF password")
        self.gbif_password_input.setEchoMode(QLineEdit.Password)

        self.btn_request_gbif_download = QPushButton("DOI 다운로드 요청")
        self.btn_request_gbif_download.setProperty("role", "success")
        self.btn_request_gbif_download.clicked.connect(self.request_gbif_download)

        download_layout.addWidget(self.gbif_username_input)
        download_layout.addWidget(self.gbif_email_input)
        download_layout.addWidget(self.gbif_password_input)
        download_layout.addWidget(self.btn_request_gbif_download)
        search_layout.addLayout(download_layout)

        download_result_layout = QHBoxLayout()
        download_result_layout.setSpacing(8)

        self.gbif_download_url_input = QLineEdit()
        self.gbif_download_url_input.setPlaceholderText("마지막 GBIF 다운로드 URL")
        self.gbif_download_url_input.setReadOnly(True)

        self.btn_open_gbif_download_url = QPushButton("URL 열기")
        self.btn_open_gbif_download_url.setProperty("role", "secondary")
        self.btn_open_gbif_download_url.clicked.connect(self.open_gbif_download_url)
        self.btn_open_gbif_download_url.setEnabled(False)

        self.btn_copy_gbif_download_url = QPushButton("URL 복사")
        self.btn_copy_gbif_download_url.setProperty("role", "secondary")
        self.btn_copy_gbif_download_url.clicked.connect(self.copy_gbif_download_url)
        self.btn_copy_gbif_download_url.setEnabled(False)

        download_result_layout.addWidget(QLabel("다운로드 URL"))
        download_result_layout.addWidget(self.gbif_download_url_input, 1)
        download_result_layout.addWidget(self.btn_open_gbif_download_url)
        download_result_layout.addWidget(self.btn_copy_gbif_download_url)
        search_layout.addLayout(download_result_layout)

        self.gbif_summary_label = QLabel("검색 결과가 아직 없습니다.")
        self.gbif_summary_label.setObjectName("metaValue")
        self.gbif_summary_label.setWordWrap(True)
        search_layout.addWidget(self.gbif_summary_label)
        layout.addWidget(search_card)

        self.gbif_result_table = QTableWidget()
        self._configure_table(self.gbif_result_table)
        self.gbif_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.gbif_result_table.verticalHeader().setVisible(True)
        self.gbif_result_table.verticalHeader().setDefaultSectionSize(28)
        layout.addWidget(self.gbif_result_table)

        return panel

    def fetch_gbif_analysis(self):
        species_name = self.gbif_species_input.text().strip()
        country_code = self.gbif_country_input.text().strip().upper()
        if not species_name:
            QMessageBox.warning(self, "GBIF 분석", "검색할 학명 또는 종명을 입력하세요.")
            return
        if country_code and not self.validate_country_code(country_code):
            QMessageBox.warning(self, "GBIF 분석", "국가코드는 KR, JP, US처럼 2자리 영문으로 입력하세요.")
            return

        self.btn_fetch_gbif.setEnabled(False)
        self.btn_fetch_gbif.setText("가져오는 중...")
        try:
            result = GbifService.fetch_occurrences(
                species_name,
                country_code=country_code,
                limit=self.gbif_limit_input.value(),
            )
            self.gbif_df = result.dataframe
            self.gbif_search_summary = {
                "matchedName": result.matched_name,
                "taxonKey": result.taxon_key,
                "rank": result.rank,
                "status": result.status,
                "totalRecords": result.total_records,
                "shownRecords": len(result.dataframe),
                "countryCode": country_code or "ALL",
            }
            self.show_gbif_results(self.gbif_df)
            self.update_gbif_year_range(self.gbif_df)
            self.btn_open_gbif_report.setEnabled(not self.gbif_df.empty)
            self.btn_export_gbif_csv.setEnabled(not self.gbif_df.empty)
            self.btn_export_gbif_excel.setEnabled(not self.gbif_df.empty)
            self.btn_export_gbif_geojson.setEnabled(not self.gbif_df.empty)
            limit_note = ""
            if result.total_records > len(result.dataframe):
                limit_note = "\n빠른 분석은 현재 설정한 건수까지만 표시합니다. 논문용 전체 자료는 DOI 다운로드 요청을 사용하세요."
            self.gbif_summary_label.setText(
                f"일치 이름: {result.matched_name} / taxonKey: {result.taxon_key} / "
                f"전체 좌표 기록: {result.total_records:,}건 / 표시: {len(result.dataframe):,}건"
                f"{limit_note}"
            )
            if self.gbif_df.empty:
                QMessageBox.information(self, "GBIF 분석", "일치하는 좌표 기록이 없습니다.")
            else:
                self.open_gbif_analysis_report()
        except Exception as e:
            self.gbif_df = None
            self.gbif_search_summary = None
            self.update_gbif_year_range(pd.DataFrame())
            self.btn_open_gbif_report.setEnabled(False)
            self.btn_export_gbif_csv.setEnabled(False)
            self.btn_export_gbif_excel.setEnabled(False)
            self.btn_export_gbif_geojson.setEnabled(False)
            QMessageBox.critical(self, "GBIF 분석 오류", f"GBIF 데이터를 가져오는 중 오류가 발생했습니다.\n\n{e}")
        finally:
            self.btn_fetch_gbif.setEnabled(True)
            self.btn_fetch_gbif.setText("GBIF 가져오기")

    def request_gbif_download(self):
        species_name = self.gbif_species_input.text().strip()
        country_code = self.gbif_country_input.text().strip().upper()
        username = self.gbif_username_input.text().strip()
        email = self.gbif_email_input.text().strip()
        password = self.gbif_password_input.text()

        if not species_name:
            QMessageBox.warning(self, "GBIF 다운로드 요청", "검색할 학명 또는 종명을 입력하세요.")
            return
        if country_code and not self.validate_country_code(country_code):
            QMessageBox.warning(self, "GBIF 다운로드 요청", "국가코드는 KR, JP, US처럼 2자리 영문으로 입력하세요.")
            return

        self.btn_request_gbif_download.setEnabled(False)
        self.btn_request_gbif_download.setText("요청 중...")
        try:
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            result = GbifService.request_occurrence_download(
                scientific_name=species_name,
                country_code=country_code,
                username=username,
                password=password,
                email=email,
                year_from=year_from,
                year_to=year_to,
                month_from=month_from,
                month_to=month_to,
            )
            self.settings.setValue("gbif/username", username)
            self.settings.setValue("gbif/email", email)
            self.set_gbif_download_url(result.download_url)
            QDesktopServices.openUrl(QUrl(result.download_url))
            QMessageBox.information(
                self,
                "GBIF 다운로드 요청 완료",
                "GBIF occurrence download 요청을 보냈습니다.\n\n"
                f"Download key: {result.key}\n"
                f"상태 API: {result.status_url}\n"
                f"다운로드/DOI 페이지: {result.download_url}\n\n"
                "GBIF가 파일 준비를 마치면 계정 이메일로 알림을 보냅니다. "
                "논문에는 다운로드 페이지의 공식 DOI citation 문구를 사용하세요.",
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "GBIF 다운로드 요청 오류",
                f"GBIF 다운로드 요청 중 오류가 발생했습니다.\n\n{e}",
            )
        finally:
            self.btn_request_gbif_download.setEnabled(True)
            self.btn_request_gbif_download.setText("DOI 다운로드 요청")

    def set_gbif_download_url(self, url: str):
        self.gbif_download_url_input.setText(url)
        has_url = bool(url.strip())
        self.btn_open_gbif_download_url.setEnabled(has_url)
        self.btn_copy_gbif_download_url.setEnabled(has_url)

    def open_gbif_download_url(self):
        url = self.gbif_download_url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "GBIF 다운로드 URL", "열 수 있는 다운로드 URL이 없습니다.")
            return
        QDesktopServices.openUrl(QUrl(url))

    def copy_gbif_download_url(self):
        url = self.gbif_download_url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "GBIF 다운로드 URL", "복사할 다운로드 URL이 없습니다.")
            return
        QApplication.clipboard().setText(url)
        QMessageBox.information(self, "URL 복사", "GBIF 다운로드 URL을 클립보드에 복사했습니다.")

    def show_gbif_results(self, df: pd.DataFrame):
        self.gbif_result_table.clear()
        self.gbif_result_table.setRowCount(len(df))
        self.gbif_result_table.setColumnCount(len(df.columns))
        self.gbif_result_table.setHorizontalHeaderLabels([str(col) for col in df.columns.tolist()])

        for row_idx in range(len(df)):
            for col_idx, _ in enumerate(df.columns):
                value = df.iloc[row_idx, col_idx]
                self.gbif_result_table.setItem(
                    row_idx,
                    col_idx,
                    QTableWidgetItem("" if pd.isna(value) else str(value)),
                )

        self.gbif_result_table.resizeColumnsToContents()
        self.gbif_result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def update_gbif_year_range(self, df: pd.DataFrame):
        years = self._gbif_year_values(df)
        if years.empty:
            for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
                widget.setRange(0, 9999)
                widget.setValue(0)
            return

        min_year = int(years.min())
        max_year = int(years.max())
        for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
            widget.setRange(0, 9999)
        self.gbif_year_from_input.setValue(min_year)
        self.gbif_year_to_input.setValue(max_year)

        self.update_gbif_period_controls()

    def update_gbif_period_controls(self, *_):
        for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
            widget.setEnabled(True)

    def get_gbif_filter_ranges(self) -> tuple[int | None, int | None, int, int]:
        year_from = self.gbif_year_from_input.value() if self.gbif_year_from_input.isEnabled() else 0
        year_to = self.gbif_year_to_input.value() if self.gbif_year_to_input.isEnabled() else 0
        month_from = self.gbif_month_from_input.value()
        month_to = self.gbif_month_to_input.value()

        if year_from and year_to and year_from > year_to:
            raise ValueError("연도 범위는 시작 연도가 끝 연도보다 클 수 없습니다.")
        if month_from > month_to:
            raise ValueError("월 범위는 시작 월이 끝 월보다 클 수 없습니다.")

        return year_from or None, year_to or None, month_from, month_to

    def open_gbif_analysis_report(self):
        if self.gbif_df is None or self.gbif_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "먼저 GBIF 데이터를 가져오세요.")
            return

        report_path = OUTPUT_DIR / "gbif_analysis_report.html"
        try:
            chart_type = self.gbif_chart_type_combo.currentData() or "bar"
            period_type = self.gbif_period_combo.currentData() or "year"
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            report_path.write_text(
                self._gbif_analysis_report_html(
                    self.gbif_df,
                    self.gbif_search_summary or {},
                    chart_type,
                    period_type,
                    year_from,
                    year_to,
                    month_from,
                    month_to,
                ),
                encoding="utf-8",
            )
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(report_path)))
        except Exception as e:
            QMessageBox.critical(self, "GBIF 분석 오류", f"리포트를 생성하는 중 오류가 발생했습니다.\n\n{e}")

    def export_gbif_analysis_csv(self):
        if self.gbif_df is None or self.gbif_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "GBIF 분석 데이터 저장",
            str(Path(self._get_last_save_dir()) / "gbif_occurrences.csv"),
            "CSV Files (*.csv)",
        )
        if not file_path:
            return

        try:
            self.gbif_df.to_csv(file_path, index=False, encoding="utf-8-sig")
            self._remember_save_dir(file_path)
            QMessageBox.information(self, "저장 완료", f"GBIF 분석 데이터를 저장했습니다.\n\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"GBIF 분석 데이터 저장 중 오류가 발생했습니다.\n\n{e}")

    def export_gbif_analysis_excel(self):
        if self.gbif_df is None or self.gbif_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "GBIF 분석 엑셀 저장",
            str(Path(self._get_last_save_dir()) / "gbif_analysis.xlsx"),
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            chart_type = self.gbif_chart_type_combo.currentData() or "bar"
            period_type = self.gbif_period_combo.currentData() or "summary"
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            filtered_df = self._filter_gbif_dataframe(
                self.gbif_df,
                year_from,
                year_to,
                month_from,
                month_to,
            )
            chart_list = self._gbif_temporal_charts(
                filtered_df,
                period_type,
                year_from,
                year_to,
                month_from,
                month_to,
            )
            self._write_gbif_analysis_workbook(
                file_path,
                filtered_df,
                chart_list,
                chart_type,
                {
                    **(self.gbif_search_summary or {}),
                    "yearFrom": year_from or "ALL",
                    "yearTo": year_to or "ALL",
                    "monthFrom": month_from,
                    "monthTo": month_to,
                    "downloadUrl": self.gbif_download_url_input.text().strip(),
                },
            )
            self._remember_save_dir(file_path)
            QMessageBox.information(
                self,
                "저장 완료",
                f"GBIF 분석 데이터와 그래프를 엑셀로 저장했습니다.\n\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"GBIF 분석 엑셀 저장 중 오류가 발생했습니다.\n\n{e}")

    def export_gbif_analysis_geojson(self):
        if self.gbif_df is None or self.gbif_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "QGIS용 GeoJSON 저장",
            str(Path(self._get_last_save_dir()) / "gbif_occurrences.geojson"),
            "GeoJSON Files (*.geojson);;JSON Files (*.json)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith((".geojson", ".json")):
            file_path += ".geojson"

        try:
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            filtered_df = self._filter_gbif_dataframe(
                self.gbif_df,
                year_from,
                year_to,
                month_from,
                month_to,
            )
            geojson = self._gbif_dataframe_to_geojson(filtered_df)
            feature_count = len(geojson["features"])
            if feature_count == 0:
                QMessageBox.warning(
                    self,
                    "QGIS 저장",
                    "저장할 수 있는 유효한 좌표가 없습니다.",
                )
                return

            Path(file_path).write_text(
                json.dumps(geojson, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self._remember_save_dir(file_path)
            QMessageBox.information(
                self,
                "저장 완료",
                f"QGIS에서 열 수 있는 GeoJSON 파일을 저장했습니다.\n\n{file_path}\n\n"
                f"저장된 좌표: {feature_count:,}개",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"QGIS용 GeoJSON 저장 중 오류가 발생했습니다.\n\n{e}")

    @staticmethod
    def _gbif_dataframe_to_geojson(df: pd.DataFrame) -> dict:
        features = []
        for _, row in df.iterrows():
            lat = pd.to_numeric(row.get("decimalLatitude"), errors="coerce")
            lon = pd.to_numeric(row.get("decimalLongitude"), errors="coerce")
            if pd.isna(lat) or pd.isna(lon):
                continue
            if not (-90 <= float(lat) <= 90 and -180 <= float(lon) <= 180):
                continue

            properties = {}
            for column, value in row.items():
                if column in {"decimalLatitude", "decimalLongitude"}:
                    continue
                if pd.isna(value):
                    properties[str(column)] = None
                else:
                    properties[str(column)] = value.item() if hasattr(value, "item") else value

            features.append(
                {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [float(lon), float(lat)],
                    },
                    "properties": properties,
                }
            )

        return {
            "type": "FeatureCollection",
            "name": "gbif_occurrences",
            "features": features,
        }

    @staticmethod
    def _write_gbif_analysis_workbook(
        file_path: str,
        df: pd.DataFrame,
        chart_list: list[dict],
        chart_type: str,
        summary: dict,
    ):
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows

        workbook = Workbook()
        summary_ws = workbook.active
        summary_ws.title = "Summary"

        summary_ws.append(["항목", "값"])
        for key, value in summary.items():
            summary_ws.append([key, value])
        summary_ws.append(["filteredRecords", len(df)])

        data_ws = workbook.create_sheet("Occurrences")
        for row in dataframe_to_rows(df, index=False, header=True):
            data_ws.append(row)

        chart_data_ws = workbook.create_sheet("ChartData")
        charts_ws = workbook.create_sheet("Charts")
        chart_data_row = 1
        chart_anchor_row = 1

        for chart_index, chart_data in enumerate(chart_list, start=1):
            labels = chart_data.get("labels", [])
            series = chart_data.get("series", [])
            title = f"{chart_data.get('title', 'Chart')} 기록 수"

            start_row = chart_data_row
            chart_data_ws.cell(row=start_row, column=1, value=title)
            header_row = start_row + 1
            chart_data_ws.cell(row=header_row, column=1, value="구분")
            for series_index, item in enumerate(series, start=2):
                chart_data_ws.cell(row=header_row, column=series_index, value=item.get("name", f"Series {series_index - 1}"))

            for label_index, label in enumerate(labels, start=header_row + 1):
                chart_data_ws.cell(row=label_index, column=1, value=label)
                for series_index, item in enumerate(series, start=2):
                    values = item.get("values", [])
                    value = values[label_index - header_row - 1] if label_index - header_row - 1 < len(values) else 0
                    chart_data_ws.cell(row=label_index, column=series_index, value=value)

            end_row = header_row + len(labels)
            end_col = 1 + len(series)
            if labels and series:
                chart = LineChart() if chart_type == "line" else BarChart()
                chart.title = title
                chart.y_axis.title = "기록 수"
                chart.x_axis.title = "구분"
                chart.height = 8
                chart.width = 16
                data_ref = Reference(
                    chart_data_ws,
                    min_col=2,
                    max_col=end_col,
                    min_row=header_row,
                    max_row=end_row,
                )
                category_ref = Reference(
                    chart_data_ws,
                    min_col=1,
                    min_row=header_row + 1,
                    max_row=end_row,
                )
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(category_ref)
                charts_ws.add_chart(chart, f"A{chart_anchor_row}")
                chart_anchor_row += 18

            chart_data_row = end_row + 3

        for worksheet in workbook.worksheets:
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells[:200]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)

        workbook.save(file_path)

    @staticmethod
    def _gbif_analysis_report_html(
        df: pd.DataFrame,
        summary: dict,
        chart_type: str = "bar",
        period_type: str = "year",
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
    ) -> str:
        df = MainWindow._filter_gbif_dataframe(
            df,
            year_from,
            year_to,
            month_from,
            month_to,
        )
        points = []
        for _, row in df.iterrows():
            lat = pd.to_numeric(row.get("decimalLatitude"), errors="coerce")
            lon = pd.to_numeric(row.get("decimalLongitude"), errors="coerce")
            if pd.isna(lat) or pd.isna(lon):
                continue
            points.append(
                {
                    "lat": float(lat),
                    "lon": float(lon),
                    "scientificName": str(row.get("scientificName", "") or ""),
                    "year": str(row.get("year", "") or ""),
                    "eventDate": str(row.get("eventDate", "") or ""),
                    "countryCode": str(row.get("countryCode", "") or ""),
                    "locality": str(row.get("locality", "") or ""),
                    "gbifID": str(row.get("gbifID", "") or ""),
                }
            )

        temporal_charts = MainWindow._gbif_temporal_charts(
            df,
            period_type,
            year_from,
            year_to,
            month_from,
            month_to,
        )

        points_json = json.dumps(points, ensure_ascii=False)
        temporal_json = json.dumps(temporal_charts, ensure_ascii=False)
        summary_json = json.dumps(summary, ensure_ascii=False)
        chart_type_json = json.dumps(chart_type if chart_type in {"bar", "line"} else "bar")
        return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GBIF Analysis Report</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
  <style>
    html, body {{
      margin: 0;
      min-height: 100%;
      background: #eef3f7;
      color: #1e293b;
      font-family: "Segoe UI", "Malgun Gothic", sans-serif;
    }}
    header {{
      background: #fbfdff;
      border-bottom: 1px solid #d8e2ec;
      padding: 18px 22px;
    }}
    h1 {{
      font-size: 22px;
      margin: 0 0 6px;
    }}
    .meta {{
      color: #52667a;
      font-size: 13px;
      line-height: 1.5;
    }}
    main {{
      display: grid;
      grid-template-columns: minmax(320px, 1.15fr) minmax(320px, 0.85fr);
      gap: 14px;
      padding: 14px;
    }}
    section {{
      background: #ffffff;
      border: 1px solid #d8e2ec;
      border-radius: 8px;
      overflow: hidden;
    }}
    .section-title {{
      border-bottom: 1px solid #e2e8f0;
      font-size: 15px;
      font-weight: 800;
      padding: 12px 14px;
    }}
    #map {{
      height: calc(100vh - 150px);
      min-height: 440px;
    }}
    .chart {{
      overflow: auto;
      padding: 12px 14px 16px;
    }}
    .chart-panel {{
      max-height: calc(100vh - 150px);
      overflow: auto;
    }}
    .chart-block {{
      border-bottom: 1px solid #e2e8f0;
    }}
    .chart-block:last-child {{
      border-bottom: none;
    }}
    .chart svg {{
      display: block;
      height: auto;
      max-width: 100%;
    }}
    .bar-row {{
      align-items: center;
      display: grid;
      grid-template-columns: 62px 1fr 52px;
      gap: 10px;
      margin: 7px 0;
    }}
    .bar {{
      background: #dbeafe;
      border-radius: 4px;
      height: 18px;
      overflow: hidden;
    }}
    .bar-fill {{
      background: #0f766e;
      height: 100%;
    }}
    .count {{
      color: #475569;
      font-variant-numeric: tabular-nums;
      text-align: right;
    }}
    .axis-label {{
      fill: #475569;
      font-size: 12px;
    }}
    .line-point {{
      fill: #ffffff;
      stroke-width: 2;
    }}
    .legend {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px 12px;
      margin-bottom: 10px;
    }}
    .legend-item {{
      align-items: center;
      color: #475569;
      display: inline-flex;
      font-size: 12px;
      gap: 5px;
    }}
    .legend-swatch {{
      border-radius: 999px;
      display: inline-block;
      height: 10px;
      width: 10px;
    }}
    @media (max-width: 900px) {{
      main {{
        grid-template-columns: 1fr;
      }}
      #map {{
        height: 520px;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>GBIF 분석 리포트</h1>
    <div class="meta" id="summary"></div>
  </header>
  <main>
    <section>
      <div class="section-title">좌표 지도</div>
      <div id="map"></div>
    </section>
    <section>
      <div class="section-title">기록 수 비교</div>
      <div class="chart-panel" id="chartPanel"></div>
    </section>
  </main>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <script>
    const points = {points_json};
    const chartList = {temporal_json};
    const summary = {summary_json};
    const chartType = {chart_type_json};
    const escapeHtml = (value) => String(value)
      .replaceAll("&", "&amp;")
      .replaceAll("<", "&lt;")
      .replaceAll(">", "&gt;")
      .replaceAll('"', "&quot;")
      .replaceAll("'", "&#039;");

    document.getElementById("summary").innerHTML = `
      일치 이름: <b>${{escapeHtml(summary.matchedName || "")}}</b> /
      taxonKey: ${{escapeHtml(summary.taxonKey || "")}} /
      국가: ${{escapeHtml(summary.countryCode || "ALL")}} /
      전체 좌표 기록: ${{Number(summary.totalRecords || 0).toLocaleString()}}건 /
      표시: ${{Number(summary.shownRecords || 0).toLocaleString()}}건
    `;

    const map = L.map("map");
    L.tileLayer("https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png", {{
      maxZoom: 19,
      attribution: "&copy; OpenStreetMap contributors"
    }}).addTo(map);

    const bounds = [];
    points.forEach((item) => {{
      const position = [item.lat, item.lon];
      bounds.push(position);
      const popup = `
        <b>${{escapeHtml(item.scientificName)}}</b><br>
        Year: ${{escapeHtml(item.year || "")}}<br>
        Event date: ${{escapeHtml(item.eventDate || "")}}<br>
        Country: ${{escapeHtml(item.countryCode || "")}}<br>
        Locality: ${{escapeHtml(item.locality || "")}}<br>
        GBIF ID: ${{escapeHtml(item.gbifID || "")}}
      `;
      L.circleMarker(position, {{
        radius: 5,
        color: "#0f766e",
        fillColor: "#14b8a6",
        fillOpacity: 0.58,
        weight: 1
      }}).bindPopup(popup).addTo(map);
    }});

    if (bounds.length === 1) {{
      map.setView(bounds[0], 8);
    }} else if (bounds.length > 1) {{
      map.fitBounds(bounds, {{ padding: [30, 30], maxZoom: 9 }});
    }} else {{
      map.setView([20, 0], 2);
    }}

    const chartPanel = document.getElementById("chartPanel");
    const colors = ["#0f766e", "#2563eb", "#b45309", "#7c3aed", "#be123c", "#0891b2", "#4d7c0f", "#9333ea"];
    const renderChart = (chartData) => {{
      const labels = chartData.labels || [];
      const series = chartData.series || [];
      const allCounts = series.flatMap((line) => line.values || []);
      const maxCount = Math.max(1, ...allCounts);
      const legend = () => series.length > 1
      ? `<div class="legend">${{series.map((line, index) => `
          <span class="legend-item">
            <span class="legend-swatch" style="background: ${{colors[index % colors.length]}}"></span>
            ${{escapeHtml(line.name)}}
          </span>
        `).join("")}}</div>`
      : "";
      const renderBarChart = () => legend() + series.flatMap((line, seriesIndex) =>
      labels.map((label, index) => {{
        const value = Number((line.values || [])[index] || 0);
        const rowLabel = series.length > 1 ? `${{line.name}} ${{label}}` : label;
        return `
          <div class="bar-row">
            <div>${{escapeHtml(rowLabel)}}</div>
            <div class="bar"><div class="bar-fill" style="width: ${{Math.max(2, value / maxCount * 100)}}%; background: ${{colors[seriesIndex % colors.length]}}"></div></div>
            <div class="count">${{value.toLocaleString()}}</div>
          </div>
        `;
      }})
    ).join("");

      const renderLineChart = () => {{
      const width = 680;
      const height = 320;
      const padding = {{ top: 24, right: 28, bottom: 42, left: 52 }};
      const plotWidth = width - padding.left - padding.right;
      const plotHeight = height - padding.top - padding.bottom;
      const xStep = labels.length > 1 ? plotWidth / (labels.length - 1) : 0;
      const pointFor = (value, index) => {{
        const x = padding.left + index * xStep;
        const y = padding.top + plotHeight - (value / maxCount * plotHeight);
        return {{ x, y }};
      }};
      const lines = series.map((line, seriesIndex) => {{
        const color = colors[seriesIndex % colors.length];
        const pointsAttr = (line.values || []).map((value, index) => {{
          const point = pointFor(Number(value || 0), index);
          return `${{point.x}},${{point.y}}`;
        }}).join(" ");
        const markers = (line.values || []).map((value, index) => {{
          const numericValue = Number(value || 0);
          const point = pointFor(numericValue, index);
          const labelY = point.y - 10 < 12 ? point.y + 22 : point.y - 10;
          const valueLabel = series.length === 1 || numericValue > 0
            ? `<text class="axis-label" x="${{point.x}}" y="${{labelY}}" text-anchor="middle">${{numericValue.toLocaleString()}}</text>`
            : "";
          return `
            <circle class="line-point" cx="${{point.x}}" cy="${{point.y}}" r="4" stroke="${{color}}"></circle>
            ${{valueLabel}}
          `;
        }}).join("");
        return `
          <polyline points="${{pointsAttr}}" fill="none" stroke="${{color}}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"></polyline>
          ${{markers}}
        `;
      }}).join("");
      const axisLabels = labels.map((label, index) => {{
        const point = pointFor(0, index);
        const visible = labels.length <= 18 || index % Math.ceil(labels.length / 18) === 0;
        return visible
          ? `<text class="axis-label" x="${{point.x}}" y="${{height - 16}}" text-anchor="middle">${{escapeHtml(label)}}</text>`
          : "";
      }}).join("");
      return legend() + `
        <svg viewBox="0 0 ${{width}} ${{height}}" role="img" aria-label="연도별 기록 수 선 그래프">
          <line x1="${{padding.left}}" y1="${{padding.top}}" x2="${{padding.left}}" y2="${{height - padding.bottom}}" stroke="#cbd5e1"></line>
          <line x1="${{padding.left}}" y1="${{height - padding.bottom}}" x2="${{width - padding.right}}" y2="${{height - padding.bottom}}" stroke="#cbd5e1"></line>
          <text class="axis-label" x="8" y="${{padding.top + 6}}">기록 수</text>
          ${{lines}}
          ${{axisLabels}}
        </svg>
      `;
    }};

      const body = labels.length && series.length
      ? (chartType === "line" ? renderLineChart() : renderBarChart())
      : `${{chartData.title}} 정보가 있는 기록이 없습니다.`;
      return `
        <div class="chart-block">
          <div class="section-title">${{escapeHtml(chartData.title)}} 기록 수</div>
          <div class="chart">${{body}}</div>
        </div>
      `;
    }};

    chartPanel.innerHTML = chartList.length
      ? chartList.map(renderChart).join("")
      : "표시할 그래프 데이터가 없습니다.";
  </script>
</body>
</html>
"""

    @staticmethod
    def _gbif_temporal_charts(
        df: pd.DataFrame,
        period_type: str,
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
    ) -> list[dict]:
        if period_type == "summary":
            return [
                MainWindow._gbif_temporal_chart_data(
                    df,
                    "year",
                    year_from,
                    year_to,
                    month_from,
                    month_to,
                ),
                MainWindow._gbif_temporal_chart_data(
                    df,
                    "month",
                    year_from,
                    year_to,
                    month_from,
                    month_to,
                ),
                MainWindow._gbif_temporal_chart_data(
                    df,
                    "season",
                    year_from,
                    year_to,
                    month_from,
                    month_to,
                ),
            ]

        return [
            MainWindow._gbif_temporal_chart_data(
                df,
                period_type,
                year_from,
                year_to,
                month_from,
                month_to,
            )
        ]

    @staticmethod
    def _gbif_temporal_chart_data(
        df: pd.DataFrame,
        period_type: str,
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
    ) -> dict:
        if period_type == "month":
            month_names = [
                "1월",
                "2월",
                "3월",
                "4월",
                "5월",
                "6월",
                "7월",
                "8월",
                "9월",
                "10월",
                "11월",
                "12월",
            ]
            month_keys = list(range(month_from, month_to + 1))
            series = MainWindow._gbif_period_series_by_year(
                df,
                year_from,
                year_to,
                month_keys,
                lambda month_values, key: int((month_values == key).sum()),
            )
            return {
                "title": "월별",
                "labels": [month_names[month - 1] for month in month_keys],
                "series": series,
            }

        if period_type == "season":
            season_defs = [
                ("봄", {3, 4, 5}),
                ("여름", {6, 7, 8}),
                ("가을", {9, 10, 11}),
                ("겨울", {12, 1, 2}),
            ]
            season_defs = [
                (label, {month for month in months if month_from <= month <= month_to})
                for label, months in season_defs
            ]
            season_defs = [(label, months) for label, months in season_defs if months]
            series = MainWindow._gbif_period_series_by_year(
                df,
                year_from,
                year_to,
                [months for _, months in season_defs],
                lambda month_values, months: int(month_values[month_values.isin(months)].count()),
            )
            return {
                "title": "계절별",
                "labels": [label for label, _ in season_defs],
                "series": series,
            }

        years = pd.to_numeric(df.get("year"), errors="coerce").dropna().astype(int)
        year_counts = years.value_counts().sort_index()
        return {
            "title": "연도별",
            "labels": [str(year) for year in year_counts.index],
            "series": [
                {
                    "name": "전체",
                    "values": [int(count) for count in year_counts.values],
                }
            ],
        }

    @staticmethod
    def _gbif_period_series_by_year(
        df: pd.DataFrame,
        year_from: int | None,
        year_to: int | None,
        keys: list,
        count_for_key,
    ) -> list[dict]:
        year_values = MainWindow._gbif_year_values(df)
        candidate_years = sorted(year_values.unique())
        if year_from is not None:
            candidate_years = [year for year in candidate_years if year >= year_from]
        if year_to is not None:
            candidate_years = [year for year in candidate_years if year <= year_to]

        if not candidate_years:
            month_values = MainWindow._gbif_month_values(df)
            return [
                {
                    "name": "전체",
                    "values": [count_for_key(month_values, key) for key in keys],
                }
            ]

        series = []
        for year in candidate_years:
            year_df = df.loc[year_values.reindex(df.index).eq(year)]
            month_values = MainWindow._gbif_month_values(year_df)
            series.append(
                {
                    "name": str(year),
                    "values": [count_for_key(month_values, key) for key in keys],
                }
            )
        return series

    @staticmethod
    def _filter_gbif_dataframe(
        df: pd.DataFrame,
        year_from: int | None,
        year_to: int | None,
        month_from: int,
        month_to: int,
    ) -> pd.DataFrame:
        filtered_df = df.copy()
        if filtered_df.empty:
            return filtered_df

        if year_from is not None or year_to is not None:
            year_values = MainWindow._gbif_year_values(filtered_df).reindex(filtered_df.index)
            mask = pd.Series(True, index=filtered_df.index)
            if year_from is not None:
                mask &= year_values.ge(year_from).fillna(False)
            if year_to is not None:
                mask &= year_values.le(year_to).fillna(False)
            filtered_df = filtered_df.loc[mask]

        if month_from > 1 or month_to < 12:
            month_values = MainWindow._gbif_month_values(filtered_df).reindex(filtered_df.index)
            mask = month_values.ge(month_from).fillna(False) & month_values.le(month_to).fillna(False)
            filtered_df = filtered_df.loc[mask]

        return filtered_df

    @staticmethod
    def _gbif_year_values(df: pd.DataFrame) -> pd.Series:
        if "year" not in df.columns:
            return pd.Series(dtype="int64")

        return pd.to_numeric(df["year"], errors="coerce").dropna().astype(int)

    @staticmethod
    def _gbif_month_values(df: pd.DataFrame) -> pd.Series:
        if "month" in df.columns:
            month_values = pd.to_numeric(df["month"], errors="coerce")
        else:
            month_values = pd.Series(float("nan"), index=df.index, dtype="float64")

        if "eventDate" in df.columns:
            parsed_months = pd.to_datetime(
                df["eventDate"],
                errors="coerce",
                utc=True,
            ).dt.month
            month_values = month_values.fillna(parsed_months)

        return month_values.dropna().astype(int).loc[lambda values: values.between(1, 12)]

    def _build_auto_mapping_settings_tab(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        layout.addWidget(
            self._build_section_header(
                "자동 매핑 설정",
                "원본 파일의 컬럼명이 아래 후보 이름을 포함하면 해당 GBIF 필드로 자동 선택됩니다. 여러 이름은 쉼표로 구분하세요.",
            )
        )

        self.auto_mapping_path_label = QLabel(f"설정 파일: {CUSTOM_AUTO_MAPPING_PATH}")
        self.auto_mapping_path_label.setObjectName("metaValue")
        self.auto_mapping_path_label.setWordWrap(True)
        layout.addWidget(self.auto_mapping_path_label)

        self.auto_mapping_table = QTableWidget()
        self.auto_mapping_table.setColumnCount(2)
        self.auto_mapping_table.setHorizontalHeaderLabels(
            ["GBIF 필드", "후보 컬럼명"]
        )
        self.auto_mapping_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.auto_mapping_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.auto_mapping_table.verticalHeader().setVisible(False)
        self.auto_mapping_table.setAlternatingRowColors(True)
        layout.addWidget(self.auto_mapping_table)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)

        self.btn_reload_auto_mapping = QPushButton("다시 불러오기")
        self.btn_reload_auto_mapping.setProperty("role", "secondary")
        self.btn_reload_auto_mapping.clicked.connect(self.load_auto_mapping_settings)

        self.btn_reset_auto_mapping = QPushButton("기본값으로 초기화")
        self.btn_reset_auto_mapping.setProperty("role", "secondary")
        self.btn_reset_auto_mapping.clicked.connect(self.reset_auto_mapping_settings)

        self.btn_save_auto_mapping = QPushButton("자동 매핑 설정 저장")
        self.btn_save_auto_mapping.setProperty("role", "primary")
        self.btn_save_auto_mapping.clicked.connect(self.save_auto_mapping_settings)

        button_layout.addWidget(self.btn_reload_auto_mapping)
        button_layout.addWidget(self.btn_reset_auto_mapping)
        button_layout.addStretch()
        button_layout.addWidget(self.btn_save_auto_mapping)
        layout.addLayout(button_layout)

        self.load_auto_mapping_settings()
        return panel

    def load_auto_mapping_settings(self):
        try:
            rules = load_auto_mapping_rules()
        except Exception as e:
            QMessageBox.warning(
                self,
                "자동 매핑 설정",
                f"자동 매핑 설정을 불러오지 못했습니다.\n기본값으로 표시합니다.\n\n{e}",
            )
            rules = get_default_auto_mapping_rules()

        self.auto_mapping_table.setRowCount(len(ALL_DWC_FIELDS))

        for row, field in enumerate(ALL_DWC_FIELDS):
            key_item = QTableWidgetItem(field["key"])
            key_item.setFlags(key_item.flags() & ~Qt.ItemIsEditable)
            self.auto_mapping_table.setItem(row, 0, key_item)

            keywords = ", ".join(rules.get(field["key"], []))
            self.auto_mapping_table.setItem(row, 1, QTableWidgetItem(keywords))

        self.auto_mapping_table.resizeRowsToContents()

    def _rules_from_auto_mapping_table(self) -> dict[str, list[str]]:
        rules = {}
        for row in range(self.auto_mapping_table.rowCount()):
            key_item = self.auto_mapping_table.item(row, 0)
            value_item = self.auto_mapping_table.item(row, 1)
            if key_item is None:
                continue

            key = key_item.text().strip()
            value = value_item.text() if value_item is not None else ""
            rules[key] = [keyword.strip() for keyword in value.split(",") if keyword.strip()]

        return rules

    def save_auto_mapping_settings(self):
        try:
            save_auto_mapping_rules(self._rules_from_auto_mapping_table())

            if self.source_df is not None:
                self.build_mapping_ui(self.source_df.columns.tolist())
                self.btn_coordinate_preview.setEnabled(False)
                self.btn_export.setEnabled(False)

            QMessageBox.information(
                self,
                "저장 완료",
                f"자동 매핑 설정을 저장했습니다.\n\n{CUSTOM_AUTO_MAPPING_PATH}",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"자동 매핑 설정 저장 중 오류가 발생했습니다.\n\n{e}")

    def reset_auto_mapping_settings(self):
        reply = QMessageBox.question(
            self,
            "기본값으로 초기화",
            "로컬 자동 매핑 설정을 삭제하고 기본값으로 되돌릴까요?",
        )
        if reply != QMessageBox.Yes:
            return

        try:
            reset_auto_mapping_rules()
            self.load_auto_mapping_settings()

            if self.source_df is not None:
                self.build_mapping_ui(self.source_df.columns.tolist())
                self.btn_coordinate_preview.setEnabled(False)
                self.btn_export.setEnabled(False)

            QMessageBox.information(self, "초기화 완료", "자동 매핑 설정을 기본값으로 되돌렸습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"자동 매핑 설정 초기화 중 오류가 발생했습니다.\n\n{e}")

    def _configure_table(self, table: QTableWidget):
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectItems)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setWordWrap(False)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.horizontalHeader().setMinimumSectionSize(110)
        table.setShowGrid(False)

    def _get_active_fields(self) -> list[dict]:
        return [get_dwc_field(key) for key in self.active_field_keys]

    def _update_mapping_summary(self):
        self.mapping_summary.setText(f"현재 {len(self.active_field_keys)}개 필드")

    def show_about(self):
        lines = [
            f"<h3>{APP_NAME}</h3>",
            f"<p>Version {APP_VERSION} · {APP_SUBTITLE}</p>",
            f"<p><b>Made by {AUTHOR_NAME}</b></p>",
            f"<p>{COPYRIGHT_TEXT}</p>",
        ]

        if AUTHOR_EMAIL:
            lines.append(
                f'<p>email: '
                f'<a href="mailto:{AUTHOR_EMAIL}">{AUTHOR_EMAIL}</a></p>'
            )
        else:
            lines.append("<p>문의사항은 이메일로 문의해주세요.</p>")

        if AUTHOR_GITHUB:
            lines.append(f'<p>GitHub: <a href="{AUTHOR_GITHUB}">{AUTHOR_GITHUB}</a></p>')
        
        lines.append( f'<p>문의사항은 이메일로 문의해주세요!')

        message = QMessageBox(self)
        message.setWindowTitle("정보")
        message.setTextFormat(Qt.RichText)
        message.setText("\n".join(lines))
        message.exec()

    def _snapshot_mapping_state(self) -> tuple[dict, dict]:
        combo_state = {}
        manual_state = {}

        for key, combo in self.combo_boxes.items():
            try:
                combo_state[key] = combo.currentText()
            except RuntimeError:
                continue

        for key, widget in self.manual_inputs.items():
            try:
                manual_state[key] = widget.text()
            except RuntimeError:
                continue

        return combo_state, manual_state

    def _restore_mapping_state(self, combo_state: dict, manual_state: dict):
        for key, value in combo_state.items():
            combo = self.combo_boxes.get(key)
            if combo and value:
                index = combo.findText(value)
                if index >= 0:
                    combo.setCurrentIndex(index)
                else:
                    combo.setCurrentText(value)

        for key, value in manual_state.items():
            line_edit = self.manual_inputs.get(key)
            if line_edit is not None:
                line_edit.setText(value)

    def _reset_active_fields(self):
        self.active_field_keys = [field["key"] for field in get_default_dwc_fields()]
        self._update_mapping_summary()

    def _get_last_open_dir(self) -> str:
        last_dir = self.settings.value("file_dialog/last_open_dir", "", str)
        if last_dir and Path(last_dir).is_dir():
            return last_dir
        return ""

    def _remember_open_dir(self, file_path: str) -> None:
        parent_dir = Path(file_path).expanduser().parent
        if parent_dir.is_dir():
            self.settings.setValue("file_dialog/last_open_dir", str(parent_dir))

    def _get_last_save_dir(self) -> str:
        last_dir = self.settings.value("file_dialog/last_save_dir", "", str)
        if last_dir and Path(last_dir).is_dir():
            return last_dir
        return str(OUTPUT_DIR)

    def _remember_save_dir(self, file_path: str) -> None:
        parent_dir = Path(file_path).expanduser().parent
        if parent_dir.is_dir():
            self.settings.setValue("file_dialog/last_save_dir", str(parent_dir))

    def load_excel(self):
        file_path = self._select_source_file()
        if not file_path:
            return

        try:
            selected_sheet = self._select_source_sheet(file_path)
            if not selected_sheet:
                return

            raw_df = ExcelService.read_excel_raw(file_path, sheet_name=selected_sheet)
            self._apply_loaded_file(file_path, selected_sheet, raw_df)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일을 불러오지 못했습니다.\n\n{e}")

    def _select_source_file(self) -> str | None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "파일 선택",
            self._get_last_open_dir(),
            "Excel/CSV Files (*.xlsx *.xls *.csv)",
        )
        if not file_path:
            return None

        self._remember_open_dir(file_path)
        return file_path

    def _select_source_sheet(self, file_path: str) -> str | None:
        sheet_names = ExcelService.get_sheet_names(file_path)
        if not sheet_names:
            raise ValueError("시트가 없는 엑셀 파일입니다.")

        selected_sheet, ok = QInputDialog.getItem(
            self,
            "시트 선택",
            "불러올 시트를 선택하세요:",
            sheet_names,
            0,
            False,
        )
        if not ok or not selected_sheet:
            return None

        return selected_sheet

    def _apply_loaded_file(self, file_path: str, sheet_name: str, raw_df: pd.DataFrame) -> None:
        self._set_loaded_file_state(file_path, sheet_name, raw_df)
        self._reset_loaded_file_ui(file_path, sheet_name, raw_df)

    def _set_loaded_file_state(
        self,
        file_path: str,
        sheet_name: str,
        raw_df: pd.DataFrame,
    ) -> None:
        self.raw_df = raw_df
        self.current_file_path = file_path
        self.current_sheet_name = sheet_name
        self.source_df = None
        self.result_df = None
        self.selected_header_row = None
        self._reset_active_fields()

    def _reset_loaded_file_ui(
        self,
        file_path: str,
        sheet_name: str,
        raw_df: pd.DataFrame,
    ) -> None:
        self._update_loaded_file_labels(file_path, sheet_name)
        self.show_preview_raw(raw_df)
        self.clear_mapping_ui()
        self.result_table.clear()
        self.result_table.setRowCount(0)
        self.result_table.setColumnCount(0)
        self.reset_missing_summary_panel()
        self._set_file_loaded_buttons_enabled()

    def _update_loaded_file_labels(self, file_path: str, sheet_name: str) -> None:
        self.lbl_file.setText(f"파일: {Path(file_path).name}")
        self.lbl_sheet.setText(sheet_name)
        self.lbl_header.setText("선택된 헤더 행: 없음")
        self.preview_position_label.setText("현재 선택 위치: 없음")

    def _set_file_loaded_buttons_enabled(self) -> None:
        self.btn_apply_header.setEnabled(True)
        self.btn_add_field.setEnabled(False)
        self.btn_source_coordinate_preview.setEnabled(False)
        self.btn_preview_result.setEnabled(False)
        self.btn_coordinate_preview.setEnabled(False)
        self.btn_export.setEnabled(False)
    
    def show_preview_raw(self, df):
        preview_df = df

        self.preview_table.clear()
        self.preview_table.setRowCount(len(preview_df))
        self.preview_table.setColumnCount(len(preview_df.columns))
        self.preview_table.setHorizontalHeaderLabels(
            [f"Column {i + 1}" for i in range(len(preview_df.columns))]
        )

        for row_idx in range(len(preview_df)):
            for col_idx in range(len(preview_df.columns)):
                value = preview_df.iloc[row_idx, col_idx]
                self.preview_table.setItem(
                    row_idx,
                    col_idx,
                    QTableWidgetItem("" if value is None else str(value)),
                )

        self.preview_table.resizeColumnsToContents()
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def show_preview_with_header(self, df):
        preview_df = df

        self.preview_table.clear()
        self.preview_table.setRowCount(len(preview_df))
        self.preview_table.setColumnCount(len(preview_df.columns))
        self.preview_table.setHorizontalHeaderLabels([str(col) for col in preview_df.columns.tolist()])

        for row_idx in range(len(preview_df)):
            for col_idx, _ in enumerate(preview_df.columns):
                value = preview_df.iloc[row_idx, col_idx]
                self.preview_table.setItem(
                    row_idx,
                    col_idx,
                    QTableWidgetItem("" if value is None else str(value)),
                )

        self.preview_table.resizeColumnsToContents()
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def select_header_row(self, row, col):
        if self.raw_df is None:
            return

        self.selected_header_row = row
        self.lbl_header.setText(f"선택된 헤더 행: {row + 1}행")
        self.preview_position_label.setText(
            f"현재 선택 위치: {row + 1}행, {col + 1}열"
        )
        self.highlight_selected_row(row)

    def highlight_selected_row(self, row):
        for r in range(self.preview_table.rowCount()):
            for c in range(self.preview_table.columnCount()):
                item = self.preview_table.item(r, c)
                if item:
                    item.setBackground(QColor("#fff4bf") if r == row else QColor("#ffffff"))

    def apply_selected_header(self):
        # 헤더 적용 전에 필요한 상태인지 체크
        if self.current_file_path is None or self.raw_df is None:
            QMessageBox.warning(self, "경고", "먼저 엑셀 파일을 업로드해주세요.")
            return

        # 헤더로 사용할 행이 선택되었는지 체크
        if self.selected_header_row is None:
            QMessageBox.warning(self, "경고", "헤더로 사용할 행을 먼저 클릭하세요.")
            return

        try:
            # 선택된 행을 헤더로 적용해서 데이터프레임 다시 불러오기
            source_df = ExcelService.read_excel_with_header(
                self.current_file_path,
                header_row=self.selected_header_row,
                sheet_name=self.current_sheet_name,
            )
            
            # 열 이름 정리 (공백 제거, 중복 방지 등)
            source_df.columns = ExcelService.clean_column_names(source_df.columns)
            # 헤더 적용된 데이터프레임으로 UI 업데이트
            self._apply_header_data(source_df)

            # 적용 완료 메시지
            QMessageBox.information(
                self,
                "완료",
                f"{self.selected_header_row + 1}행을 헤더로 적용했습니다.",
            )

        except Exception as e:
            QMessageBox.critical(self, "오류", f"헤더 적용 중 오류가 발생했습니다.\n\n{e}")

    # 헤더가 적용된 데이터프레임으로 UI 업데이트
    def _apply_header_data(self, source_df: pd.DataFrame) -> None:
        self.source_df = source_df
        self.show_preview_with_header(source_df)
        self.preview_position_label.setText(
            f"현재 선택 위치: {self.selected_header_row + 1}행 (헤더 적용됨, 열 선택은 마지막 클릭 기준)"
        )
        # 헤더가 적용된 데이터프레임의 열 이름을 기준으로 매핑 UI 구축
        self.build_mapping_ui(source_df.columns.tolist())
        # 헤더가 적용되었으므로 관련 버튼 활성화
        self._set_header_applied_buttons_enabled()

    def _set_header_applied_buttons_enabled(self) -> None:
        self.btn_add_field.setEnabled(True)
        self.btn_source_coordinate_preview.setEnabled(True)
        self.btn_preview_result.setEnabled(True)
        self.btn_coordinate_preview.setEnabled(False)
        self.btn_export.setEnabled(False)

    def add_optional_field(self):
        if self.source_df is None:
            QMessageBox.warning(self, "경고", "헤더를 먼저 적용해주세요.")
            return

        optional_fields = get_optional_dwc_fields(self.active_field_keys)
        if not optional_fields:
            QMessageBox.information(self, "안내", "추가할 수 있는 GBIF 컬럼이 더 없습니다.")
            return

        display_items = [
            f'{field["label"]} - {field["description"]}'
            for field in optional_fields
        ]
        field_by_display = {
            f'{field["label"]} - {field["description"]}': field
            for field in optional_fields
        }

        selected_display, ok = QInputDialog.getItem(
            self,
            "컬럼 추가",
            "추가할 GBIF 컬럼을 선택하세요:",
            display_items,
            0,
            False,
        )
        if not ok or not selected_display:
            return

        selected_field = field_by_display[selected_display]
        self.active_field_keys.append(selected_field["key"])

        combo_state, manual_state = self._snapshot_mapping_state()
        self.build_mapping_ui(self.source_df.columns.tolist(), combo_state, manual_state)
        self.btn_coordinate_preview.setEnabled(False)
        self.btn_export.setEnabled(False)

    def build_mapping_ui(
        self,
        source_columns: list[str],
        combo_state: dict | None = None,
        manual_state: dict | None = None,
    ):
        self.clear_mapping_ui()

        for field in self._get_active_fields():
            frame = QFrame()
            frame.setFrameShape(QFrame.StyledPanel)
            frame.setProperty("class", "mappingCard")

            row_layout = QHBoxLayout(frame)
            row_layout.setContentsMargins(16, 13, 16, 13)
            row_layout.setSpacing(16)

            label_layout = QVBoxLayout()
            label_layout.setSpacing(4)

            lbl_name = QLabel(field["label"])
            lbl_name.setProperty("class", "fieldName")

            lbl_desc = QLabel(field["description"])
            lbl_desc.setWordWrap(True)
            lbl_desc.setProperty("class", "fieldDescription")

            label_layout.addWidget(lbl_name)
            label_layout.addWidget(lbl_desc)

            if field["key"] == "basisOfRecord":
                combo = QComboBox()
                combo.setMinimumWidth(300)
                combo.addItems(
                    [
                        "",
                        "HumanObservation",
                        "MachineObservation",
                        "PreservedSpecimen",
                        "LivingSpecimen",
                        "MaterialSample",
                    ]
                )
                row_layout.addLayout(label_layout, stretch=3)
                row_layout.addWidget(combo, stretch=2)
                self.combo_boxes[field["key"]] = combo

            elif field["key"] == "collectionCode":
                sub_layout = QVBoxLayout()
                sub_layout.setSpacing(8)

                combo = QComboBox()
                combo.setMinimumWidth(300)
                combo.addItem("")
                combo.addItems([str(col) for col in source_columns])

                line_edit = QLineEdit()
                line_edit.setPlaceholderText("또는 직접 입력 (예: PL)")

                sub_layout.addWidget(combo)
                sub_layout.addWidget(line_edit)
                row_layout.addLayout(label_layout, stretch=3)
                row_layout.addLayout(sub_layout, stretch=2)

                self.combo_boxes[field["key"]] = combo
                self.manual_inputs[field["key"]] = line_edit

            elif field["key"] == "countryCode":
                sub_layout = QVBoxLayout()
                sub_layout.setSpacing(8)

                combo = QComboBox()
                combo.setMinimumWidth(300)
                combo.addItems(ISO_COUNTRY_CODES)
                combo.setToolTip("ISO 3166-1 alpha-2 국가 코드")

                line_edit = QLineEdit()
                line_edit.setPlaceholderText("또는 직접 입력 (예: KR)")
                line_edit.setToolTip("직접 입력값이 있으면 그 값을 우선 사용합니다.")

                sub_layout.addWidget(combo)
                sub_layout.addWidget(line_edit)
                row_layout.addLayout(label_layout, stretch=3)
                row_layout.addLayout(sub_layout, stretch=2)

                self.combo_boxes[field["key"]] = combo
                self.manual_inputs[field["key"]] = line_edit

            else:
                combo = QComboBox()
                combo.setMinimumWidth(300)
                combo.addItem("")
                combo.addItems([str(col) for col in source_columns])

                matched_col = auto_match_column(field["key"], source_columns)
                if matched_col:
                    combo.setCurrentText(matched_col)
                    combo.setProperty("autoMatched", True)

                row_layout.addLayout(label_layout, stretch=3)
                row_layout.addWidget(combo, stretch=2)
                self.combo_boxes[field["key"]] = combo

            self.mapping_layout.addWidget(frame)

        if combo_state or manual_state:
            self._restore_mapping_state(combo_state or {}, manual_state or {})

        self._update_mapping_summary()

    def clear_mapping_ui(self):
        while self.mapping_layout.count():
            item = self.mapping_layout.takeAt(0)

            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
                continue

            layout = item.layout()
            if layout is not None:
                self._clear_layout(layout)

        self.combo_boxes.clear()
        self.manual_inputs.clear()

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)

            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
                continue

            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)

    def get_selected_mapping(self) -> dict:
        combo_map = {}

        for dwc_key, combo in self.combo_boxes.items():
            if combo is None:
                continue

            try:
                combo_map[dwc_key] = combo.currentText()
            except RuntimeError:
                continue

        return MappingService.build_mapping(combo_map)

    def export_mapped_excel(self):
        if self.result_df is None:
            QMessageBox.warning(self, "경고", "먼저 '결과 확인'으로 변환 결과를 생성하세요.")
            return

        try:
            final_df = self.get_result_df_from_table()
            if final_df is None:
                QMessageBox.warning(self, "경고", "저장할 결과가 없습니다.")
                return

            output_path, _ = QFileDialog.getSaveFileName(
                self,
                "저장할 파일 경로 선택",
                str(Path(self._get_last_save_dir()) / "dwc_mapped_output.xlsx"),
                "Excel Files (*.xlsx)",
            )

            if not output_path:
                return

            ExcelService.save_excel(final_df, output_path)
            self._remember_save_dir(output_path)
            QMessageBox.information(self, "완료", f"엑셀 파일이 저장되었습니다.\n\n{output_path}")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류가 발생했습니다.\n\n{e}")

    def _select_coordinate_column(self, title: str, prompt: str, preferred_key: str) -> str | None:
        if self.source_df is None:
            return None

        columns = [str(column) for column in self.source_df.columns.tolist()]
        if not columns:
            return None

        current_index = 0
        matched_column = auto_match_column(preferred_key, columns)
        if matched_column in columns:
            current_index = columns.index(matched_column)

        selected_column, ok = QInputDialog.getItem(
            self,
            title,
            prompt,
            columns,
            current_index,
            False,
        )
        if not ok or not selected_column:
            return None
        return selected_column

    def preview_source_coordinates_on_map(self):
        if self.source_df is None:
            QMessageBox.warning(self, "경고", "먼저 헤더를 적용해 주세요.")
            return

        lat_column = self._select_coordinate_column(
            "원본 좌표 확인",
            "위도 컬럼을 선택하세요:",
            "decimalLatitude",
        )
        if not lat_column:
            return

        lon_column = self._select_coordinate_column(
            "원본 좌표 확인",
            "경도 컬럼을 선택하세요:",
            "decimalLongitude",
        )
        if not lon_column:
            return

        if lat_column == lon_column:
            QMessageBox.warning(self, "경고", "위도와 경도는 서로 다른 컬럼을 선택해 주세요.")
            return

        label_columns = [
            str(column)
            for column in self.source_df.columns.tolist()
            if str(column) not in {lat_column, lon_column}
        ][:8]
        self._preview_coordinate_dataframe(
            self.source_df,
            lat_column,
            lon_column,
            label_columns,
            "원본 좌표 확인",
        )

    def preview_coordinates_on_map(self):
        if self.result_df is None:
            QMessageBox.warning(self, "경고", "먼저 '결과 확인'으로 변환 결과를 생성하세요.")
            return

        final_df = self.get_result_df_from_table()
        if final_df is None or final_df.empty:
            QMessageBox.warning(self, "경고", "확인할 결과 데이터가 없습니다.")
            return

        required_columns = {"decimalLatitude", "decimalLongitude"}
        if not required_columns.issubset(final_df.columns):
            QMessageBox.warning(
                self,
                "좌표 컬럼 없음",
                "decimalLatitude와 decimalLongitude 컬럼이 결과에 있어야 좌표를 확인할 수 있습니다.",
            )
            return

        label_columns = [
            "scientificName",
            "vernacularName",
            "eventDate",
            "locality",
            "verbatimLocality",
            "catalogNumber",
            "occurrenceID",
        ]
        self._preview_coordinate_dataframe(
            final_df,
            "decimalLatitude",
            "decimalLongitude",
            label_columns,
            "좌표 확인",
        )

    def _preview_coordinate_dataframe(
        self,
        df: pd.DataFrame,
        lat_column: str,
        lon_column: str,
        label_columns: list[str],
        dialog_title: str,
    ):
        markers = []
        invalid_coordinates = []

        for idx, row in df.iterrows():
            row_number = int(idx) + 1
            raw_lat = row.get(lat_column)
            raw_lon = row.get(lon_column)
            lat = pd.to_numeric(raw_lat, errors="coerce")
            lon = pd.to_numeric(raw_lon, errors="coerce")
            invalid_reasons = []

            if pd.isna(raw_lat) or str(raw_lat).strip() == "":
                invalid_reasons.append("위도 없음")
            elif pd.isna(lat):
                invalid_reasons.append("위도 숫자 아님")
            elif not (-90 <= lat <= 90):
                invalid_reasons.append("위도 범위 초과")

            if pd.isna(raw_lon) or str(raw_lon).strip() == "":
                invalid_reasons.append("경도 없음")
            elif pd.isna(lon):
                invalid_reasons.append("경도 숫자 아님")
            elif not (-180 <= lon <= 180):
                invalid_reasons.append("경도 범위 초과")

            if invalid_reasons:
                invalid_coordinates.append(
                    {
                        "rowNumber": row_number,
                        "latitude": "" if pd.isna(raw_lat) else str(raw_lat).strip(),
                        "longitude": "" if pd.isna(raw_lon) else str(raw_lon).strip(),
                        "reason": ", ".join(invalid_reasons),
                    }
                )
                continue

            details = []
            for column in label_columns:
                if column not in df.columns:
                    continue
                value = row.get(column)
                if pd.isna(value) or str(value).strip() == "":
                    continue
                details.append({"label": column, "value": str(value).strip()})

            markers.append(
                {
                    "rowNumber": row_number,
                    "lat": float(lat),
                    "lon": float(lon),
                    "details": details,
                }
            )

        map_path = OUTPUT_DIR / "coordinate_preview_map.html"
        invalid_report_path = OUTPUT_DIR / "coordinate_invalid_rows.csv"

        if not markers:
            invalid_message = ""
            if invalid_coordinates:
                try:
                    pd.DataFrame(invalid_coordinates).to_csv(
                        invalid_report_path,
                        index=False,
                        encoding="utf-8-sig",
                    )
                    invalid_message = f"\n\n제외 좌표 목록: {invalid_report_path}"
                except Exception as e:
                    QMessageBox.critical(self, "오류", f"제외 좌표 목록 저장 중 오류가 발생했습니다.\n\n{e}")
                    return
            QMessageBox.warning(
                self,
                "좌표 없음",
                f"지도에 표시할 수 있는 {lat_column} / {lon_column} 값이 없습니다."
                f"{invalid_message}",
            )
            return

        try:
            if invalid_coordinates:
                pd.DataFrame(invalid_coordinates).to_csv(
                    invalid_report_path,
                    index=False,
                    encoding="utf-8-sig",
                )
            map_path.write_text(
                self._coordinate_preview_map_html(markers, invalid_coordinates),
                encoding="utf-8",
            )
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(map_path)))
            invalid_lines = ""
            if invalid_coordinates:
                preview_rows = ", ".join(
                    str(item["rowNumber"]) for item in invalid_coordinates[:20]
                )
                more_count = len(invalid_coordinates) - 20
                if more_count > 0:
                    preview_rows += f" 외 {more_count}개"
                invalid_lines = (
                    f"\n제외된 행: {preview_rows}"
                    f"\n제외 좌표 목록: {invalid_report_path}"
                )
            QMessageBox.information(
                self,
                dialog_title,
                f"좌표 지도 미리보기를 생성했습니다.\n\n{map_path}\n\n"
                f"표시된 좌표: {len(markers)}개\n"
                f"제외된 좌표: {len(invalid_coordinates)}개"
                f"{invalid_lines}",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"{dialog_title} 중 오류가 발생했습니다.\n\n{e}")

    @staticmethod
    def _coordinate_preview_map_html(markers: list[dict], invalid_coordinates: list[dict]) -> str:
        marker_json = json.dumps(markers, ensure_ascii=False)
        invalid_json = json.dumps(invalid_coordinates[:50], ensure_ascii=False)
        return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Coordinate Preview Map</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
  <style>
    html, body, #map {{
      height: 100%;
      margin: 0;
    }}
    body {{
      font-family: "Segoe UI", "Malgun Gothic", sans-serif;
    }}
    .summary {{
      background: #ffffff;
      border: 1px solid #d8e2ec;
      border-radius: 8px;
      box-shadow: 0 8px 24px rgba(15, 23, 42, 0.14);
      color: #1e293b;
      font-size: 13px;
      line-height: 1.45;
      padding: 10px 12px;
    }}
    .summary strong {{
      display: block;
      font-size: 14px;
      margin-bottom: 4px;
    }}
    .invalid-list {{
      border-top: 1px solid #e2e8f0;
      margin-top: 8px;
      max-height: 180px;
      overflow: auto;
      padding-top: 6px;
    }}
    .invalid-row {{
      margin: 3px 0;
      white-space: nowrap;
    }}
    .popup-title {{
      color: #0f766e;
      font-weight: 800;
      margin-bottom: 6px;
    }}
    .popup-row {{
      margin: 2px 0;
    }}
    .popup-row b {{
      color: #334155;
    }}
  </style>
</head>
<body>
  <div id="map"></div>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <script>
    const markers = {marker_json};
    const invalidCoordinates = {invalid_json};
    const invalidCount = {len(invalid_coordinates)};
    const map = L.map("map");

    L.tileLayer("https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png", {{
      maxZoom: 19,
      attribution: "&copy; OpenStreetMap contributors"
    }}).addTo(map);

    const bounds = [];
    const markerLayer = L.layerGroup().addTo(map);
    const escapeHtml = (value) => String(value)
      .replaceAll("&", "&amp;")
      .replaceAll("<", "&lt;")
      .replaceAll(">", "&gt;")
      .replaceAll('"', "&quot;")
      .replaceAll("'", "&#039;");

    markers.forEach((item) => {{
      const position = [item.lat, item.lon];
      bounds.push(position);
      const detailRows = item.details.map((detail) =>
        `<div class="popup-row"><b>${{escapeHtml(detail.label)}}:</b> ${{escapeHtml(detail.value)}}</div>`
      ).join("");
      const popup = `
        <div class="popup-title">Row ${{item.rowNumber}}</div>
        <div class="popup-row"><b>Latitude:</b> ${{item.lat}}</div>
        <div class="popup-row"><b>Longitude:</b> ${{item.lon}}</div>
        ${{detailRows}}
      `;
      L.circleMarker(position, {{
        radius: 7,
        color: "#0f766e",
        fillColor: "#14b8a6",
        fillOpacity: 0.72,
        weight: 2
      }}).bindPopup(popup).addTo(markerLayer);
    }});

    if (bounds.length === 1) {{
      map.setView(bounds[0], 12);
    }} else {{
      map.fitBounds(bounds, {{ padding: [36, 36], maxZoom: 14 }});
    }}

    const summary = L.control({{ position: "topright" }});
    summary.onAdd = function() {{
      const div = L.DomUtil.create("div", "summary");
      const invalidRows = invalidCoordinates.map((item) =>
        `<div class="invalid-row">Row ${{item.rowNumber}}: ${{escapeHtml(item.reason)}}</div>`
      ).join("");
      const moreInvalid = invalidCount > invalidCoordinates.length
        ? `<div class="invalid-row">...and ${{invalidCount - invalidCoordinates.length}} more</div>`
        : "";
      div.innerHTML = `
        <strong>Coordinate Preview</strong>
        Displayed: ${{markers.length}}<br>
        Invalid / blank: ${{invalidCount}}
        ${{invalidCount ? `<div class="invalid-list">${{invalidRows}}${{moreInvalid}}</div>` : ""}}
      `;
      return div;
    }};
    summary.addTo(map);
  </script>
</body>
</html>
"""

    def get_manual_values(self) -> dict:
        values = {}

        for key, widget in self.manual_inputs.items():
            if widget is None:
                continue

            try:
                text = widget.text().strip()
                if text:
                    values[key] = text
            except RuntimeError:
                continue

        return values

    def _get_missing_value_summary(self, df) -> tuple[int, list[str], list[tuple[int, int, str]]]:
        blank_mask = df.fillna("").astype(str).apply(lambda col: col.str.strip() == "")
        missing_counts = blank_mask.sum()
        missing_counts = missing_counts[missing_counts > 0].sort_values(ascending=False)

        if missing_counts.empty:
            return 0, [], []

        total_missing = int(missing_counts.sum())
        summary_lines = ["컬럼별 빈칸 수"]
        summary_lines.extend(
            f"{column}: {int(count)}개 비어 있음"
            for column, count in missing_counts.head(8).items()
        )

        if len(missing_counts) > 8:
            summary_lines.append(f"외 {len(missing_counts) - 8}개 컬럼")

        missing_locations = []
        preview_count = 0
        for row_idx in range(len(df)):
            for col_idx, column in enumerate(df.columns):
                value = df.iat[row_idx, col_idx]
                if pd.isna(value) or str(value).strip() == "":
                    missing_locations.append((row_idx, col_idx, str(column)))
                    preview_count += 1
                    if preview_count >= 12:
                        remaining = total_missing - preview_count
                        if remaining > 0:
                            summary_lines.append(f"외 {remaining}개 위치")
                        return total_missing, summary_lines, missing_locations

        return total_missing, summary_lines, missing_locations

    def _clear_missing_summary_content(self):
        while self.missing_summary_content_layout.count():
            item = self.missing_summary_content_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def reset_missing_summary_panel(self):
        self._clear_missing_summary_content()
        self.missing_summary_label = QLabel("결과 확인 후 빈칸 요약이 여기에 표시됩니다.")
        self.missing_summary_label.setWordWrap(True)
        self.missing_summary_label.setProperty("class", "sectionDescription")
        self.missing_summary_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.missing_summary_content_layout.addWidget(self.missing_summary_label)

    def _update_missing_summary_panel(
        self,
        total_missing: int,
        missing_summary: list[str],
        missing_locations: list[tuple[int, int, str]] | None = None,
    ):
        self._clear_missing_summary_content()

        self.missing_summary_label = QLabel()
        self.missing_summary_label.setWordWrap(True)
        self.missing_summary_label.setProperty("class", "sectionDescription")
        self.missing_summary_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.missing_summary_content_layout.addWidget(self.missing_summary_label)

        if total_missing > 0:
            self.missing_summary_label.setText(
                "총 빈칸 수: "
                f"{total_missing}개\n"
                + "\n".join(missing_summary)
            )

            if missing_locations:
                heading = QLabel("빈칸 위치 예시")
                heading.setProperty("class", "sectionDescription")
                self.missing_summary_content_layout.addWidget(heading)

                for row_idx, col_idx, column in missing_locations:
                    button = QPushButton(f"{column}: {row_idx + 1}행 {col_idx + 1}열")
                    button.setProperty("role", "link")
                    button.setCursor(Qt.PointingHandCursor)
                    button.clicked.connect(
                        lambda checked=False, r=row_idx, c=col_idx: self.focus_result_cell(r, c)
                    )
                    self.missing_summary_content_layout.addWidget(button)
        else:
            self.missing_summary_label.setText("빈칸 없이 모두 채워져 있습니다.")

        self.missing_summary_content_layout.addStretch()

    def focus_result_cell(self, row_idx: int, col_idx: int):
        if row_idx >= self.result_table.rowCount() or col_idx >= self.result_table.columnCount():
            return

        if hasattr(self, "main_tabs"):
            self.main_tabs.setCurrentIndex(0)

        self.result_table.setFocus()
        self.result_table.clearSelection()
        self.result_table.setCurrentCell(row_idx, col_idx)
        self.result_table.scrollToItem(
            self.result_table.item(row_idx, col_idx),
            QAbstractItemView.PositionAtCenter,
        )

    @staticmethod
    def validate_country_code(code: str) -> bool:
        return len(code) == 2 and code.isalpha()

    def preview_mapped_result(self):
        if self.source_df is None:
            QMessageBox.warning(self, "경고", "헤더를 먼저 적용해주세요.")
            return

        basis_combo = self.combo_boxes.get("basisOfRecord")
        if basis_combo is None or not basis_combo.currentText().strip():
            QMessageBox.warning(self, "경고", "basisOfRecord는 반드시 선택해야 합니다.")
            return

        mapping = self.get_selected_mapping()
        manual_values = self.get_manual_values()
        fixed_values = {"basisOfRecord": basis_combo.currentText().strip()}

        country_manual = manual_values.get("countryCode")
        if country_manual:
            fixed_values["countryCode"] = country_manual
        else:
            combo = self.combo_boxes.get("countryCode")
            if combo:
                try:
                    country_value = combo.currentText().strip()
                    if country_value:
                        fixed_values["countryCode"] = country_value
                except RuntimeError:
                    pass

        for key, value in manual_values.items():
            if key != "countryCode":
                fixed_values[key] = value

        try:
            self.result_df = MappingService.convert_dataframe(
                source_df=self.source_df,
                mapping=mapping,
                dwc_fields=self._get_active_fields(),
                fixed_values=fixed_values,
            )

            self.show_result_preview(self.result_df)
            self.btn_export.setEnabled(True)
            self.btn_coordinate_preview.setEnabled(True)
            total_missing, missing_summary, missing_locations = self._get_missing_value_summary(self.result_df)
            self._update_missing_summary_panel(total_missing, missing_summary, missing_locations)

            if total_missing > 0:
                QMessageBox.warning(
                    self,
                    "빈칸 확인 필요",
                    "결과를 아래 편집 테이블에 표시했습니다.\n"
                    "빈칸이 있는 데이터가 있어 저장 전에 확인하는 것을 권장합니다.\n\n"
                    f"총 빈칸 수: {total_missing}개\n"
                    + "\n".join(missing_summary),
                )
            else:
                QMessageBox.information(
                    self,
                    "완료",
                    "결과를 아래 편집 테이블에 표시했습니다.\n필요한 값은 직접 수정한 뒤 저장하세요.",
                )

        except Exception as e:
            QMessageBox.critical(self, "오류", f"결과 확인 중 오류가 발생했습니다.\n\n{e}")

    def show_result_preview(self, df):
        preview_df = df

        self.result_table.clear()
        self.result_table.setRowCount(len(preview_df))
        self.result_table.setColumnCount(len(preview_df.columns))
        self.result_table.setHorizontalHeaderLabels([str(col) for col in preview_df.columns.tolist()])

        for row_idx in range(len(preview_df)):
            for col_idx, _ in enumerate(preview_df.columns):
                value = preview_df.iloc[row_idx, col_idx]
                self.result_table.setItem(
                    row_idx,
                    col_idx,
                    QTableWidgetItem("" if value is None else str(value)),
                )

        self.result_table.resizeColumnsToContents()
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def get_result_df_from_table(self):
        if self.result_df is None:
            return None

        updated_df = self.result_df.copy()

        for row in range(self.result_table.rowCount()):
            for col in range(self.result_table.columnCount()):
                item = self.result_table.item(row, col)
                updated_df.iat[row, col] = item.text() if item else ""

        return updated_df
