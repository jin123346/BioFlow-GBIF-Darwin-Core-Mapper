from pathlib import Path
import json

import pandas as pd

from PySide6.QtCore import QSettings, Qt, QTimer, QUrl, Signal
from PySide6.QtGui import QColor, QDesktopServices, QPainter, QPen, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QCompleter,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QScrollArea,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.config.app_info import APP_NAME, AUTHOR_NAME
from app.config.country_code import GBIF_CONTINENT_OPTIONS, GBIF_COUNTRY_OPTIONS, ISO_COUNTRY_CODES
from app.services.climate_service import ClimateService
from app.services.gbif_service import GbifFetchCancelled, GbifOccurrenceCriteria, GbifService
from app.utils.paths import OUTPUT_DIR


GBIF_REGION_PRESETS = [
    ("서울", (126.76, 37.41, 127.18, 37.70), ["seoul", "서울", "서울시", "서울특별시"]),
    ("부산", (128.75, 34.98, 129.31, 35.40), ["busan", "부산", "부산시", "부산광역시"]),
    ("대구", (128.35, 35.73, 128.77, 36.02), ["daegu", "대구", "대구시", "대구광역시"]),
    ("인천", (124.60, 36.88, 126.80, 37.98), ["incheon", "인천", "인천시", "인천광역시"]),
    ("광주", (126.65, 35.05, 127.02, 35.28), ["gwangju", "광주", "광주시", "광주광역시"]),
    ("대전", (127.25, 36.18, 127.55, 36.50), ["daejeon", "대전", "대전시", "대전광역시"]),
    ("울산", (129.00, 35.32, 129.47, 35.72), ["ulsan", "울산", "울산시", "울산광역시"]),
    ("세종", (127.13, 36.40, 127.38, 36.73), ["sejong", "세종", "세종시", "세종특별자치시"]),
    ("경기", (126.35, 36.85, 127.85, 38.30), ["gyeonggi", "경기", "경기도"]),
    ("강원", (127.05, 37.02, 129.38, 38.62), ["gangwon", "강원", "강원도", "강원특별자치도"]),
    ("충북", (127.25, 36.00, 128.72, 37.35), ["chungbuk", "충북", "충청북도"]),
    ("충남", (125.95, 35.95, 127.65, 37.10), ["chungnam", "충남", "충청남도"]),
    ("전북", (126.38, 35.30, 127.90, 36.18), ["jeonbuk", "전북", "전라북도", "전북특별자치도"]),
    ("전남", (125.00, 33.90, 127.85, 35.55), ["jeonnam", "전남", "전라남도"]),
    ("경북", (128.00, 35.55, 131.90, 37.55), ["gyeongbuk", "경북", "경상북도"]),
    ("경남", (127.55, 34.55, 129.25, 35.95), ["gyeongnam", "경남", "경상남도"]),
    ("제주", (126.05, 33.06, 126.98, 33.58), ["jeju", "jejudo", "jejuisland", "제주", "제주도", "제주특별자치도"]),
    ("수원", (126.93, 37.20, 127.10, 37.35), ["suwon", "수원", "수원시"]),
    ("성남", (127.00, 37.32, 127.20, 37.50), ["seongnam", "성남", "성남시"]),
    ("고양", (126.68, 37.58, 126.95, 37.75), ["goyang", "고양", "고양시"]),
    ("용인", (127.05, 37.08, 127.40, 37.35), ["yongin", "용인", "용인시"]),
    ("화성", (126.55, 37.00, 127.15, 37.32), ["hwaseong", "화성", "화성시"]),
    ("청주", (127.25, 36.45, 127.75, 36.85), ["cheongju", "청주", "청주시"]),
    ("천안", (126.95, 36.70, 127.35, 37.05), ["cheonan", "천안", "천안시"]),
    ("전주", (127.00, 35.72, 127.25, 35.95), ["jeonju", "전주", "전주시"]),
    ("군산", (126.45, 35.80, 126.95, 36.15), ["gunsan", "군산", "군산시"]),
    ("목포", (126.25, 34.70, 126.55, 34.90), ["mokpo", "목포", "목포시"]),
    ("여수", (127.45, 34.50, 128.05, 35.05), ["yeosu", "여수", "여수시"]),
    ("순천", (127.25, 34.80, 127.65, 35.15), ["suncheon", "순천", "순천시"]),
    ("포항", (129.15, 35.85, 129.65, 36.35), ["pohang", "포항", "포항시"]),
    ("경주", (128.90, 35.60, 129.50, 36.05), ["gyeongju", "경주", "경주시"]),
    ("안동", (128.45, 36.35, 129.05, 36.85), ["andong", "안동", "안동시"]),
    ("창원", (128.35, 35.00, 128.90, 35.40), ["changwon", "창원", "창원시"]),
    ("김해", (128.65, 35.10, 129.05, 35.40), ["gimhae", "김해", "김해시"]),
    ("진주", (127.90, 35.05, 128.35, 35.35), ["jinju", "진주", "진주시"]),
    ("거제", (128.45, 34.65, 128.80, 35.05), ["geoje", "거제", "거제시"]),
    ("통영", (128.25, 34.70, 128.55, 34.95), ["tongyeong", "통영", "통영시"]),
    ("춘천", (127.55, 37.75, 128.05, 38.10), ["chuncheon", "춘천", "춘천시"]),
    ("원주", (127.75, 37.15, 128.25, 37.55), ["wonju", "원주", "원주시"]),
    ("강릉", (128.65, 37.55, 129.15, 37.95), ["gangneung", "강릉", "강릉시"]),
    ("서귀포", (126.15, 33.10, 126.95, 33.38), ["seogwipo", "서귀포", "서귀포시"]),
]


def _normalize_region_key(value: str) -> str:
    return value.casefold().replace(" ", "").replace("-", "").replace("_", "")


GBIF_REGION_BBOX_ALIASES = {
    _normalize_region_key(alias): bbox
    for _, bbox, aliases in GBIF_REGION_PRESETS
    for alias in aliases
}
GBIF_REGION_COMPLETER_ITEMS = sorted(
    {alias for label, _, aliases in GBIF_REGION_PRESETS for alias in [label, *aliases]},
    key=str.casefold,
)
GBIF_GEOMETRY_HELP_TEXT = (
    "한국 주요 지역명은 대략 bbox로 자동 변환합니다.\n"
    "예: 서울, 부산, 제주, seoul, busan, jeju\n\n"
    "그 외 지역은 bbox 또는 WKT POLYGON을 직접 입력하세요.\n"
    "bbox 형식: W,S,E,N = 서쪽 경도, 남쪽 위도, 동쪽 경도, 북쪽 위도\n"
    "예: 126.05,33.06,126.98,33.58"
)


class CheckableComboBox(QComboBox):
    selectionChanged = Signal()

    def __init__(self, placeholder: str, parent: QWidget | None = None):
        super().__init__(parent)
        self.placeholder = placeholder
        self.setModel(QStandardItemModel(self))
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setPlaceholderText(placeholder)
        self.view().pressed.connect(self._toggle_item)
        self.setMinimumWidth(190)
        self._update_text()

    def add_check_item(self, label: str, value: str) -> None:
        item = QStandardItem(label)
        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
        item.setData(Qt.Unchecked, Qt.CheckStateRole)
        item.setData(value, Qt.UserRole)
        self.model().appendRow(item)
        self._update_text()

    def add_check_items(self, items: list[tuple[str, str]]) -> None:
        for label, value in items:
            self.add_check_item(label, value)

    def selected_values(self) -> list[str]:
        values = []
        for row in range(self.model().rowCount()):
            item = self.model().item(row)
            if item.checkState() == Qt.Checked:
                values.append(str(item.data(Qt.UserRole)))
        return values

    def clear_selection(self) -> None:
        changed = False
        for row in range(self.model().rowCount()):
            item = self.model().item(row)
            if item.checkState() == Qt.Checked:
                item.setCheckState(Qt.Unchecked)
                changed = True
        self._update_text()
        if changed:
            self.selectionChanged.emit()

    def clear_items(self) -> None:
        self.model().clear()
        self._update_text()

    def _toggle_item(self, index) -> None:
        item = self.model().itemFromIndex(index)
        if item is None:
            return
        item.setCheckState(Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked)
        self._update_text()
        self.selectionChanged.emit()

    def _update_text(self) -> None:
        labels = []
        for row in range(self.model().rowCount()):
            item = self.model().item(row)
            if item.checkState() == Qt.Checked:
                labels.append(item.text())
        if not labels:
            self.lineEdit().setText("")
            return
        text = labels[0] if len(labels) == 1 else f"{labels[0]} +{len(labels) - 1}"
        self.lineEdit().setText(text)


class GbifLiveChartWidget(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.chart_list: list[dict] = []
        self.chart_type = "bar"
        self.message = "GBIF 데이터를 가져오면 그래프가 표시됩니다."
        self.colors = [
            QColor("#0f766e"),
            QColor("#2563eb"),
            QColor("#b45309"),
            QColor("#7c3aed"),
            QColor("#be123c"),
            QColor("#0891b2"),
        ]
        self.setMinimumHeight(260)

    def clear_chart(self, message: str | None = None) -> None:
        self.chart_list = []
        self.setMinimumHeight(260)
        self.message = message or "GBIF 데이터를 가져오면 그래프가 표시됩니다."
        self.update()

    def update_chart(self, chart_list: list[dict], chart_type: str) -> None:
        self.chart_list = chart_list
        self.chart_type = chart_type if chart_type in {"bar", "line"} else "bar"
        self.setMinimumHeight(260 + max(0, min(len(chart_list), 3) - 1) * 180)
        self.message = "현재 조건에 맞는 그래프 데이터가 없습니다."
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = self.rect().adjusted(12, 12, -12, -12)
        painter.fillRect(self.rect(), QColor("#ffffff"))

        if not self.chart_list:
            painter.setPen(QColor("#64748b"))
            painter.drawText(rect, Qt.AlignCenter, self.message)
            return

        visible_charts = self.chart_list[:3]
        chart_height = max(120, rect.height() // max(1, len(visible_charts)))
        for index, chart in enumerate(visible_charts):
            top = rect.top() + index * chart_height
            bottom = rect.bottom() if index == len(visible_charts) - 1 else top + chart_height - 8
            chart_rect = rect.adjusted(0, top - rect.top(), 0, bottom - rect.bottom())
            self._draw_chart(painter, chart_rect, chart)

    def _draw_chart(self, painter, rect, chart):
        labels = [str(label) for label in chart.get("labels", [])]
        series = chart.get("series", [])
        if not labels or not series:
            painter.setPen(QColor("#64748b"))
            painter.drawText(rect, Qt.AlignCenter, self.message)
            return

        title_rect = rect.adjusted(0, 0, 0, -rect.height() + 26)
        painter.setPen(QColor("#0f172a"))
        title_font = painter.font()
        title_font.setBold(True)
        title_font.setPointSize(max(title_font.pointSize(), 10))
        painter.setFont(title_font)
        painter.drawText(title_rect, Qt.AlignLeft | Qt.AlignVCenter, str(chart.get("title", "Live chart")))

        font = painter.font()
        font.setBold(False)
        font.setPointSize(9)
        painter.setFont(font)

        legend_height = self._draw_legend(painter, rect.adjusted(0, 28, 0, 0), series)
        plot_rect = rect.adjusted(6, 38 + legend_height, -8, -24)
        painter.setPen(QPen(QColor("#cbd5e1"), 1))
        painter.drawLine(plot_rect.bottomLeft(), plot_rect.bottomRight())
        painter.drawLine(plot_rect.bottomLeft(), plot_rect.topLeft())

        values_by_series = [
            [self._chart_number(value) for value in item.get("values", [])[: len(labels)]]
            for item in series
        ]
        flat_values = [value for values in values_by_series for value in values]
        min_value = min([0, *flat_values])
        max_value = max([1, *flat_values])

        if self.chart_type == "line":
            self._draw_line_chart(painter, plot_rect, labels, series, values_by_series, min_value, max_value)
        else:
            self._draw_bar_chart(painter, plot_rect, labels, series, values_by_series, min_value, max_value)

    def _draw_legend(self, painter, rect, series) -> int:
        if len(series) <= 1:
            return 0
        x = rect.left()
        y = rect.top()
        row_height = 20
        max_width = rect.width()
        painter.setPen(QColor("#475569"))
        for index, item in enumerate(series):
            name = str(item.get("name", ""))[:42]
            text_width = painter.fontMetrics().horizontalAdvance(name)
            item_width = min(max(90, text_width + 24), 230)
            if x + item_width > rect.left() + max_width:
                x = rect.left()
                y += row_height
            color = self.colors[index % len(self.colors)]
            painter.fillRect(x, y + 5, 10, 10, color)
            painter.drawText(x + 16, y, item_width - 18, row_height, Qt.AlignVCenter | Qt.AlignLeft, name)
            x += item_width + 10
        return y - rect.top() + row_height + 4

    def _draw_bar_chart(self, painter, plot_rect, labels, series, values_by_series, min_value, max_value):
        group_count = len(labels)
        series_count = max(1, len(series))
        group_width = plot_rect.width() / max(1, group_count)
        bar_width = max(3, min(28, (group_width - 8) / series_count))
        value_range = max(max_value - min_value, 1)
        baseline = plot_rect.bottom() - int(((0 - min_value) / value_range) * max(1, plot_rect.height() - 20))
        painter.setPen(QPen(QColor("#e2e8f0"), 1))
        painter.drawLine(plot_rect.left(), baseline, plot_rect.right(), baseline)

        for label_index, label in enumerate(labels):
            group_left = plot_rect.left() + label_index * group_width
            for series_index, values in enumerate(values_by_series):
                value = values[label_index] if label_index < len(values) else 0
                value_y = plot_rect.bottom() - int(((value - min_value) / value_range) * max(1, plot_rect.height() - 20))
                height = abs(value_y - baseline)
                left = int(group_left + 4 + series_index * bar_width)
                top = min(value_y, baseline)
                painter.fillRect(left, top, int(bar_width - 2), height, self.colors[series_index % len(self.colors)])
                if group_count <= 14 and value:
                    painter.setPen(QColor("#334155"))
                    painter.drawText(left - 8, top - 16, int(bar_width + 24), 14, Qt.AlignCenter, self._format_chart_value(value))

            if group_count <= 18 or label_index % max(1, group_count // 12) == 0:
                painter.setPen(QColor("#475569"))
                painter.drawText(int(group_left), baseline + 4, int(group_width), 18, Qt.AlignCenter, label)

    def _draw_line_chart(self, painter, plot_rect, labels, series, values_by_series, min_value, max_value):
        point_count = len(labels)
        if point_count == 1:
            step = 0
        else:
            step = plot_rect.width() / (point_count - 1)
        value_range = max(max_value - min_value, 1)

        for series_index, values in enumerate(values_by_series):
            color = self.colors[series_index % len(self.colors)]
            painter.setPen(QPen(color, 2))
            previous = None
            for value_index, value in enumerate(values):
                x = plot_rect.left() + value_index * step
                y = plot_rect.bottom() - ((value - min_value) / value_range) * max(1, plot_rect.height() - 20)
                if previous is not None:
                    painter.drawLine(int(previous[0]), int(previous[1]), int(x), int(y))
                painter.setBrush(QColor("#ffffff"))
                painter.drawEllipse(int(x) - 3, int(y) - 3, 6, 6)
                previous = (x, y)

        painter.setPen(QColor("#475569"))
        for label_index, label in enumerate(labels):
            if point_count <= 18 or label_index % max(1, point_count // 12) == 0:
                x = plot_rect.left() + label_index * step
                painter.drawText(int(x) - 25, plot_rect.bottom() + 4, 50, 18, Qt.AlignCenter, label)

    @staticmethod
    def _format_chart_value(value: float) -> str:
        return f"{int(value):,}" if float(value).is_integer() else f"{value:,.1f}"

    @staticmethod
    def _chart_number(value) -> float:
        if value is None:
            return 0.0
        try:
            if pd.isna(value):
                return 0.0
        except TypeError:
            pass
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0


class GbifAnalysisTab(QWidget):
    def __init__(self, settings: QSettings | None = None, parent: QWidget | None = None):
        super().__init__(parent)
        self.settings = settings or QSettings(AUTHOR_NAME, APP_NAME)
        self.gbif_df = None
        self.gbif_raw_df = None
        self.gbif_filtered_df = None
        self.gbif_climate_df = None
        self.gbif_climate_location = None
        self.gbif_table_preview_df = None
        self.gbif_preview_loaded_rows = 0
        self.gbif_preview_chunk_size = 1000
        self._loading_gbif_preview_rows = False
        self.gbif_search_summary = None
        self.gbif_report_path = None
        self._building_gbif_basis_filter = False
        self._syncing_gbif_quick_month = False

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setWidget(self._build_gbif_analysis_tab())
        root_layout.addWidget(scroll_area)

        self.gbif_cleaning_timer = QTimer(self)
        self.gbif_cleaning_timer.setSingleShot(True)
        self.gbif_cleaning_timer.setInterval(450)
        self.gbif_cleaning_timer.timeout.connect(self.apply_gbif_cleaning_silently)

    def _build_section_header(self, title: str, description: str) -> QFrame:
        frame = QFrame()
        frame.setProperty("class", "sectionCard")

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(4)

        title_label = QLabel(title)
        title_label.setProperty("class", "sectionTitle")

        desc_label = QLabel(description)
        desc_label.setProperty("class", "sectionDescription")
        desc_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(desc_label)
        return frame

    def _configure_table(self, table: QTableWidget):
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectItems)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setWordWrap(False)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.horizontalHeader().setMinimumSectionSize(110)
        table.setShowGrid(False)

    def _get_last_save_dir(self) -> str:
        last_dir = self.settings.value("file_dialog/last_save_dir", "", str)
        if last_dir and Path(last_dir).is_dir():
            return last_dir
        return str(OUTPUT_DIR)

    def _remember_save_dir(self, file_path: str) -> None:
        parent_dir = Path(file_path).expanduser().parent
        if parent_dir.is_dir():
            self.settings.setValue("file_dialog/last_save_dir", str(parent_dir))

    @staticmethod
    def validate_country_code(code: str) -> bool:
        return code in ISO_COUNTRY_CODES

    def _build_gbif_analysis_tab(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        layout.addWidget(
            self._build_section_header(
                "GBIF occurrence 데이터 분석",
                "분류군, 데이터셋, 국가코드, geometry 조건을 조합해 GBIF 좌표 기록을 가져오고 분석합니다.",
            )
        )

        search_card = QGroupBox("1. 검색 조건 조합")
        search_card.setProperty("class", "sectionCard")
        search_layout = QVBoxLayout(search_card)
        search_layout.setContentsMargins(16, 14, 16, 14)
        search_layout.setSpacing(10)

        input_layout = QHBoxLayout()
        input_layout.setSpacing(8)

        self.gbif_species_input = QLineEdit()
        self.gbif_species_input.setPlaceholderText("분류군 조건: 학명/상위분류군 예 Quercus mongolica, Insecta, Aves")

        self.gbif_taxon_key_input = QLineEdit()
        self.gbif_taxon_key_input.setPlaceholderText("taxonKey")
        self.gbif_taxon_key_input.setFixedWidth(96)

        self.gbif_dataset_key_input = QLineEdit()
        self.gbif_dataset_key_input.setPlaceholderText("데이터셋 조건: datasetKey")
        self.gbif_dataset_key_input.setFixedWidth(210)

        self.gbif_country_combo = CheckableComboBox("Country")
        self.gbif_country_combo.add_check_items(
            [(f"{name} ({code})", code) for code, name in GBIF_COUNTRY_OPTIONS]
        )
        self.gbif_country_combo.setFixedWidth(230)

        self.gbif_continent_combo = CheckableComboBox("Continent")
        self.gbif_continent_combo.add_check_items(
            [(name, code) for code, name in GBIF_CONTINENT_OPTIONS]
        )
        self.gbif_continent_combo.setFixedWidth(170)
        self.gbif_country_combo.setMaxVisibleItems(18)
        self.gbif_country_combo.setFixedWidth(230)

        self.gbif_filter_country_combo = CheckableComboBox("Filter Country")
        self.gbif_filter_country_combo.add_check_items(
            [(f"{name} ({code})", code) for code, name in GBIF_COUNTRY_OPTIONS]
        )
        self.gbif_filter_country_combo.setFixedWidth(230)
        self.gbif_filter_country_combo.setMaxVisibleItems(18)

        self.gbif_filter_continent_combo = CheckableComboBox("Filter Continent")
        self.gbif_filter_continent_combo.add_check_items(
            [(name, code) for code, name in GBIF_CONTINENT_OPTIONS]
        )
        self.gbif_filter_continent_combo.setFixedWidth(170)

        self.gbif_filter_species_combo = CheckableComboBox("Filter Species")
        self.gbif_filter_species_combo.setFixedWidth(280)
        self.gbif_filter_species_combo.setMaxVisibleItems(18)

        self.gbif_geometry_input = QLineEdit()
        self.gbif_geometry_input.setPlaceholderText("지역 조건: 서울, 부산, jeju 또는 bbox W,S,E,N")
        self.gbif_geometry_input.setToolTip(GBIF_GEOMETRY_HELP_TEXT)
        region_completer = QCompleter(GBIF_REGION_COMPLETER_ITEMS, self.gbif_geometry_input)
        region_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.gbif_geometry_input.setCompleter(region_completer)

        geometry_help_icon = QLabel("!")
        geometry_help_icon.setAlignment(Qt.AlignCenter)
        geometry_help_icon.setFixedSize(20, 20)
        geometry_help_icon.setToolTip(GBIF_GEOMETRY_HELP_TEXT)
        geometry_help_icon.setStyleSheet(
            "QLabel {"
            "background: #fff7ed;"
            "border: 1px solid #f59e0b;"
            "border-radius: 10px;"
            "color: #92400e;"
            "font-weight: 800;"
            "}"
        )

        self.gbif_use_cache_check = QCheckBox("캐시 사용")
        self.gbif_use_cache_check.setChecked(True)

        self.gbif_chart_type_combo = QComboBox()
        self.gbif_chart_type_combo.addItem("막대 그래프", "bar")
        self.gbif_chart_type_combo.addItem("선 그래프", "line")
        self.gbif_chart_type_combo.setFixedWidth(120)
        self.gbif_chart_type_combo.currentIndexChanged.connect(self.update_gbif_live_chart)

        self.gbif_fill_missing_periods_check = QCheckBox("빈 기간 0 채움")
        self.gbif_fill_missing_periods_check.setChecked(True)
        self.gbif_fill_missing_periods_check.stateChanged.connect(self.update_gbif_live_chart)

        self.gbif_cumulative_check = QCheckBox("누적")
        self.gbif_cumulative_check.stateChanged.connect(self.update_gbif_live_chart)

        self.gbif_period_combo = QComboBox()
        self.gbif_period_combo.addItem("종합", "summary")
        self.gbif_period_combo.addItem("연도별", "year")
        self.gbif_period_combo.addItem("월별", "month")
        self.gbif_period_combo.addItem("계절별", "season")
        self.gbif_period_combo.addItem("종별 기록 수", "species")
        self.gbif_period_combo.addItem("데이터셋별 기록 수", "dataset")
        self.gbif_period_combo.addItem("자료유형별 기록 수", "basis")
        self.gbif_period_combo.setFixedWidth(130)
        self.gbif_period_combo.currentIndexChanged.connect(self.update_gbif_period_controls)

        self.gbif_quick_month_combo = QComboBox()
        self.gbif_quick_month_combo.addItem("전체 월", 0)
        for month in range(1, 13):
            self.gbif_quick_month_combo.addItem(f"{month}월", month)
        self.gbif_quick_month_combo.setFixedWidth(100)
        self.gbif_quick_month_combo.currentIndexChanged.connect(self.apply_gbif_quick_month_filter)

        self.gbif_year_from_input = QSpinBox()
        self.gbif_year_from_input.setRange(0, 9999)
        self.gbif_year_from_input.setSpecialValueText("전체")
        self.gbif_year_from_input.setFixedWidth(86)

        self.gbif_year_to_input = QSpinBox()
        self.gbif_year_to_input.setRange(0, 9999)
        self.gbif_year_to_input.setSpecialValueText("전체")
        self.gbif_year_to_input.setFixedWidth(86)

        self.gbif_month_from_input = QSpinBox()
        self.gbif_month_from_input.setRange(1, 12)
        self.gbif_month_from_input.setValue(1)
        self.gbif_month_from_input.setSuffix("월")
        self.gbif_month_from_input.setFixedWidth(70)

        self.gbif_month_to_input = QSpinBox()
        self.gbif_month_to_input.setRange(1, 12)
        self.gbif_month_to_input.setValue(12)
        self.gbif_month_to_input.setSuffix("월")
        self.gbif_month_to_input.setFixedWidth(70)

        self.btn_fetch_gbif = QPushButton("GBIF 가져오기")
        self.btn_fetch_gbif.setProperty("role", "primary")
        self.btn_fetch_gbif.clicked.connect(self.fetch_gbif_analysis)

        self.btn_refresh_gbif = QPushButton("캐시 새로고침")
        self.btn_refresh_gbif.setProperty("role", "secondary")
        self.btn_refresh_gbif.clicked.connect(self.refresh_gbif_analysis)

        self.btn_draw_gbif_graph = QPushButton("그래프 그리기")
        self.btn_draw_gbif_graph.setProperty("role", "primary")
        self.btn_draw_gbif_graph.clicked.connect(self.draw_gbif_graph_report)
        self.btn_draw_gbif_graph.setEnabled(False)

        self.btn_open_gbif_report = QPushButton("지도/그래프 열기")
        self.btn_open_gbif_report.setProperty("role", "secondary")
        self.btn_open_gbif_report.clicked.connect(self.open_gbif_analysis_report)
        self.btn_open_gbif_report.setEnabled(False)

        self.btn_export_gbif_csv = QPushButton("CSV 저장")
        self.btn_export_gbif_csv.setProperty("role", "secondary")
        self.btn_export_gbif_csv.clicked.connect(self.export_gbif_analysis_csv)
        self.btn_export_gbif_csv.setEnabled(False)

        self.btn_export_gbif_excel = QPushButton("엑셀 저장")
        self.btn_export_gbif_excel.setProperty("role", "secondary")
        self.btn_export_gbif_excel.clicked.connect(self.export_gbif_analysis_excel)
        self.btn_export_gbif_excel.setEnabled(False)

        self.btn_export_gbif_timeseries_csv = QPushButton("Tableau 데이터")
        self.btn_export_gbif_timeseries_csv.setProperty("role", "secondary")
        self.btn_export_gbif_timeseries_csv.clicked.connect(self.export_gbif_timeseries_csv)
        self.btn_export_gbif_timeseries_csv.setEnabled(False)

        self.btn_fetch_gbif_climate = QPushButton("기후 데이터")
        self.btn_fetch_gbif_climate.setProperty("role", "secondary")
        self.btn_fetch_gbif_climate.clicked.connect(self.fetch_gbif_climate_data)
        self.btn_fetch_gbif_climate.setEnabled(False)

        self.btn_export_gbif_geojson = QPushButton("QGIS 저장")
        self.btn_export_gbif_geojson.setProperty("role", "secondary")
        self.btn_export_gbif_geojson.clicked.connect(self.export_gbif_analysis_geojson)
        self.btn_export_gbif_geojson.setEnabled(False)

        input_layout.addWidget(self.gbif_species_input, 1)
        input_layout.addWidget(self.gbif_taxon_key_input)
        input_layout.addWidget(self.gbif_dataset_key_input)
        input_layout.addWidget(self.gbif_country_combo)
        input_layout.addWidget(self.gbif_continent_combo)
        input_layout.addWidget(self.gbif_geometry_input, 1)
        input_layout.addWidget(geometry_help_icon)
        input_layout.addWidget(self.gbif_use_cache_check)
        input_layout.addWidget(self.btn_fetch_gbif)
        input_layout.addWidget(self.btn_refresh_gbif)
        search_layout.addLayout(input_layout)

        geometry_help_label = QLabel(
            "지역 조건: 한국 주요 지역명은 자동 bbox로 검색합니다. "
            "지원하지 않는 지역은 W,S,E,N 예: 126.05,33.06,126.98,33.58 또는 WKT POLYGON을 입력하세요."
        )
        geometry_help_label.setProperty("class", "sectionDescription")
        geometry_help_label.setWordWrap(True)
        search_layout.addWidget(geometry_help_label)

        analysis_layout = QHBoxLayout()
        analysis_layout.setSpacing(8)
        analysis_layout.addWidget(QLabel("그래프 기준"))
        analysis_layout.addWidget(self.gbif_period_combo)
        analysis_layout.addWidget(self.gbif_chart_type_combo)
        analysis_layout.addWidget(self.gbif_fill_missing_periods_check)
        analysis_layout.addWidget(self.gbif_cumulative_check)
        analysis_layout.addWidget(QLabel("월 보기"))
        analysis_layout.addWidget(self.gbif_quick_month_combo)
        analysis_layout.addWidget(self.btn_draw_gbif_graph)
        analysis_layout.addWidget(self.btn_open_gbif_report)
        analysis_layout.addWidget(self.btn_export_gbif_csv)
        analysis_layout.addWidget(self.btn_export_gbif_excel)
        analysis_layout.addWidget(self.btn_export_gbif_timeseries_csv)
        analysis_layout.addWidget(self.btn_fetch_gbif_climate)
        analysis_layout.addWidget(self.btn_export_gbif_geojson)
        analysis_layout.addStretch()
        search_layout.addLayout(analysis_layout)

        cleaning_group = QGroupBox("2. 데이터 정제 조건")
        cleaning_group.setProperty("class", "sectionCard")
        cleaning_layout = QVBoxLayout(cleaning_group)
        cleaning_layout.setContentsMargins(16, 14, 16, 14)
        cleaning_layout.setSpacing(10)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(8)
        filter_layout.addWidget(QLabel("연도 범위"))
        filter_layout.addWidget(self.gbif_year_from_input)
        filter_layout.addWidget(QLabel("~"))
        filter_layout.addWidget(self.gbif_year_to_input)
        filter_layout.addSpacing(14)
        filter_layout.addWidget(QLabel("월 범위"))
        filter_layout.addWidget(self.gbif_month_from_input)
        filter_layout.addWidget(QLabel("~"))
        filter_layout.addWidget(self.gbif_month_to_input)
        filter_layout.addSpacing(14)
        filter_layout.addWidget(QLabel("Country"))
        filter_layout.addWidget(self.gbif_filter_country_combo)
        filter_layout.addWidget(QLabel("Continent"))
        filter_layout.addWidget(self.gbif_filter_continent_combo)
        filter_layout.addWidget(QLabel("Species"))
        filter_layout.addWidget(self.gbif_filter_species_combo)
        filter_layout.addStretch()
        cleaning_layout.addLayout(filter_layout)

        clean_layout = QHBoxLayout()
        clean_layout.setSpacing(8)

        self.gbif_basis_container = QWidget()
        self.gbif_basis_layout = QVBoxLayout(self.gbif_basis_container)
        self.gbif_basis_layout.setContentsMargins(0, 0, 0, 0)
        self.gbif_basis_layout.setSpacing(4)
        self.gbif_basis_checkboxes = {}

        self.gbif_remove_missing_year_check = QCheckBox("연도 없음 제거")
        self.gbif_remove_missing_year_check.setChecked(True)

        self.gbif_remove_missing_month_check = QCheckBox("월 없음 제거")

        self.gbif_remove_missing_name_check = QCheckBox("학명 없음 제거")
        self.gbif_remove_missing_name_check.setChecked(True)

        self.gbif_dedupe_gbif_id_check = QCheckBox("중복 GBIF ID 제거")
        self.gbif_dedupe_gbif_id_check.setChecked(True)

        self.btn_apply_gbif_cleaning = QPushButton("정제하기")
        self.btn_apply_gbif_cleaning.setProperty("role", "secondary")
        self.btn_apply_gbif_cleaning.clicked.connect(self.apply_gbif_cleaning)
        self.btn_apply_gbif_cleaning.setEnabled(False)

        clean_layout.addWidget(QLabel("제거 조건"))
        clean_layout.addWidget(self.gbif_remove_missing_year_check)
        clean_layout.addWidget(self.gbif_remove_missing_month_check)
        clean_layout.addWidget(self.gbif_remove_missing_name_check)
        clean_layout.addWidget(self.gbif_dedupe_gbif_id_check)
        clean_layout.addWidget(self.btn_apply_gbif_cleaning)
        clean_layout.addStretch()
        cleaning_layout.addLayout(clean_layout)

        cleaning_layout.addWidget(QLabel("자료유형 basisOfRecord"))
        cleaning_layout.addWidget(self.gbif_basis_container)

        exclude_layout = QHBoxLayout()
        exclude_layout.setSpacing(8)

        self.gbif_exclude_dataset_input = QLineEdit()
        self.gbif_exclude_dataset_input.setPlaceholderText("제외 datasetName 키워드, 쉼표 구분")

        self.gbif_exclude_institution_input = QLineEdit()
        self.gbif_exclude_institution_input.setPlaceholderText("제외 institutionCode 키워드, 쉼표 구분")

        exclude_layout.addWidget(QLabel("제외"))
        exclude_layout.addWidget(self.gbif_exclude_dataset_input, 1)
        exclude_layout.addWidget(self.gbif_exclude_institution_input, 1)
        cleaning_layout.addLayout(exclude_layout)

        self.gbif_filtered_count_label = QLabel("정제 조건에 해당하는 데이터: 0건")
        self.gbif_filtered_count_label.setObjectName("metaValue")
        cleaning_layout.addWidget(self.gbif_filtered_count_label)

        search_layout.addWidget(cleaning_group)

        download_layout = QHBoxLayout()
        download_layout.setSpacing(8)

        self.gbif_username_input = QLineEdit()
        self.gbif_username_input.setPlaceholderText("GBIF username")
        self.gbif_username_input.setText(self.settings.value("gbif/username", "", str))

        self.gbif_email_input = QLineEdit()
        self.gbif_email_input.setPlaceholderText("GBIF email")
        self.gbif_email_input.setText(self.settings.value("gbif/email", "", str))

        self.gbif_password_input = QLineEdit()
        self.gbif_password_input.setPlaceholderText("GBIF password")
        self.gbif_password_input.setEchoMode(QLineEdit.Password)

        self.btn_request_gbif_download = QPushButton("DOI 다운로드 요청")
        self.btn_request_gbif_download.setProperty("role", "success")
        self.btn_request_gbif_download.clicked.connect(self.request_gbif_download)

        download_layout.addWidget(self.gbif_username_input)
        download_layout.addWidget(self.gbif_email_input)
        download_layout.addWidget(self.gbif_password_input)
        download_layout.addWidget(self.btn_request_gbif_download)
        search_layout.addLayout(download_layout)

        self.gbif_summary_label = QLabel("검색 결과가 아직 없습니다.")
        self.gbif_summary_label.setObjectName("metaValue")
        self.gbif_summary_label.setWordWrap(True)
        search_layout.addWidget(self.gbif_summary_label)
        self.gbif_climate_label = QLabel("기후 데이터가 아직 연결되지 않았습니다.")
        self.gbif_climate_label.setObjectName("metaValue")
        self.gbif_climate_label.setWordWrap(True)
        search_layout.addWidget(self.gbif_climate_label)
        self.gbif_live_chart = GbifLiveChartWidget()
        search_layout.addWidget(self.gbif_live_chart)
        layout.addWidget(search_card)
        self._connect_gbif_cleaning_signals()

        self.gbif_result_table = QTableWidget()
        self.gbif_result_table.setMinimumHeight(320)
        self._configure_table(self.gbif_result_table)
        self.gbif_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.gbif_result_table.verticalHeader().setVisible(True)
        self.gbif_result_table.verticalHeader().setDefaultSectionSize(28)
        self.gbif_result_table.verticalScrollBar().valueChanged.connect(self.maybe_append_gbif_preview_rows)
        layout.addWidget(self.gbif_result_table)

        return panel

    def fetch_gbif_analysis(self):
        self._fetch_gbif_analysis(use_cache=self.gbif_use_cache_check.isChecked())

    def refresh_gbif_analysis(self):
        self._fetch_gbif_analysis(use_cache=False)

    def current_gbif_search_criteria(self) -> GbifOccurrenceCriteria:
        taxon_key_text = self.gbif_taxon_key_input.text().strip()
        taxon_key = None
        if taxon_key_text:
            try:
                taxon_key = int(taxon_key_text)
            except ValueError as e:
                raise ValueError("taxonKey는 숫자로 입력하세요.") from e

        scientific_name = self.gbif_species_input.text().strip()
        dataset_key = self.gbif_dataset_key_input.text().strip()
        geometry = self._normalize_gbif_geometry(self.gbif_geometry_input.text().strip())
        country_codes = self.gbif_country_combo.selected_values()
        continent_codes = self.gbif_continent_combo.selected_values()

        invalid_country_codes = [code for code in country_codes if not self.validate_country_code(code)]
        if invalid_country_codes:
            raise ValueError("국가코드는 KR, JP, US처럼 2자리 영문으로 입력하세요.")
        if taxon_key is None and not scientific_name and not dataset_key and not country_codes and not continent_codes and not geometry:
            raise ValueError("분류군, datasetKey, 국가코드, geometry 중 하나 이상을 입력하세요.")

        return GbifOccurrenceCriteria(
            scientific_name=scientific_name,
            taxon_key=taxon_key,
            dataset_key=dataset_key,
            country_codes=country_codes,
            continent_codes=continent_codes,
            geometry=geometry,
            limit=100000,
        )

    @staticmethod
    def _normalize_gbif_geometry(value: str) -> str:
        if not value:
            return ""

        alias_key = _normalize_region_key(value)
        if alias_key in GBIF_REGION_BBOX_ALIASES:
            return GbifAnalysisTab._bbox_to_polygon(GBIF_REGION_BBOX_ALIASES[alias_key])

        upper_value = value.upper()
        if upper_value.startswith(("POLYGON", "MULTIPOLYGON")):
            if "((" not in value or "))" not in value:
                raise ValueError("geometry WKT는 POLYGON((...)) 또는 MULTIPOLYGON((...)) 형태로 입력하세요.")
            return value

        parts = [part.strip() for part in value.split(",")]
        if len(parts) != 4:
            raise ValueError(
                "지역 조건은 한국 주요 지역명(예: 서울, 부산, 제주, seoul, busan, jeju), "
                "bbox(W,S,E,N), 또는 WKT POLYGON으로 입력하세요. "
                "지원하지 않는 지명은 직접 bbox를 입력해야 합니다."
            )
        try:
            west, south, east, north = [float(part) for part in parts]
        except ValueError:
            raise ValueError("bbox는 W,S,E,N 순서의 숫자 4개로 입력하세요. 예: 126.05,33.06,126.98,33.58")

        if not (-180 <= west <= 180 and -180 <= east <= 180 and -90 <= south <= 90 and -90 <= north <= 90):
            raise ValueError("bbox 좌표 범위를 확인하세요. 경도는 -180~180, 위도는 -90~90이어야 합니다.")
        if west >= east or south >= north:
            raise ValueError("bbox는 W,S,E,N 순서여야 하며 west < east, south < north여야 합니다.")

        return GbifAnalysisTab._bbox_to_polygon((west, south, east, north))

    @staticmethod
    def _bbox_to_polygon(bbox: tuple[float, float, float, float]) -> str:
        west, south, east, north = bbox
        return (
            "POLYGON(("
            f"{west} {south}, {east} {south}, {east} {north}, {west} {north}, {west} {south}"
            "))"
        )

    @staticmethod
    def _format_selected_codes(values: list[str]) -> str:
        return ", ".join(values) if values else "ALL"

    @staticmethod
    def _refined_gbif_codes(search_codes: list[str], filter_codes: list[str]) -> list[str]:
        if not filter_codes:
            return search_codes
        if not search_codes:
            return filter_codes
        search_set = {code.upper() for code in search_codes}
        return [code for code in filter_codes if code.upper() in search_set]

    def _fetch_gbif_analysis(self, use_cache: bool):
        try:
            criteria = self.current_gbif_search_criteria()
        except ValueError as e:
            QMessageBox.warning(self, "GBIF 분석", str(e))
            return

        self.reset_gbif_analysis_state(clear_raw=True)
        self.reset_gbif_cleaning_controls()
        self.btn_fetch_gbif.setEnabled(False)
        self.btn_refresh_gbif.setEnabled(False)
        self.btn_fetch_gbif.setText("가져오는 중...")
        progress_dialog = QProgressDialog(
            "GBIF 데이터를 가져오는 중입니다...",
            "취소",
            0,
            0,
            self,
        )
        progress_dialog.setWindowTitle("GBIF 분석")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setMinimumDuration(0)
        progress_dialog.show()
        QApplication.processEvents()

        def update_progress(fetched_count: int, total_records: int, requested_limit: int):
            target_count = min(total_records or requested_limit, requested_limit)
            progress_dialog.setMaximum(max(target_count, 1))
            progress_dialog.setValue(min(fetched_count, target_count))
            progress_dialog.setLabelText(
                "GBIF 데이터를 가져오는 중입니다...\n"
                f"가져옴: {fetched_count:,}건 / 목표: {target_count:,}건"
            )
            QApplication.processEvents()
            return not progress_dialog.wasCanceled()

        try:
            result = GbifService.fetch_occurrences_by_criteria(
                criteria,
                use_cache=use_cache,
                progress_callback=update_progress,
            )
            self.gbif_df = result.dataframe
            self.gbif_search_summary = {
                "matchedName": result.matched_name,
                "taxonKey": result.taxon_key,
                "rank": result.rank,
                "status": result.status,
                "totalRecords": result.total_records,
                "shownRecords": len(result.dataframe),
                "countryCode": self._format_selected_codes(criteria.country_codes),
                "continent": self._format_selected_codes(criteria.continent_codes),
                "datasetKey": criteria.dataset_key or "ALL",
                "geometry": criteria.geometry or "ALL",
                "basisOfRecord": "ALL",
            }
            self.gbif_raw_df = result.dataframe
            self.gbif_df = self.gbif_raw_df
            self.update_gbif_year_range(self.gbif_raw_df)
            self.build_gbif_basis_filter(self.gbif_raw_df)
            self.build_gbif_species_filter(self.gbif_raw_df)
            self.update_gbif_filtered_df(show_table=True)
            has_data = self.gbif_raw_df is not None and not self.gbif_raw_df.empty
            self.set_gbif_analysis_buttons_enabled(has_data)
            limit_note = ""
            if result.total_records > len(result.dataframe):
                limit_note = "\n빠른 분석은 최대 100,000건까지 표시합니다. 논문용 전체 자료는 DOI 다운로드 요청을 사용하세요."
            cache_text = "캐시" if result.from_cache else "GBIF API"
            self.gbif_summary_label.setText(
                f"일치 이름: {result.matched_name} / taxonKey: {result.taxon_key} / "
                f"datasetKey: {criteria.dataset_key or 'ALL'} / 국가: {self._format_selected_codes(criteria.country_codes)} / "
                f"대륙: {self._format_selected_codes(criteria.continent_codes)} / "
                f"전체 좌표 기록: {result.total_records:,}건 / 가져옴: {len(result.dataframe):,}건 / "
                f"정제 후: {len(self.gbif_filtered_df):,}건 / 출처: {cache_text}"
                f"{limit_note}"
            )
            if not has_data:
                QMessageBox.information(self, "GBIF 분석", "일치하는 좌표 기록이 없습니다.")
        except GbifFetchCancelled:
            self.reset_gbif_analysis_state(clear_raw=True)
            self.gbif_search_summary = None
            self.update_gbif_year_range(pd.DataFrame())
            self.set_gbif_analysis_buttons_enabled(False)
            self.gbif_summary_label.setText("GBIF 데이터 가져오기가 취소되었습니다.")
        except Exception as e:
            self.reset_gbif_analysis_state(clear_raw=True)
            self.gbif_search_summary = None
            self.update_gbif_year_range(pd.DataFrame())
            self.set_gbif_analysis_buttons_enabled(False)
            QMessageBox.critical(self, "GBIF 분석 오류", f"GBIF 데이터를 가져오는 중 오류가 발생했습니다.\n\n{e}")
        finally:
            progress_dialog.close()
            self.btn_fetch_gbif.setEnabled(True)
            self.btn_refresh_gbif.setEnabled(True)
            self.btn_fetch_gbif.setText("GBIF 가져오기")

    def reset_gbif_analysis_state(self, clear_raw: bool = False):
        if clear_raw:
            self.gbif_raw_df = None
            self.gbif_df = None
        self.gbif_filtered_df = None
        self.gbif_climate_df = None
        self.gbif_climate_location = None
        self.gbif_report_path = None
        self.gbif_table_preview_df = None
        self.gbif_preview_loaded_rows = 0
        self._loading_gbif_preview_rows = False
        self.gbif_result_table.clear()
        self.gbif_result_table.setRowCount(0)
        self.gbif_result_table.setColumnCount(0)
        if hasattr(self, "gbif_live_chart"):
            self.gbif_live_chart.clear_chart()
        if hasattr(self, "gbif_climate_label"):
            self.gbif_climate_label.setText("기후 데이터가 아직 연결되지 않았습니다.")
        self.gbif_filtered_count_label.setText("정제 조건에 해당하는 데이터: 0건")
        self.clear_gbif_basis_filter()
        self.clear_gbif_species_filter()
        self.set_gbif_analysis_buttons_enabled(False)

    def reset_gbif_cleaning_controls(self):
        self._building_gbif_basis_filter = True
        try:
            self.gbif_year_from_input.setRange(0, 9999)
            self.gbif_year_to_input.setRange(0, 9999)
            self.gbif_year_from_input.setValue(0)
            self.gbif_year_to_input.setValue(0)
            self.gbif_month_from_input.setValue(1)
            self.gbif_month_to_input.setValue(12)
            self._sync_gbif_quick_month_combo()
            self.gbif_filter_country_combo.clear_selection()
            self.gbif_filter_continent_combo.clear_selection()
            self.gbif_filter_species_combo.clear_selection()
            self.gbif_fill_missing_periods_check.setChecked(True)
            self.gbif_cumulative_check.setChecked(False)
            self.gbif_remove_missing_year_check.setChecked(True)
            self.gbif_remove_missing_month_check.setChecked(False)
            self.gbif_remove_missing_name_check.setChecked(True)
            self.gbif_dedupe_gbif_id_check.setChecked(True)
            self.gbif_exclude_dataset_input.clear()
            self.gbif_exclude_institution_input.clear()
        finally:
            self._building_gbif_basis_filter = False

    def apply_gbif_quick_month_filter(self, *_):
        if getattr(self, "_syncing_gbif_quick_month", False):
            return
        month = int(self.gbif_quick_month_combo.currentData() or 0)
        self._syncing_gbif_quick_month = True
        try:
            if month:
                self.gbif_month_from_input.setValue(month)
                self.gbif_month_to_input.setValue(month)
            else:
                self.gbif_month_from_input.setValue(1)
                self.gbif_month_to_input.setValue(12)
        finally:
            self._syncing_gbif_quick_month = False
        if self.gbif_raw_df is None:
            self.auto_apply_gbif_cleaning()
            return
        self.gbif_cleaning_timer.stop()
        self.update_gbif_filtered_df(show_table=True)

    def _sync_gbif_quick_month_combo(self):
        if not hasattr(self, "gbif_quick_month_combo"):
            return
        month_from = self.gbif_month_from_input.value()
        month_to = self.gbif_month_to_input.value()
        month = month_from if month_from == month_to else 0
        if month_from == 1 and month_to == 12:
            month = 0
        self._syncing_gbif_quick_month = True
        try:
            index = self.gbif_quick_month_combo.findData(month)
            if index >= 0:
                self.gbif_quick_month_combo.setCurrentIndex(index)
        finally:
            self._syncing_gbif_quick_month = False

    def set_gbif_analysis_buttons_enabled(self, enabled: bool):
        self.btn_draw_gbif_graph.setEnabled(enabled)
        self.btn_apply_gbif_cleaning.setEnabled(enabled)
        self.btn_open_gbif_report.setEnabled(enabled)
        self.btn_export_gbif_csv.setEnabled(enabled)
        self.btn_export_gbif_excel.setEnabled(enabled)
        self.btn_export_gbif_timeseries_csv.setEnabled(enabled)
        self.btn_fetch_gbif_climate.setEnabled(enabled)
        self.btn_export_gbif_geojson.setEnabled(enabled)

    def _connect_gbif_cleaning_signals(self):
        for widget in [
            self.gbif_year_from_input,
            self.gbif_year_to_input,
            self.gbif_month_from_input,
            self.gbif_month_to_input,
        ]:
            widget.valueChanged.connect(self.auto_apply_gbif_cleaning)

        self.gbif_filter_country_combo.selectionChanged.connect(self.auto_apply_gbif_cleaning)
        self.gbif_filter_continent_combo.selectionChanged.connect(self.auto_apply_gbif_cleaning)
        self.gbif_filter_species_combo.selectionChanged.connect(self.auto_apply_gbif_cleaning)

        for widget in [
            self.gbif_remove_missing_year_check,
            self.gbif_remove_missing_month_check,
            self.gbif_remove_missing_name_check,
            self.gbif_dedupe_gbif_id_check,
        ]:
            widget.stateChanged.connect(self.auto_apply_gbif_cleaning)

        self.gbif_exclude_dataset_input.textChanged.connect(self.auto_apply_gbif_cleaning)
        self.gbif_exclude_institution_input.textChanged.connect(self.auto_apply_gbif_cleaning)

    def auto_apply_gbif_cleaning(self, *_):
        if getattr(self, "_building_gbif_basis_filter", False):
            return
        if not getattr(self, "_syncing_gbif_quick_month", False):
            self._sync_gbif_quick_month_combo()
        if self.gbif_raw_df is None:
            return
        self.gbif_filtered_count_label.setText("정제 조건 변경 감지: 잠시 후 적용됩니다...")
        self.gbif_cleaning_timer.start()

    def apply_gbif_cleaning_silently(self):
        if self.gbif_raw_df is None:
            return
        self.update_gbif_filtered_df(show_table=True)

    def clear_gbif_basis_filter(self):
        self._building_gbif_basis_filter = True
        try:
            while self.gbif_basis_layout.count():
                item = self.gbif_basis_layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
            self.gbif_basis_checkboxes = {}
            empty_label = QLabel("데이터를 가져오면 자료유형 목록이 표시됩니다.")
            empty_label.setProperty("class", "sectionDescription")
            self.gbif_basis_layout.addWidget(empty_label)
        finally:
            self._building_gbif_basis_filter = False

    def build_gbif_basis_filter(self, df: pd.DataFrame):
        self._building_gbif_basis_filter = True
        try:
            while self.gbif_basis_layout.count():
                item = self.gbif_basis_layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
            self.gbif_basis_checkboxes = {}

            if "basisOfRecord" not in df.columns or df.empty:
                self.gbif_basis_layout.addWidget(QLabel("자료유형 값이 없습니다."))
                return

            counts = (
                df["basisOfRecord"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "UNKNOWN")
                .value_counts()
                .sort_index()
            )
            for basis, count in counts.items():
                checkbox = QCheckBox(f"{self._format_basis_label(basis)}    {int(count):,}")
                checkbox.setChecked(True)
                checkbox.stateChanged.connect(self.auto_apply_gbif_cleaning)
                self.gbif_basis_checkboxes[basis] = checkbox
                self.gbif_basis_layout.addWidget(checkbox)
        finally:
            self._building_gbif_basis_filter = False

    def clear_gbif_species_filter(self):
        if not hasattr(self, "gbif_filter_species_combo"):
            return
        self.gbif_filter_species_combo.clear_items()

    def build_gbif_species_filter(self, df: pd.DataFrame):
        if not hasattr(self, "gbif_filter_species_combo"):
            return
        self._building_gbif_basis_filter = True
        try:
            self.gbif_filter_species_combo.clear_items()
            species_values = self._gbif_species_values(df)
            if species_values.empty:
                return
            counts = species_values.value_counts().sort_index()
            for species_name, count in counts.items():
                label = f"{species_name[:80]}    {int(count):,}"
                self.gbif_filter_species_combo.add_check_item(label, species_name)
        finally:
            self._building_gbif_basis_filter = False

    @staticmethod
    def _format_basis_label(value: str) -> str:
        if value == "UNKNOWN":
            return "Unknown"
        return value.replace("_", " ").title()

    def selected_gbif_basis_values(self) -> set[str]:
        if not self.gbif_basis_checkboxes:
            return set()
        return {
            basis
            for basis, checkbox in self.gbif_basis_checkboxes.items()
            if checkbox.isChecked()
        }

    def selected_gbif_species_values(self) -> list[str]:
        if not hasattr(self, "gbif_filter_species_combo"):
            return []
        return self.gbif_filter_species_combo.selected_values()

    def current_gbif_time_series_options(self) -> tuple[bool, bool]:
        fill_missing = (
            hasattr(self, "gbif_fill_missing_periods_check")
            and self.gbif_fill_missing_periods_check.isChecked()
        )
        cumulative = hasattr(self, "gbif_cumulative_check") and self.gbif_cumulative_check.isChecked()
        return fill_missing, cumulative

    @staticmethod
    def _gbif_climate_charts(
        climate_df: pd.DataFrame | None,
        year_from: int | None,
        year_to: int | None,
        month_from: int,
        month_to: int,
    ) -> list[dict]:
        if climate_df is None or climate_df.empty:
            return []
        df = climate_df.copy()
        if year_from is not None:
            df = df.loc[df["year"].ge(year_from)]
        if year_to is not None:
            df = df.loc[df["year"].le(year_to)]
        df = df.loc[df["month"].ge(month_from) & df["month"].le(month_to)]
        if df.empty:
            return []
        labels = [f"{int(row.year)}-{int(row.month):02d}" for row in df.itertuples()]
        charts = []
        if "temperatureC" in df.columns:
            charts.append(
                {
                    "title": "월별 평균기온 (NASA POWER, C)",
                    "labels": labels,
                    "series": [
                        {
                            "name": "temperatureC",
                            "values": [round(float(value), 2) if pd.notna(value) else 0 for value in df["temperatureC"]],
                        }
                    ],
                }
            )
        if "precipitationTotalMm" in df.columns:
            charts.append(
                {
                    "title": "월별 총강수량 (NASA POWER, mm)",
                    "labels": labels,
                    "series": [
                        {
                            "name": "precipitationTotalMm",
                            "values": [
                                round(float(value), 2) if pd.notna(value) else 0 for value in df["precipitationTotalMm"]
                            ],
                        }
                    ],
                }
            )
        return charts

    def update_gbif_filtered_df(self, show_table: bool = True):
        self.gbif_filtered_df = self.get_current_gbif_analysis_df()
        if show_table:
            self.show_gbif_results(self.gbif_filtered_df)
        self.update_gbif_filtered_count_label()
        self.update_gbif_live_chart()
        self.gbif_report_path = None

    def update_gbif_live_chart(self, *_):
        if not hasattr(self, "gbif_live_chart"):
            return
        if self.gbif_filtered_df is None or self.gbif_filtered_df.empty:
            self.gbif_live_chart.clear_chart("현재 조건에 맞는 데이터가 없습니다.")
            return
        try:
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            fill_missing, cumulative = self.current_gbif_time_series_options()
            chart_list = self._gbif_temporal_charts(
                self.gbif_filtered_df,
                self.gbif_period_combo.currentData() or "summary",
                year_from,
                year_to,
                month_from,
                month_to,
                self.selected_gbif_species_values(),
                fill_missing,
                cumulative,
            )
            climate_charts = self._gbif_climate_charts(
                self.gbif_climate_df,
                year_from,
                year_to,
                month_from,
                month_to,
            )
            if climate_charts:
                chart_list = chart_list[:1] + climate_charts
            self.gbif_live_chart.update_chart(
                chart_list,
                self.gbif_chart_type_combo.currentData() or "bar",
            )
        except Exception as e:
            self.gbif_live_chart.clear_chart(str(e))

    def update_gbif_filtered_count_label(self):
        raw_count = 0 if self.gbif_raw_df is None else len(self.gbif_raw_df)
        filtered_count = 0 if self.gbif_filtered_df is None else len(self.gbif_filtered_df)
        preview_total = 0 if self.gbif_table_preview_df is None else len(self.gbif_table_preview_df)
        preview_count = min(self.gbif_preview_loaded_rows, preview_total)
        self.gbif_filtered_count_label.setText(
            f"정제 조건에 해당하는 데이터: {filtered_count:,}건 / 원본 {raw_count:,}건 / "
            f"테이블 미리보기: {preview_count:,}건"
        )

    def fetch_gbif_climate_data(self):
        if self.gbif_filtered_df is None:
            self.update_gbif_filtered_df(show_table=False)
        if self.gbif_filtered_df is None or self.gbif_filtered_df.empty:
            QMessageBox.warning(self, "GBIF 기후 데이터", "현재 정제/필터 조건에 맞는 GBIF 데이터가 없습니다.")
            return

        try:
            latitude, longitude = self._gbif_filtered_coordinate_center(self.gbif_filtered_df)
            start_year, end_year = self._gbif_climate_year_range(self.gbif_filtered_df)
        except ValueError as e:
            QMessageBox.warning(self, "GBIF 기후 데이터", str(e))
            return

        self.btn_fetch_gbif_climate.setEnabled(False)
        self.btn_fetch_gbif_climate.setText("기후 조회 중...")
        try:
            result = ClimateService.fetch_monthly_power(
                latitude=latitude,
                longitude=longitude,
                start_year=start_year,
                end_year=end_year,
                use_cache=self.gbif_use_cache_check.isChecked(),
            )
            self.gbif_climate_df = result.dataframe
            self.gbif_climate_location = {
                "latitude": result.latitude,
                "longitude": result.longitude,
                "source": result.source,
                "fromCache": result.from_cache,
            }
            source_text = "캐시" if result.from_cache else result.source
            self.gbif_climate_label.setText(
                f"기후 데이터: {result.source} / 중심 좌표 {result.latitude:.4f}, {result.longitude:.4f} / "
                f"{result.start_year}-{result.end_year} / {len(result.dataframe):,}개월 / 출처: {source_text}"
            )
            self.update_gbif_live_chart()
        except Exception as e:
            QMessageBox.critical(self, "GBIF 기후 데이터 오류", f"기후 데이터를 가져오는 중 오류가 발생했습니다.\n\n{e}")
        finally:
            self.btn_fetch_gbif_climate.setEnabled(self.gbif_raw_df is not None and not self.gbif_raw_df.empty)
            self.btn_fetch_gbif_climate.setText("기후 데이터")

    @staticmethod
    def _gbif_filtered_coordinate_center(df: pd.DataFrame) -> tuple[float, float]:
        if "decimalLatitude" not in df.columns or "decimalLongitude" not in df.columns:
            raise ValueError("기후 데이터를 가져오려면 decimalLatitude/decimalLongitude 컬럼이 필요합니다.")
        latitudes = pd.to_numeric(df["decimalLatitude"], errors="coerce")
        longitudes = pd.to_numeric(df["decimalLongitude"], errors="coerce")
        valid_mask = latitudes.notna() & longitudes.notna()
        if not valid_mask.any():
            raise ValueError("기후 데이터를 가져올 수 있는 유효한 좌표가 없습니다.")
        return float(latitudes.loc[valid_mask].mean()), float(longitudes.loc[valid_mask].mean())

    @staticmethod
    def _gbif_climate_year_range(df: pd.DataFrame) -> tuple[int, int]:
        years = GbifAnalysisTab._gbif_year_values(df)
        if years.empty:
            raise ValueError("기후 데이터를 가져오려면 year 값이 필요합니다.")
        return int(years.min()), int(years.max())

    def draw_gbif_graph_report(self):
        if self.gbif_raw_df is None or self.gbif_filtered_df is None or self.gbif_filtered_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "그래프를 그릴 데이터가 없습니다.")
            return

        try:
            self.gbif_report_path = self.write_gbif_analysis_report()
            QMessageBox.information(
                self,
                "그래프 생성",
                f"지도/그래프 리포트를 생성했습니다.\n\n{self.gbif_report_path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "GBIF 분석 오류", f"그래프 생성 중 오류가 발생했습니다.\n\n{e}")

    def request_gbif_download(self):
        try:
            criteria = self.current_gbif_search_criteria()
        except ValueError as e:
            QMessageBox.warning(self, "GBIF 다운로드 요청", str(e))
            return

        username = self.gbif_username_input.text().strip()
        email = self.gbif_email_input.text().strip()
        password = self.gbif_password_input.text()

        self.btn_request_gbif_download.setEnabled(False)
        self.btn_request_gbif_download.setText("요청 중...")
        try:
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            selected_country_codes = self.gbif_filter_country_combo.selected_values()
            selected_continent_codes = self.gbif_filter_continent_combo.selected_values()
            country_codes = self._refined_gbif_codes(criteria.country_codes, selected_country_codes)
            continent_codes = self._refined_gbif_codes(criteria.continent_codes, selected_continent_codes)
            if selected_country_codes and not country_codes:
                raise ValueError("Search country and filter country do not overlap.")
            if selected_continent_codes and not continent_codes:
                raise ValueError("Search continent and filter continent do not overlap.")
            result = GbifService.request_occurrence_download(
                scientific_name=criteria.scientific_name,
                taxon_key=criteria.taxon_key,
                dataset_key=criteria.dataset_key,
                country_codes=country_codes,
                continent_codes=continent_codes,
                geometry=criteria.geometry,
                username=username,
                password=password,
                email=email,
                basis_of_record=self._single_selected_gbif_basis_value(),
                year_from=year_from,
                year_to=year_to,
                month_from=month_from,
                month_to=month_to,
            )
            self.settings.setValue("gbif/username", username)
            self.settings.setValue("gbif/email", email)
            QMessageBox.information(
                self,
                "GBIF 다운로드 요청 완료",
                "GBIF occurrence download 요청을 보냈습니다.\n\n"
                f"Download key: {result.key}\n"
                f"상태 API: {result.status_url}\n"
                "GBIF가 파일 준비를 마치면 계정 이메일로 다운로드/DOI 안내를 보냅니다. "
                "논문에는 다운로드 페이지의 공식 DOI citation 문구를 사용하세요.",
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "GBIF 다운로드 요청 오류",
                f"GBIF 다운로드 요청 중 오류가 발생했습니다.\n\n{e}",
            )
        finally:
            self.btn_request_gbif_download.setEnabled(True)
            self.btn_request_gbif_download.setText("DOI 다운로드 요청")

    def apply_gbif_cleaning(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 정제", "먼저 GBIF 데이터를 가져오세요.")
            return

        try:
            self.gbif_cleaning_timer.stop()
            self.update_gbif_filtered_df(show_table=True)
            QMessageBox.information(
                self,
                "GBIF 정제",
                f"정제 옵션을 적용했습니다.\n\n"
                f"원본: {len(self.gbif_raw_df):,}건\n"
                f"정제 후: {len(self.gbif_filtered_df):,}건",
            )
        except Exception as e:
            QMessageBox.critical(self, "GBIF 정제 오류", f"정제 옵션 적용 중 오류가 발생했습니다.\n\n{e}")

    def get_current_gbif_analysis_df(self) -> pd.DataFrame:
        if self.gbif_raw_df is None:
            return pd.DataFrame()

        year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
        filtered_df = self._filter_gbif_dataframe(
            self.gbif_raw_df,
            year_from,
            year_to,
            month_from,
            month_to,
            self.gbif_filter_country_combo.selected_values(),
            self.gbif_filter_continent_combo.selected_values(),
        )
        return self._apply_gbif_cleaning_options(filtered_df)

    def _apply_gbif_cleaning_options(self, df: pd.DataFrame) -> pd.DataFrame:
        cleaned_df = df.copy()

        if self.gbif_dedupe_gbif_id_check.isChecked() and "gbifID" in cleaned_df.columns:
            cleaned_df = cleaned_df.drop_duplicates(subset=["gbifID"], keep="first")

        if self.gbif_remove_missing_year_check.isChecked() and "year" in cleaned_df.columns:
            years = pd.to_numeric(cleaned_df["year"], errors="coerce")
            cleaned_df = cleaned_df.loc[years.notna()]

        if self.gbif_remove_missing_month_check.isChecked():
            month_values = self._gbif_month_values(cleaned_df).reindex(cleaned_df.index)
            cleaned_df = cleaned_df.loc[month_values.notna()]

        if self.gbif_remove_missing_name_check.isChecked() and "scientificName" in cleaned_df.columns:
            names = cleaned_df["scientificName"].fillna("").astype(str).str.strip()
            cleaned_df = cleaned_df.loc[names != ""]

        selected_basis = self.selected_gbif_basis_values()
        if selected_basis and "basisOfRecord" in cleaned_df.columns:
            basis_values = (
                cleaned_df["basisOfRecord"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "UNKNOWN")
            )
            cleaned_df = cleaned_df.loc[basis_values.isin(selected_basis)]

        selected_species = self.selected_gbif_species_values()
        if selected_species:
            species_values = self._gbif_species_values(cleaned_df).reindex(cleaned_df.index)
            cleaned_df = cleaned_df.loc[species_values.isin(selected_species)]

        cleaned_df = self._exclude_gbif_keywords(
            cleaned_df,
            "datasetName",
            self.gbif_exclude_dataset_input.text(),
        )
        cleaned_df = self._exclude_gbif_keywords(
            cleaned_df,
            "institutionCode",
            self.gbif_exclude_institution_input.text(),
        )

        return cleaned_df.reset_index(drop=True)

    @staticmethod
    def _exclude_gbif_keywords(df: pd.DataFrame, column: str, raw_keywords: str) -> pd.DataFrame:
        keywords = [keyword.strip().lower() for keyword in raw_keywords.split(",") if keyword.strip()]
        if not keywords or column not in df.columns:
            return df

        values = df[column].fillna("").astype(str).str.lower()
        exclude_mask = pd.Series(False, index=df.index)
        for keyword in keywords:
            exclude_mask |= values.str.contains(keyword, regex=False)
        return df.loc[~exclude_mask]

    def _single_selected_gbif_basis_value(self) -> str:
        selected_basis = self.selected_gbif_basis_values()
        if len(selected_basis) == 1:
            value = next(iter(selected_basis))
            return "" if value == "UNKNOWN" else value
        return ""

    def show_gbif_results(self, df: pd.DataFrame):
        self.gbif_table_preview_df = df.reset_index(drop=True)
        self.gbif_preview_loaded_rows = 0
        self.gbif_result_table.setUpdatesEnabled(False)
        self.gbif_result_table.clear()
        self.gbif_result_table.setRowCount(0)
        self.gbif_result_table.setColumnCount(len(self.gbif_table_preview_df.columns))
        self.gbif_result_table.setHorizontalHeaderLabels([str(col) for col in self.gbif_table_preview_df.columns.tolist()])
        self.gbif_result_table.setUpdatesEnabled(True)
        self.append_gbif_preview_rows()

    def maybe_append_gbif_preview_rows(self, value: int):
        if self._loading_gbif_preview_rows or self.gbif_table_preview_df is None:
            return
        scrollbar = self.gbif_result_table.verticalScrollBar()
        if value >= scrollbar.maximum() - 5:
            self.append_gbif_preview_rows()

    def append_gbif_preview_rows(self):
        if self.gbif_table_preview_df is None or self.gbif_table_preview_df.empty:
            self.update_gbif_filtered_count_label()
            return
        if self.gbif_preview_loaded_rows >= len(self.gbif_table_preview_df):
            return

        self._loading_gbif_preview_rows = True
        start_row = self.gbif_preview_loaded_rows
        end_row = min(start_row + self.gbif_preview_chunk_size, len(self.gbif_table_preview_df))

        try:
            self.gbif_result_table.setUpdatesEnabled(False)
            self.gbif_result_table.setRowCount(end_row)
            for row_idx in range(start_row, end_row):
                for col_idx, _ in enumerate(self.gbif_table_preview_df.columns):
                    value = self.gbif_table_preview_df.iloc[row_idx, col_idx]
                    self.gbif_result_table.setItem(
                        row_idx,
                        col_idx,
                        QTableWidgetItem("" if pd.isna(value) else str(value)),
                    )
            self.gbif_preview_loaded_rows = end_row
            if start_row == 0:
                self.gbif_result_table.resizeColumnsToContents()
                self.gbif_result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        finally:
            self.gbif_result_table.setUpdatesEnabled(True)
            self._loading_gbif_preview_rows = False
            self.update_gbif_filtered_count_label()

    def update_gbif_year_range(self, df: pd.DataFrame):
        years = self._gbif_year_values(df)
        if years.empty:
            for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
                widget.setRange(0, 9999)
                widget.setValue(0)
            return

        min_year = int(years.min())
        max_year = int(years.max())
        for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
            widget.setRange(0, 9999)
        self.gbif_year_from_input.setValue(min_year)
        self.gbif_year_to_input.setValue(max_year)

        self.update_gbif_period_controls()

    def update_gbif_period_controls(self, *_):
        for widget in (self.gbif_year_from_input, self.gbif_year_to_input):
            widget.setEnabled(True)
        self.update_gbif_live_chart()

    def get_gbif_filter_ranges(self) -> tuple[int | None, int | None, int, int]:
        year_from = self.gbif_year_from_input.value() if self.gbif_year_from_input.isEnabled() else 0
        year_to = self.gbif_year_to_input.value() if self.gbif_year_to_input.isEnabled() else 0
        month_from = self.gbif_month_from_input.value()
        month_to = self.gbif_month_to_input.value()

        if year_from and year_to and year_from > year_to:
            raise ValueError("연도 범위는 시작 연도가 끝 연도보다 클 수 없습니다.")
        if month_from > month_to:
            raise ValueError("월 범위는 시작 월이 끝 월보다 클 수 없습니다.")

        return year_from or None, year_to or None, month_from, month_to

    def open_gbif_analysis_report(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "먼저 GBIF 데이터를 가져오세요.")
            return
        if self.gbif_report_path and Path(self.gbif_report_path).exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.gbif_report_path)))
            return

        try:
            self.gbif_report_path = self.write_gbif_analysis_report()
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.gbif_report_path)))
        except Exception as e:
            QMessageBox.critical(self, "GBIF 분석 오류", f"리포트를 생성하는 중 오류가 발생했습니다.\n\n{e}")

    def write_gbif_analysis_report(self) -> Path:
        if self.gbif_filtered_df is None:
            self.update_gbif_filtered_df(show_table=False)
        if self.gbif_filtered_df is None or self.gbif_filtered_df.empty:
            raise ValueError("현재 정제/필터 조건에 맞는 데이터가 없습니다.")

        report_path = OUTPUT_DIR / "gbif_analysis_report.html"
        chart_type = self.gbif_chart_type_combo.currentData() or "bar"
        period_type = self.gbif_period_combo.currentData() or "summary"
        year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
        fill_missing, cumulative = self.current_gbif_time_series_options()
        report_path.write_text(
            self._gbif_analysis_report_html(
                self.gbif_filtered_df,
                self.gbif_search_summary or {},
                chart_type,
                period_type,
                year_from,
                year_to,
                month_from,
                month_to,
                self.selected_gbif_species_values(),
                fill_missing,
                cumulative,
            ),
            encoding="utf-8",
        )
        return report_path

    def export_gbif_analysis_csv(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "GBIF 분석 데이터 저장",
            str(Path(self._get_last_save_dir()) / "gbif_occurrences.csv"),
            "CSV Files (*.csv)",
        )
        if not file_path:
            return

        try:
            analysis_df = self.gbif_filtered_df if self.gbif_filtered_df is not None else self.get_current_gbif_analysis_df()
            if analysis_df.empty:
                QMessageBox.warning(self, "GBIF 분석", "현재 정제/필터 조건에 맞는 데이터가 없습니다.")
                return
            analysis_df.to_csv(file_path, index=False, encoding="utf-8-sig")
            self._remember_save_dir(file_path)
            QMessageBox.information(self, "저장 완료", f"GBIF 분석 데이터를 저장했습니다.\n\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"GBIF 분석 데이터 저장 중 오류가 발생했습니다.\n\n{e}")

    def export_gbif_analysis_excel(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "GBIF 분석 엑셀 저장",
            str(Path(self._get_last_save_dir()) / "gbif_analysis.xlsx"),
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            chart_type = self.gbif_chart_type_combo.currentData() or "bar"
            period_type = self.gbif_period_combo.currentData() or "summary"
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            fill_missing, cumulative = self.current_gbif_time_series_options()
            filtered_df = self.gbif_filtered_df if self.gbif_filtered_df is not None else self.get_current_gbif_analysis_df()
            if filtered_df.empty:
                QMessageBox.warning(self, "GBIF 분석", "현재 정제/필터 조건에 맞는 데이터가 없습니다.")
                return
            chart_list = self._gbif_temporal_charts(
                filtered_df,
                period_type,
                year_from,
                year_to,
                month_from,
                month_to,
                self.selected_gbif_species_values(),
                fill_missing,
                cumulative,
            )
            self._write_gbif_analysis_workbook(
                file_path,
                filtered_df,
                chart_list,
                chart_type,
                {
                    **(self.gbif_search_summary or {}),
                    "yearFrom": year_from or "ALL",
                    "yearTo": year_to or "ALL",
                    "monthFrom": month_from,
                    "monthTo": month_to,
                },
            )
            self._remember_save_dir(file_path)
            QMessageBox.information(
                self,
                "저장 완료",
                f"GBIF 분석 데이터와 그래프를 엑셀로 저장했습니다.\n\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"GBIF 분석 엑셀 저장 중 오류가 발생했습니다.\n\n{e}")

    def export_gbif_timeseries_csv(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "GBIF Tableau 데이터 저장",
            str(Path(self._get_last_save_dir()) / "gbif_tableau_timeseries.xlsx"),
            "Excel Files (*.xlsx);;JSON Files (*.json);;CSV Files (*.csv)",
        )
        if not file_path:
            return

        try:
            filtered_df = self.gbif_filtered_df if self.gbif_filtered_df is not None else self.get_current_gbif_analysis_df()
            if filtered_df.empty:
                QMessageBox.warning(self, "GBIF 분석", "현재 정제/필터 조건에 맞는 데이터가 없습니다.")
                return
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            fill_missing, cumulative = self.current_gbif_time_series_options()
            tableau_df = self._gbif_tableau_timeseries_dataframe(
                filtered_df,
                year_from,
                year_to,
                month_from,
                month_to,
                self.selected_gbif_species_values(),
                fill_missing,
                self.gbif_climate_df,
            )
            if file_path.lower().endswith(".xlsx"):
                tableau_df.to_excel(file_path, index=False)
            elif file_path.lower().endswith(".json"):
                Path(file_path).write_text(
                    tableau_df.to_json(orient="records", force_ascii=False, indent=2),
                    encoding="utf-8",
                )
            else:
                if not file_path.lower().endswith(".csv"):
                    file_path += ".xlsx"
                    tableau_df.to_excel(file_path, index=False)
                    self._remember_save_dir(file_path)
                    QMessageBox.information(self, "저장 완료", f"Tableau용 GBIF 시계열 데이터를 저장했습니다.\n\n{file_path}")
                    return
                tableau_df.to_csv(file_path, index=False, encoding="utf-8-sig")
            self._remember_save_dir(file_path)
            QMessageBox.information(self, "저장 완료", f"Tableau용 GBIF 시계열 데이터를 저장했습니다.\n\n{file_path}")
            return
            chart_list = self._gbif_temporal_charts(
                filtered_df,
                self.gbif_period_combo.currentData() or "summary",
                year_from,
                year_to,
                month_from,
                month_to,
                self.selected_gbif_species_values(),
                fill_missing,
                cumulative,
            )
            chart_list = chart_list + self._gbif_climate_charts(
                self.gbif_climate_df,
                year_from,
                year_to,
                month_from,
                month_to,
            )
            rows = []
            for chart in chart_list:
                labels = chart.get("labels", [])
                for series in chart.get("series", []):
                    values = series.get("values", [])
                    for label, value in zip(labels, values):
                        series_name = str(series.get("name", ""))
                        rows.append(
                            {
                                "chart": chart.get("title", ""),
                                "series": series_name,
                                "period": label,
                                "count": "" if series_name in {"temperatureC", "precipitationMm"} else int(value or 0),
                                "temperatureC": value if series_name == "temperatureC" else "",
                                "precipitationMm": value if series_name == "precipitationMm" else "",
                                "value": value,
                            }
                        )
            pd.DataFrame(rows).to_csv(file_path, index=False, encoding="utf-8-sig")
            self._remember_save_dir(file_path)
            QMessageBox.information(self, "저장 완료", f"GBIF 시계열 데이터를 저장했습니다.\n\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"GBIF 시계열 CSV 저장 중 오류가 발생했습니다.\n\n{e}")

    def _gbif_tableau_timeseries_dataframe(
        self,
        df: pd.DataFrame,
        year_from: int | None,
        year_to: int | None,
        month_from: int,
        month_to: int,
        selected_species: list[str],
        fill_missing: bool,
        climate_df: pd.DataFrame | None,
    ) -> pd.DataFrame:
        working_df = df.copy()
        year_values = self._gbif_year_values(working_df).reindex(working_df.index)
        month_values = self._gbif_month_values(working_df).reindex(working_df.index)
        valid_mask = year_values.notna() & month_values.notna()
        working_df = working_df.loc[valid_mask].copy()
        working_df["year"] = year_values.loc[valid_mask].astype(int)
        working_df["month"] = month_values.loc[valid_mask].astype(int)
        if selected_species:
            working_df["seriesName"] = self._gbif_species_values(working_df).reindex(working_df.index).fillna("Unknown")
        else:
            working_df["seriesName"] = "ALL"

        monthly_counts = (
            working_df.groupby(["seriesName", "year", "month"], dropna=False)
            .size()
            .reset_index(name="occurrenceCount")
        )
        if fill_missing and not working_df.empty:
            start_year = year_from if year_from is not None else int(working_df["year"].min())
            end_year = year_to if year_to is not None else int(working_df["year"].max())
            series_names = sorted(monthly_counts["seriesName"].dropna().unique().tolist()) or ["ALL"]
            index_rows = [
                {"seriesName": series_name, "year": year, "month": month}
                for series_name in series_names
                for year in range(start_year, end_year + 1)
                for month in range(month_from, month_to + 1)
            ]
            monthly_counts = (
                pd.DataFrame(index_rows)
                .merge(monthly_counts, on=["seriesName", "year", "month"], how="left")
                .fillna({"occurrenceCount": 0})
            )

        monthly_counts["occurrenceCount"] = monthly_counts["occurrenceCount"].astype(int)
        monthly_counts["periodGrain"] = "month"
        monthly_counts["periodLabel"] = (
            monthly_counts["year"].astype(str) + "-" + monthly_counts["month"].astype(str).str.zfill(2)
        )
        monthly_counts["periodDate"] = pd.to_datetime(
            monthly_counts["periodLabel"] + "-01",
            errors="coerce",
        ).dt.strftime("%Y-%m-%d")
        monthly_counts = self._merge_climate_for_tableau(monthly_counts, climate_df)
        annual_counts = self._annualize_tableau_timeseries(monthly_counts)
        return pd.concat([monthly_counts, annual_counts], ignore_index=True).sort_values(
            ["periodGrain", "seriesName", "year", "month"]
        )

    def _merge_climate_for_tableau(self, monthly_df: pd.DataFrame, climate_df: pd.DataFrame | None) -> pd.DataFrame:
        result_df = monthly_df.copy()
        if climate_df is None or climate_df.empty:
            result_df["temperatureC"] = ""
            result_df["precipitationMmPerDay"] = ""
            result_df["precipitationTotalMm"] = ""
            result_df["climateSource"] = ""
            result_df["climateLatitude"] = ""
            result_df["climateLongitude"] = ""
            return result_df

        climate_columns = [
            column
            for column in ["year", "month", "temperatureC", "precipitationMmPerDay", "precipitationTotalMm"]
            if column in climate_df.columns
        ]
        result_df = result_df.merge(climate_df[climate_columns], on=["year", "month"], how="left")
        location = self.gbif_climate_location or {}
        result_df["climateSource"] = location.get("source", "NASA POWER")
        result_df["climateLatitude"] = location.get("latitude", "")
        result_df["climateLongitude"] = location.get("longitude", "")
        return result_df

    @staticmethod
    def _annualize_tableau_timeseries(monthly_df: pd.DataFrame) -> pd.DataFrame:
        annual_rows = []
        for (series_name, year), group in monthly_df.groupby(["seriesName", "year"], dropna=False):
            annual_rows.append(
                {
                    "seriesName": series_name,
                    "year": int(year),
                    "month": 0,
                    "occurrenceCount": int(group["occurrenceCount"].sum()),
                    "periodGrain": "year",
                    "periodLabel": str(int(year)),
                    "periodDate": f"{int(year)}-01-01",
                    "temperatureC": pd.to_numeric(group.get("temperatureC"), errors="coerce").mean(),
                    "precipitationMmPerDay": pd.to_numeric(group.get("precipitationMmPerDay"), errors="coerce").mean(),
                    "precipitationTotalMm": pd.to_numeric(group.get("precipitationTotalMm"), errors="coerce").sum(),
                    "climateSource": group.get("climateSource", pd.Series([""])).iloc[0],
                    "climateLatitude": group.get("climateLatitude", pd.Series([""])).iloc[0],
                    "climateLongitude": group.get("climateLongitude", pd.Series([""])).iloc[0],
                }
            )
        return pd.DataFrame(annual_rows)

    def export_gbif_analysis_geojson(self):
        if self.gbif_raw_df is None or self.gbif_raw_df.empty:
            QMessageBox.warning(self, "GBIF 분석", "저장할 GBIF 데이터가 없습니다.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "QGIS용 GeoJSON 저장",
            str(Path(self._get_last_save_dir()) / "gbif_occurrences.geojson"),
            "GeoJSON Files (*.geojson);;JSON Files (*.json)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith((".geojson", ".json")):
            file_path += ".geojson"

        try:
            year_from, year_to, month_from, month_to = self.get_gbif_filter_ranges()
            filtered_df = self.gbif_filtered_df if self.gbif_filtered_df is not None else self.get_current_gbif_analysis_df()
            geojson = self._gbif_dataframe_to_geojson(filtered_df)
            feature_count = len(geojson["features"])
            if feature_count == 0:
                QMessageBox.warning(
                    self,
                    "QGIS 저장",
                    "저장할 수 있는 유효한 좌표가 없습니다.",
                )
                return

            Path(file_path).write_text(
                json.dumps(geojson, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self._remember_save_dir(file_path)
            QMessageBox.information(
                self,
                "저장 완료",
                f"QGIS에서 열 수 있는 GeoJSON 파일을 저장했습니다.\n\n{file_path}\n\n"
                f"저장된 좌표: {feature_count:,}개",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"QGIS용 GeoJSON 저장 중 오류가 발생했습니다.\n\n{e}")

    @staticmethod
    def _gbif_dataframe_to_geojson(df: pd.DataFrame) -> dict:
        features = []
        for _, row in df.iterrows():
            lat = pd.to_numeric(row.get("decimalLatitude"), errors="coerce")
            lon = pd.to_numeric(row.get("decimalLongitude"), errors="coerce")
            if pd.isna(lat) or pd.isna(lon):
                continue
            if not (-90 <= float(lat) <= 90 and -180 <= float(lon) <= 180):
                continue

            properties = {}
            for column, value in row.items():
                if column in {"decimalLatitude", "decimalLongitude"}:
                    continue
                if pd.isna(value):
                    properties[str(column)] = None
                else:
                    properties[str(column)] = value.item() if hasattr(value, "item") else value

            features.append(
                {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [float(lon), float(lat)],
                    },
                    "properties": properties,
                }
            )

        return {
            "type": "FeatureCollection",
            "name": "gbif_occurrences",
            "features": features,
        }

    @staticmethod
    def _write_gbif_analysis_workbook(
        file_path: str,
        df: pd.DataFrame,
        chart_list: list[dict],
        chart_type: str,
        summary: dict,
    ):
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows

        workbook = Workbook()
        summary_ws = workbook.active
        summary_ws.title = "Summary"

        summary_ws.append(["항목", "값"])
        for key, value in summary.items():
            summary_ws.append([key, value])
        summary_ws.append(["filteredRecords", len(df)])

        data_ws = workbook.create_sheet("Occurrences")
        for row in dataframe_to_rows(df, index=False, header=True):
            data_ws.append(row)

        chart_data_ws = workbook.create_sheet("ChartData")
        charts_ws = workbook.create_sheet("Charts")
        chart_data_row = 1
        chart_anchor_row = 1

        for chart_index, chart_data in enumerate(chart_list, start=1):
            labels = chart_data.get("labels", [])
            series = chart_data.get("series", [])
            title = f"{chart_data.get('title', 'Chart')} 기록 수"

            start_row = chart_data_row
            chart_data_ws.cell(row=start_row, column=1, value=title)
            header_row = start_row + 1
            chart_data_ws.cell(row=header_row, column=1, value="구분")
            for series_index, item in enumerate(series, start=2):
                chart_data_ws.cell(row=header_row, column=series_index, value=item.get("name", f"Series {series_index - 1}"))

            for label_index, label in enumerate(labels, start=header_row + 1):
                chart_data_ws.cell(row=label_index, column=1, value=label)
                for series_index, item in enumerate(series, start=2):
                    values = item.get("values", [])
                    value = values[label_index - header_row - 1] if label_index - header_row - 1 < len(values) else 0
                    chart_data_ws.cell(row=label_index, column=series_index, value=value)

            end_row = header_row + len(labels)
            end_col = 1 + len(series)
            if labels and series:
                chart = LineChart() if chart_type == "line" else BarChart()
                chart.title = title
                chart.y_axis.title = "기록 수"
                chart.x_axis.title = "구분"
                chart.height = 8
                chart.width = 16
                data_ref = Reference(
                    chart_data_ws,
                    min_col=2,
                    max_col=end_col,
                    min_row=header_row,
                    max_row=end_row,
                )
                category_ref = Reference(
                    chart_data_ws,
                    min_col=1,
                    min_row=header_row + 1,
                    max_row=end_row,
                )
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(category_ref)
                charts_ws.add_chart(chart, f"A{chart_anchor_row}")
                chart_anchor_row += 18

            chart_data_row = end_row + 3

        for worksheet in workbook.worksheets:
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells[:200]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)

        workbook.save(file_path)

    @staticmethod
    def _gbif_analysis_report_html(
        df: pd.DataFrame,
        summary: dict,
        chart_type: str = "bar",
        period_type: str = "year",
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
        compare_species: list[str] | None = None,
        fill_missing_periods: bool = True,
        cumulative: bool = False,
    ) -> str:
        df = GbifAnalysisTab._filter_gbif_dataframe(
            df,
            year_from,
            year_to,
            month_from,
            month_to,
        )
        points = []
        for _, row in df.iterrows():
            lat = pd.to_numeric(row.get("decimalLatitude"), errors="coerce")
            lon = pd.to_numeric(row.get("decimalLongitude"), errors="coerce")
            if pd.isna(lat) or pd.isna(lon):
                continue
            points.append(
                {
                    "lat": float(lat),
                    "lon": float(lon),
                    "scientificName": str(row.get("scientificName", "") or ""),
                    "year": str(row.get("year", "") or ""),
                    "eventDate": str(row.get("eventDate", "") or ""),
                    "countryCode": str(row.get("countryCode", "") or ""),
                    "locality": str(row.get("locality", "") or ""),
                    "gbifID": str(row.get("gbifID", "") or ""),
                }
            )

        temporal_charts = GbifAnalysisTab._gbif_temporal_charts(
            df,
            period_type,
            year_from,
            year_to,
            month_from,
            month_to,
            compare_species,
            fill_missing_periods,
            cumulative,
        )

        points_json = json.dumps(points, ensure_ascii=False)
        temporal_json = json.dumps(temporal_charts, ensure_ascii=False)
        summary_json = json.dumps(summary, ensure_ascii=False)
        chart_type_json = json.dumps(chart_type if chart_type in {"bar", "line"} else "bar")
        return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GBIF Analysis Report</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
  <style>
    html, body {{
      margin: 0;
      min-height: 100%;
      background: #eef3f7;
      color: #1e293b;
      font-family: "Segoe UI", "Malgun Gothic", sans-serif;
    }}
    header {{
      background: #fbfdff;
      border-bottom: 1px solid #d8e2ec;
      padding: 18px 22px;
    }}
    h1 {{
      font-size: 22px;
      margin: 0 0 6px;
    }}
    .meta {{
      color: #52667a;
      font-size: 13px;
      line-height: 1.5;
    }}
    main {{
      display: grid;
      grid-template-columns: minmax(320px, 1.15fr) minmax(320px, 0.85fr);
      gap: 14px;
      padding: 14px;
    }}
    section {{
      background: #ffffff;
      border: 1px solid #d8e2ec;
      border-radius: 8px;
      overflow: hidden;
    }}
    .section-title {{
      border-bottom: 1px solid #e2e8f0;
      font-size: 15px;
      font-weight: 800;
      padding: 12px 14px;
    }}
    #map {{
      height: calc(100vh - 150px);
      min-height: 440px;
    }}
    .chart {{
      overflow: auto;
      padding: 12px 14px 16px;
    }}
    .chart-panel {{
      max-height: calc(100vh - 150px);
      overflow: auto;
    }}
    .chart-block {{
      border-bottom: 1px solid #e2e8f0;
    }}
    .chart-block:last-child {{
      border-bottom: none;
    }}
    .chart svg {{
      display: block;
      height: auto;
      max-width: 100%;
    }}
    .bar-row {{
      align-items: center;
      display: grid;
      grid-template-columns: 62px 1fr 52px;
      gap: 10px;
      margin: 7px 0;
    }}
    .bar {{
      background: #dbeafe;
      border-radius: 4px;
      height: 18px;
      overflow: hidden;
    }}
    .bar-fill {{
      background: #0f766e;
      height: 100%;
    }}
    .count {{
      color: #475569;
      font-variant-numeric: tabular-nums;
      text-align: right;
    }}
    .axis-label {{
      fill: #475569;
      font-size: 12px;
    }}
    .line-point {{
      fill: #ffffff;
      stroke-width: 2;
    }}
    .legend {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px 12px;
      margin-bottom: 10px;
    }}
    .legend-item {{
      align-items: center;
      color: #475569;
      display: inline-flex;
      font-size: 12px;
      gap: 5px;
    }}
    .legend-swatch {{
      border-radius: 999px;
      display: inline-block;
      height: 10px;
      width: 10px;
    }}
    @media (max-width: 900px) {{
      main {{
        grid-template-columns: 1fr;
      }}
      #map {{
        height: 520px;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>GBIF 분석 리포트</h1>
    <div class="meta" id="summary"></div>
  </header>
  <main>
    <section>
      <div class="section-title">좌표 지도</div>
      <div id="map"></div>
    </section>
    <section>
      <div class="section-title">기록 수 비교</div>
      <div class="chart-panel" id="chartPanel"></div>
    </section>
  </main>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <script>
    const points = {points_json};
    const chartList = {temporal_json};
    const summary = {summary_json};
    const chartType = {chart_type_json};
    const escapeHtml = (value) => String(value)
      .replaceAll("&", "&amp;")
      .replaceAll("<", "&lt;")
      .replaceAll(">", "&gt;")
      .replaceAll('"', "&quot;")
      .replaceAll("'", "&#039;");

    document.getElementById("summary").innerHTML = `
      일치 이름: <b>${{escapeHtml(summary.matchedName || "")}}</b> /
      taxonKey: ${{escapeHtml(summary.taxonKey || "")}} /
      국가: ${{escapeHtml(summary.countryCode || "ALL")}} /
      대륙: ${{escapeHtml(summary.continent || "ALL")}} /
      전체 좌표 기록: ${{Number(summary.totalRecords || 0).toLocaleString()}}건 /
      표시: ${{Number(summary.shownRecords || 0).toLocaleString()}}건
    `;

    const map = L.map("map");
    L.tileLayer("https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png", {{
      maxZoom: 19,
      attribution: "&copy; OpenStreetMap contributors"
    }}).addTo(map);

    const bounds = [];
    points.forEach((item) => {{
      const position = [item.lat, item.lon];
      bounds.push(position);
      const popup = `
        <b>${{escapeHtml(item.scientificName)}}</b><br>
        Year: ${{escapeHtml(item.year || "")}}<br>
        Event date: ${{escapeHtml(item.eventDate || "")}}<br>
        Country: ${{escapeHtml(item.countryCode || "")}}<br>
        Locality: ${{escapeHtml(item.locality || "")}}<br>
        GBIF ID: ${{escapeHtml(item.gbifID || "")}}
      `;
      L.circleMarker(position, {{
        radius: 5,
        color: "#0f766e",
        fillColor: "#14b8a6",
        fillOpacity: 0.58,
        weight: 1
      }}).bindPopup(popup).addTo(map);
    }});

    if (bounds.length === 1) {{
      map.setView(bounds[0], 8);
    }} else if (bounds.length > 1) {{
      map.fitBounds(bounds, {{ padding: [30, 30], maxZoom: 9 }});
    }} else {{
      map.setView([20, 0], 2);
    }}

    const chartPanel = document.getElementById("chartPanel");
    const colors = ["#0f766e", "#2563eb", "#b45309", "#7c3aed", "#be123c", "#0891b2", "#4d7c0f", "#9333ea"];
    const renderChart = (chartData) => {{
      const labels = chartData.labels || [];
      const series = chartData.series || [];
      const allCounts = series.flatMap((line) => line.values || []);
      const maxCount = Math.max(1, ...allCounts);
      const legend = () => series.length > 1
      ? `<div class="legend">${{series.map((line, index) => `
          <span class="legend-item">
            <span class="legend-swatch" style="background: ${{colors[index % colors.length]}}"></span>
            ${{escapeHtml(line.name)}}
          </span>
        `).join("")}}</div>`
      : "";
      const renderBarChart = () => legend() + series.flatMap((line, seriesIndex) =>
      labels.map((label, index) => {{
        const value = Number((line.values || [])[index] || 0);
        const rowLabel = series.length > 1 ? `${{line.name}} ${{label}}` : label;
        return `
          <div class="bar-row">
            <div>${{escapeHtml(rowLabel)}}</div>
            <div class="bar"><div class="bar-fill" style="width: ${{Math.max(2, value / maxCount * 100)}}%; background: ${{colors[seriesIndex % colors.length]}}"></div></div>
            <div class="count">${{value.toLocaleString()}}</div>
          </div>
        `;
      }})
    ).join("");

      const renderLineChart = () => {{
      const width = 680;
      const height = 320;
      const padding = {{ top: 24, right: 28, bottom: 42, left: 52 }};
      const plotWidth = width - padding.left - padding.right;
      const plotHeight = height - padding.top - padding.bottom;
      const xStep = labels.length > 1 ? plotWidth / (labels.length - 1) : 0;
      const pointFor = (value, index) => {{
        const x = padding.left + index * xStep;
        const y = padding.top + plotHeight - (value / maxCount * plotHeight);
        return {{ x, y }};
      }};
      const lines = series.map((line, seriesIndex) => {{
        const color = colors[seriesIndex % colors.length];
        const pointsAttr = (line.values || []).map((value, index) => {{
          const point = pointFor(Number(value || 0), index);
          return `${{point.x}},${{point.y}}`;
        }}).join(" ");
        const markers = (line.values || []).map((value, index) => {{
          const numericValue = Number(value || 0);
          const point = pointFor(numericValue, index);
          const labelY = point.y - 10 < 12 ? point.y + 22 : point.y - 10;
          const valueLabel = series.length === 1 || numericValue > 0
            ? `<text class="axis-label" x="${{point.x}}" y="${{labelY}}" text-anchor="middle">${{numericValue.toLocaleString()}}</text>`
            : "";
          return `
            <circle class="line-point" cx="${{point.x}}" cy="${{point.y}}" r="4" stroke="${{color}}"></circle>
            ${{valueLabel}}
          `;
        }}).join("");
        return `
          <polyline points="${{pointsAttr}}" fill="none" stroke="${{color}}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"></polyline>
          ${{markers}}
        `;
      }}).join("");
      const axisLabels = labels.map((label, index) => {{
        const point = pointFor(0, index);
        const visible = labels.length <= 18 || index % Math.ceil(labels.length / 18) === 0;
        return visible
          ? `<text class="axis-label" x="${{point.x}}" y="${{height - 16}}" text-anchor="middle">${{escapeHtml(label)}}</text>`
          : "";
      }}).join("");
      return legend() + `
        <svg viewBox="0 0 ${{width}} ${{height}}" role="img" aria-label="연도별 기록 수 선 그래프">
          <line x1="${{padding.left}}" y1="${{padding.top}}" x2="${{padding.left}}" y2="${{height - padding.bottom}}" stroke="#cbd5e1"></line>
          <line x1="${{padding.left}}" y1="${{height - padding.bottom}}" x2="${{width - padding.right}}" y2="${{height - padding.bottom}}" stroke="#cbd5e1"></line>
          <text class="axis-label" x="8" y="${{padding.top + 6}}">기록 수</text>
          ${{lines}}
          ${{axisLabels}}
        </svg>
      `;
    }};

      const body = labels.length && series.length
      ? (chartType === "line" ? renderLineChart() : renderBarChart())
      : `${{chartData.title}} 정보가 있는 기록이 없습니다.`;
      return `
        <div class="chart-block">
          <div class="section-title">${{escapeHtml(chartData.title)}} 기록 수</div>
          <div class="chart">${{body}}</div>
        </div>
      `;
    }};

    chartPanel.innerHTML = chartList.length
      ? chartList.map(renderChart).join("")
      : "표시할 그래프 데이터가 없습니다.";
  </script>
</body>
</html>
"""

    @staticmethod
    def _gbif_temporal_charts(
        df: pd.DataFrame,
        period_type: str,
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
        compare_species: list[str] | None = None,
        fill_missing_periods: bool = True,
        cumulative: bool = False,
    ) -> list[dict]:
        if compare_species and len(compare_species) > 1 and period_type in {"year", "month", "season"}:
            return [
                GbifAnalysisTab._apply_gbif_chart_series_options(
                    GbifAnalysisTab._gbif_species_period_chart_data(
                        df,
                        period_type,
                        compare_species,
                        year_from,
                        year_to,
                        month_from,
                        month_to,
                        fill_missing_periods,
                    ),
                    cumulative,
                )
            ]

        def chart_for(chart_period_type: str) -> dict:
            return GbifAnalysisTab._apply_gbif_chart_series_options(
                GbifAnalysisTab._gbif_temporal_chart_data(
                    df,
                    chart_period_type,
                    year_from,
                    year_to,
                    month_from,
                    month_to,
                    fill_missing_periods,
                ),
                cumulative,
            )

        if period_type == "summary":
            return [
                chart_for("year"),
                chart_for("month"),
                chart_for("season"),
            ]

        return [chart_for(period_type)]

    @staticmethod
    def _gbif_species_period_chart_data(
        df: pd.DataFrame,
        period_type: str,
        species_names: list[str],
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
        fill_missing_periods: bool = True,
    ) -> dict:
        species_values = GbifAnalysisTab._gbif_species_values(df).reindex(df.index)
        selected_species = [name for name in species_names if name in set(species_values.dropna())]
        if not selected_species:
            selected_species = species_names

        if period_type == "month":
            labels = [f"{month}월" for month in range(month_from, month_to + 1)]
            keys = list(range(month_from, month_to + 1))

            def values_for_species(species_df: pd.DataFrame) -> list[int]:
                month_values = GbifAnalysisTab._gbif_month_values(species_df)
                return [int((month_values == key).sum()) for key in keys]

            title = "월별 종 비교"
        elif period_type == "season":
            season_defs = [
                ("봄", {3, 4, 5}),
                ("여름", {6, 7, 8}),
                ("가을", {9, 10, 11}),
                ("겨울", {12, 1, 2}),
            ]
            season_defs = [
                (label, {month for month in months if month_from <= month <= month_to})
                for label, months in season_defs
            ]
            season_defs = [(label, months) for label, months in season_defs if months]
            labels = [label for label, _ in season_defs]

            def values_for_species(species_df: pd.DataFrame) -> list[int]:
                month_values = GbifAnalysisTab._gbif_month_values(species_df)
                return [int(month_values[month_values.isin(months)].count()) for _, months in season_defs]

            title = "계절별 종 비교"
        else:
            year_values = GbifAnalysisTab._gbif_year_values(df)
            if fill_missing_periods and not year_values.empty:
                start_year = year_from if year_from is not None else int(year_values.min())
                end_year = year_to if year_to is not None else int(year_values.max())
                years = list(range(start_year, end_year + 1))
            else:
                years = sorted(year_values.unique())
            if year_from is not None:
                years = [year for year in years if year >= year_from]
            if year_to is not None:
                years = [year for year in years if year <= year_to]
            labels = [str(year) for year in years]

            def values_for_species(species_df: pd.DataFrame) -> list[int]:
                species_years = GbifAnalysisTab._gbif_year_values(species_df)
                return [int((species_years == year).sum()) for year in years]

            title = "연도별 종 비교"

        series = []
        for species_name in selected_species:
            species_df = df.loc[species_values.eq(species_name).fillna(False)]
            series.append(
                {
                    "name": species_name[:70],
                    "values": values_for_species(species_df),
                }
            )
        return {"title": title, "labels": labels, "series": series}

    @staticmethod
    def _apply_gbif_chart_series_options(chart: dict, cumulative: bool = False) -> dict:
        if not cumulative:
            return chart
        updated_series = []
        for series in chart.get("series", []):
            running_total = 0
            values = []
            for value in series.get("values", []):
                running_total += int(value or 0)
                values.append(running_total)
            updated_series.append({**series, "values": values})
        title = str(chart.get("title", ""))
        return {**chart, "title": f"{title} 누적", "series": updated_series}

    @staticmethod
    def _gbif_temporal_chart_data(
        df: pd.DataFrame,
        period_type: str,
        year_from: int | None = None,
        year_to: int | None = None,
        month_from: int = 1,
        month_to: int = 12,
        fill_missing_periods: bool = True,
    ) -> dict:
        if period_type == "species":
            return GbifAnalysisTab._gbif_categorical_chart_data(
                df,
                "scientificName",
                "종별",
                fallback_column="acceptedScientificName",
            )

        if period_type == "dataset":
            return GbifAnalysisTab._gbif_categorical_chart_data(
                df,
                "datasetName",
                "데이터셋별",
                fallback_column="datasetKey",
            )

        if period_type == "basis":
            return GbifAnalysisTab._gbif_categorical_chart_data(
                df,
                "basisOfRecord",
                "자료유형별",
            )

        if period_type == "month":
            month_names = [
                "1월",
                "2월",
                "3월",
                "4월",
                "5월",
                "6월",
                "7월",
                "8월",
                "9월",
                "10월",
                "11월",
                "12월",
            ]
            month_keys = list(range(month_from, month_to + 1))
            series = GbifAnalysisTab._gbif_period_series_by_year(
                df,
                year_from,
                year_to,
                month_keys,
                lambda month_values, key: int((month_values == key).sum()),
                fill_missing_periods,
            )
            return {
                "title": "월별",
                "labels": [month_names[month - 1] for month in month_keys],
                "series": series,
            }

        if period_type == "season":
            season_defs = [
                ("봄", {3, 4, 5}),
                ("여름", {6, 7, 8}),
                ("가을", {9, 10, 11}),
                ("겨울", {12, 1, 2}),
            ]
            season_defs = [
                (label, {month for month in months if month_from <= month <= month_to})
                for label, months in season_defs
            ]
            season_defs = [(label, months) for label, months in season_defs if months]
            series = GbifAnalysisTab._gbif_period_series_by_year(
                df,
                year_from,
                year_to,
                [months for _, months in season_defs],
                lambda month_values, months: int(month_values[month_values.isin(months)].count()),
                fill_missing_periods,
            )
            return {
                "title": "계절별",
                "labels": [label for label, _ in season_defs],
                "series": series,
            }

        years = pd.to_numeric(df.get("year"), errors="coerce").dropna().astype(int)
        year_counts = years.value_counts().sort_index()
        if fill_missing_periods and not years.empty:
            start_year = year_from if year_from is not None else int(years.min())
            end_year = year_to if year_to is not None else int(years.max())
            full_years = list(range(start_year, end_year + 1))
            year_counts = year_counts.reindex(full_years, fill_value=0)
        return {
            "title": "연도별",
            "labels": [str(year) for year in year_counts.index],
            "series": [
                {
                    "name": "전체",
                    "values": [int(count) for count in year_counts.values],
                }
            ],
        }

    @staticmethod
    def _gbif_categorical_chart_data(
        df: pd.DataFrame,
        column: str,
        title: str,
        fallback_column: str | None = None,
        limit: int = 20,
    ) -> dict:
        if df.empty:
            counts = pd.Series(dtype="int64")
        elif column in df.columns:
            values = df[column].fillna("").astype(str).str.strip()
            if fallback_column and fallback_column in df.columns:
                fallback_values = df[fallback_column].fillna("").astype(str).str.strip()
                values = values.mask(values == "", fallback_values)
            counts = values.replace("", "Unknown").value_counts().head(limit)
        elif fallback_column and fallback_column in df.columns:
            counts = (
                df[fallback_column]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "Unknown")
                .value_counts()
                .head(limit)
            )
        else:
            counts = pd.Series(dtype="int64")

        labels = [str(label)[:70] for label in counts.index]
        return {
            "title": title,
            "labels": labels,
            "series": [
                {
                    "name": "전체",
                    "values": [int(count) for count in counts.values],
                }
            ],
        }

    @staticmethod
    def _gbif_period_series_by_year(
        df: pd.DataFrame,
        year_from: int | None,
        year_to: int | None,
        keys: list,
        count_for_key,
        fill_missing_periods: bool = True,
    ) -> list[dict]:
        year_values = GbifAnalysisTab._gbif_year_values(df)
        if fill_missing_periods and not year_values.empty:
            start_year = year_from if year_from is not None else int(year_values.min())
            end_year = year_to if year_to is not None else int(year_values.max())
            candidate_years = list(range(start_year, end_year + 1))
        else:
            candidate_years = sorted(year_values.unique())
        if year_from is not None:
            candidate_years = [year for year in candidate_years if year >= year_from]
        if year_to is not None:
            candidate_years = [year for year in candidate_years if year <= year_to]

        if not candidate_years:
            month_values = GbifAnalysisTab._gbif_month_values(df)
            return [
                {
                    "name": "전체",
                    "values": [count_for_key(month_values, key) for key in keys],
                }
            ]

        series = []
        for year in candidate_years:
            year_df = df.loc[year_values.reindex(df.index).eq(year)]
            month_values = GbifAnalysisTab._gbif_month_values(year_df)
            series.append(
                {
                    "name": str(year),
                    "values": [count_for_key(month_values, key) for key in keys],
                }
            )
        return series

    @staticmethod
    def _filter_gbif_dataframe(
        df: pd.DataFrame,
        year_from: int | None,
        year_to: int | None,
        month_from: int,
        month_to: int,
        country_codes: list[str] | None = None,
        continent_codes: list[str] | None = None,
    ) -> pd.DataFrame:
        filtered_df = df.copy()
        if filtered_df.empty:
            return filtered_df

        if year_from is not None or year_to is not None:
            year_values = GbifAnalysisTab._gbif_year_values(filtered_df).reindex(filtered_df.index)
            mask = pd.Series(True, index=filtered_df.index)
            if year_from is not None:
                mask &= year_values.ge(year_from).fillna(False)
            if year_to is not None:
                mask &= year_values.le(year_to).fillna(False)
            filtered_df = filtered_df.loc[mask]

        if month_from > 1 or month_to < 12:
            month_values = GbifAnalysisTab._gbif_month_values(filtered_df).reindex(filtered_df.index)
            mask = month_values.ge(month_from).fillna(False) & month_values.le(month_to).fillna(False)
            filtered_df = filtered_df.loc[mask]

        if country_codes and "countryCode" in filtered_df.columns:
            country_values = filtered_df["countryCode"].fillna("").astype(str).str.upper()
            allowed_countries = {code.upper() for code in country_codes}
            filtered_df = filtered_df.loc[country_values.isin(allowed_countries)]

        if continent_codes and "continent" in filtered_df.columns:
            continent_values = filtered_df["continent"].fillna("").astype(str).str.upper()
            allowed_continents = {code.upper() for code in continent_codes}
            filtered_df = filtered_df.loc[continent_values.isin(allowed_continents)]

        return filtered_df

    @staticmethod
    def _gbif_species_values(df: pd.DataFrame) -> pd.Series:
        if df.empty:
            return pd.Series(dtype="object")
        if "scientificName" in df.columns:
            values = df["scientificName"].fillna("").astype(str).str.strip()
        else:
            values = pd.Series("", index=df.index, dtype="object")
        if "acceptedScientificName" in df.columns:
            accepted_values = df["acceptedScientificName"].fillna("").astype(str).str.strip()
            values = values.mask(values == "", accepted_values)
        return values.replace("", pd.NA).dropna()

    @staticmethod
    def _gbif_year_values(df: pd.DataFrame) -> pd.Series:
        if "year" not in df.columns:
            return pd.Series(dtype="int64")

        return pd.to_numeric(df["year"], errors="coerce").dropna().astype(int)

    @staticmethod
    def _gbif_month_values(df: pd.DataFrame) -> pd.Series:
        if "month" in df.columns:
            month_values = pd.to_numeric(df["month"], errors="coerce")
        else:
            month_values = pd.Series(float("nan"), index=df.index, dtype="float64")

        if "eventDate" in df.columns:
            parsed_months = pd.to_datetime(
                df["eventDate"],
                errors="coerce",
                utc=True,
            ).dt.month
            month_values = month_values.fillna(parsed_months)

        return month_values.dropna().astype(int).loc[lambda values: values.between(1, 12)]

